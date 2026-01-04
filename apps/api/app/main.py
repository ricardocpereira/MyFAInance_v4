import csv
import hashlib
import json
import logging
import math
import os
import re
import secrets
import smtplib
import sqlite3
import threading
import unicodedata
import xlrd
import pdfplumber
import urllib.parse
import urllib.request
from io import BytesIO
from datetime import datetime, timedelta, date
from email.message import EmailMessage
from contextlib import contextmanager

from fastapi import FastAPI, Header, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook


app = FastAPI(title="MyFAInance v2 API")
logger = logging.getLogger("myfainance")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
CODE_TTL_MINUTES = 10
SESSION_TTL_HOURS = 24
DEFAULT_CATEGORIES = ["Cash", "Emergency Funds", "Retirement Plans", "Stocks"]
DEFAULT_INVESTMENT_TAGS = [
    "ETF",
    "Value Stocks",
    "Dividends",
    "REITs",
    "Emerging Markets",
    "Children Future",
]
KNOWN_ETF_TICKERS = {
    "XAIX.DE",
    "VUAA.DE",
    "VWCE.DE",
    "IS3N.DE",
    "NQSE.DE",
    "DFEN.DE",
    "SXR8.DE",
    "SXRV.DE",
    "XESC.DE",
    "XDWT.DE",
    "EUNL.DE",
    "QDVI.DE",
    "RBOT.UK",
    "DGTL.UK",
    "EGLN.UK",
    "EMEC.DE",
    "ESIH.DE",
    "ESIT.DE",
    "IMAE.NL",
    "IESE.NL",
    "IDVY.NL",
    "GOAI.DE",
    "HDLV.DE",
    "LSMC.DE",
    "WTEM.DE",
    "VVSM.DE",
    "VGWD.DE",
    "FLXI.DE",
    "CBRS.DE",
    "CEMS.DE",
    "SEC0.DE",
    "SDGPEX.DE",
    "ASWC.DE",
}
KNOWN_REIT_TICKERS = {
    "ARE",
    "ARE.US",
    "VICI.US",
    "CUBE.US",
}
MAX_TAG_LENGTH = 40
ALLOWED_CURRENCIES = {"EUR", "USD", "GBP"}
PRICE_API_PROVIDER = os.getenv("PRICE_API_PROVIDER", "twelvedata").lower()
PRICE_API_KEY = os.getenv("PRICE_API_KEY", "")
PRICE_CACHE_TTL_MINUTES = int(os.getenv("PRICE_CACHE_TTL_MINUTES", "60"))
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
DB_PATH = os.getenv(
    "DB_PATH", os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "app.db"))
)
# Ensure DB_PATH is absolute
if not os.path.isabs(DB_PATH):
    # If relative, make it relative to the app directory
    DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", DB_PATH))

BANKING_CATEGORY_TREE = {
    "Sem categoria": ["Sem subcategoria"],
    "Alimentacao": [
        "Sem subcategoria",
        "Cafetaria",
        "Mantimentos",
        "Restauracao",
        "Takeaway",
        "Talho",
    ],
    "Cuidados pessoais": ["Sem subcategoria"],
    "Despesas bancarias": [
        "Sem subcategoria",
        "Cheques",
        "Comissao de cartao",
        "Manutencao do deposito",
    ],
    "Habitacao": [
        "Sem subcategoria",
        "Condominio",
        "Electricidade",
        "Esgotos",
        "Gas",
        "TV & Internet",
        "Agua",
    ],
    "Habitacao (recheio)": [
        "Sem subcategoria",
        "Decoracao",
        "Eletrodomesticos",
        "Mobilia",
    ],
    "Impostos": ["Sem subcategoria", "IRS", "IUC", "Imposto de Selo"],
    "Juros": ["Sem subcategoria"],
    "Automovel": ["Sem subcategoria"],
    "Lazer": ["Sem subcategoria", "Cinema", "Desporto", "Discoteca", "Museus"],
    "Saude": ["Sem subcategoria", "Consultas", "Fisioterapia", "Medicamentos", "Urgencias"],
    "Telemovel": ["Sem subcategoria"],
    "Transportes": [
        "Sem subcategoria",
        "Comboio",
        "Combustivel",
        "Estacionamento",
        "Inspecao",
        "Metro",
        "Portagem",
    ],
    "Seguros": ["Sem subcategoria"],
    "Vencimentos": ["Sem subcategoria"],
    "Vestuario": ["Sem subcategoria"],
}
BANKING_DEFAULT_CATEGORY = "Sem categoria"
BANKING_DEFAULT_SUBCATEGORY = "Sem subcategoria"
BANKING_CATEGORY_KEYWORDS = [
    ("Alimentacao", "Restauracao", ["restaurante", "restauracao", "pizza", "sushi", "burger", "hamburguer"]),
    ("Alimentacao", "Cafetaria", ["cafe", "cafeteria", "pastelaria"]),
    ("Alimentacao", "Takeaway", ["takeaway", "ubereats", "uber eats", "glovo", "deliveroo"]),
    ("Alimentacao", "Mantimentos", ["supermercado", "continente", "pingo doce", "lidl", "aldi", "auchan", "minipreco"]),
    ("Alimentacao", "Talho", ["talho", "carnes"]),
    ("Transportes", "Combustivel", ["combustivel", "gasolina", "gasoleo", "galp", "repsol", "bp"]),
    ("Transportes", "Metro", ["metro"]),
    ("Transportes", "Comboio", ["comboio", "cp"]),
    ("Transportes", "Portagem", ["portagem", "via verde"]),
    ("Transportes", "Estacionamento", ["estacionamento", "parking", "parque"]),
    ("Telemovel", BANKING_DEFAULT_SUBCATEGORY, ["vodafone", "nos", "meo", "tmn"]),
    ("Saude", "Medicamentos", ["farmacia"]),
    ("Saude", "Consultas", ["consulta", "clinica", "hospital"]),
    ("Saude", "Fisioterapia", ["fisioterapia"]),
    ("Lazer", "Cinema", ["cinema"]),
    ("Lazer", "Desporto", ["ginasio", "gym", "fitness"]),
    ("Lazer", "Discoteca", ["discoteca", "bar"]),
    ("Lazer", "Museus", ["museu"]),
    ("Habitacao", "Condominio", ["condominio"]),
    ("Habitacao", "Electricidade", ["electricidade", "edp", "electric"]),
    ("Habitacao", "Gas", ["gas"]),
    ("Habitacao", "Agua", ["agua"]),
    ("Habitacao", "TV & Internet", ["internet", "tv"]),
    ("Despesas bancarias", "Comissao de cartao", ["comissao", "cartao", "card fee"]),
    ("Despesas bancarias", "Manutencao do deposito", ["manutencao", "quota"]),
    ("Impostos", "IRS", ["irs"]),
    ("Impostos", "IUC", ["iuc"]),
    ("Impostos", "Imposto de Selo", ["imposto de selo", "selo"]),
    ("Juros", BANKING_DEFAULT_SUBCATEGORY, ["juros", "interest"]),
    ("Seguros", BANKING_DEFAULT_SUBCATEGORY, ["seguro", "insurance"]),
    ("Vencimentos", BANKING_DEFAULT_SUBCATEGORY, ["salario", "vencimento", "salary"]),
    ("Vestuario", BANKING_DEFAULT_SUBCATEGORY, ["zara", "hm", "h&m", "bershka", "pull&bear", "roupa"]),
    ("Cuidados pessoais", BANKING_DEFAULT_SUBCATEGORY, ["cabeleireiro", "barbeiro", "cosmet", "spa", "estetica"]),
    ("Automovel", BANKING_DEFAULT_SUBCATEGORY, ["oficina", "mecanico", "auto"]),
    ("Habitacao (recheio)", "Eletrodomesticos", ["eletrodomestico", "eletro"]),
    ("Habitacao (recheio)", "Mobilia", ["mobilia", "sofa", "cama"]),
    ("Habitacao (recheio)", "Decoracao", ["decoracao", "ikea"]),
]


class RegisterRequest(BaseModel):
    email: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


class VerifyRequest(BaseModel):
    email: str
    code: str


class ResendRequest(BaseModel):
    email: str


class PortfolioCreateRequest(BaseModel):
    name: str
    currency: str
    custom_categories: list[str] | None = None


class SantanderItem(BaseModel):
    section: str
    account: str
    description: str | None = None
    balance: float
    category: str
    ignore: bool = False
    invested: float | None = None
    gains: float | None = None


class SantanderCommitRequest(BaseModel):
    filename: str
    items: list[SantanderItem]


class TradeRepublicManualRequest(BaseModel):
    available_cash: str | float
    interests_received: str | float
    currency: str | None = None
    category: str | None = None


class TradeRepublicImportItem(BaseModel):
    filename: str
    file_hash: str
    snapshot_date: str | None = None
    available_cash: float
    interests_received: float
    invested: float | None = None
    value: float | None = None
    gains: float | None = None
    category: str


class TradeRepublicImportRequest(BaseModel):
    items: list[TradeRepublicImportItem]


class SaveNGrowItem(BaseModel):
    name: str
    invested: float | None = None
    current_value: float
    profit_value: float | None = None
    profit_percent: float | None = None
    category: str


class SaveNGrowCommitRequest(BaseModel):
    filename: str
    file_hash: str
    snapshot_date: str | None = None
    items: list[SaveNGrowItem]


class AforroNetItem(BaseModel):
    name: str
    invested: float
    current_value: float
    category: str


class AforroNetCommitRequest(BaseModel):
    filename: str
    file_hash: str
    snapshot_date: str | None = None
    items: list[AforroNetItem]


class AforroNetBatchCommitRequest(BaseModel):
    imports: list[AforroNetCommitRequest]


class BancoInvestItem(BaseModel):
    holder: str
    invested: float | None = None
    current_value: float
    gains: float | None = None
    category: str


class BancoInvestCommitRequest(BaseModel):
    filename: str
    file_hash: str
    snapshot_date: str | None = None
    items: list[BancoInvestItem]


class XtbImportItem(BaseModel):
    filename: str
    file_hash: str
    account_type: str
    category: str
    current_value: float
    cash_value: float
    invested: float
    profit_value: float | None = None
    profit_percent: float | None = None


class HoldingImportItem(BaseModel):
    source_file: str | None = None
    ticker: str
    name: str | None = None
    shares: float
    open_price: float
    current_price: float | None = None
    purchase_value: float | None = None
    category: str | None = None


class XtbImportCommitRequest(BaseModel):
    items: list[XtbImportItem]
    holdings: list[HoldingImportItem] | None = None
    operations: list["HoldingOperationItem"] | None = None


class HoldingOperationItem(BaseModel):
    source_file: str
    ticker: str | None = None
    operation_type: str
    operation_kind: str | None = None
    description: str | None = None
    amount: float | None = None
    trade_date: str | None = None
    currency: str | None = None


class HoldingTransactionRequest(BaseModel):
    ticker: str
    operation: str
    trade_date: str
    shares: float
    price: float
    fee: float | None = None
    note: str | None = None
    institution: str | None = None
    category: str | None = None
    name: str | None = None


class HoldingMetadataRequest(BaseModel):
    ticker: str
    sector: str | None = None
    industry: str | None = None
    country: str | None = None
    asset_type: str | None = None
    tags: list[str] | None = None


class TagRequest(BaseModel):
    name: str


class PriceRefreshRequest(BaseModel):
    tickers: list[str] | None = None
    force: bool = False


class TickerPriceUpload(BaseModel):
    ticker: str
    price: float
    currency: str = "USD"


class BulkTickerPriceUpload(BaseModel):
    prices: list[TickerPriceUpload]


class TickerMetadata(BaseModel):
    ticker: str
    name: str | None = None
    asset_class: str | None = None
    sector: str | None = None
    industry: str | None = None
    country: str | None = None
    region: str | None = None
    currency: str | None = None
    exchange: str | None = None
    dividend_yield: float | None = None
    dividend_frequency: str | None = None
    next_dividend_date: str | None = None
    next_dividend_amount: float | None = None


class BankingPreviewRow(BaseModel):
    cells: list[str | float | int | None]
    include: bool = True


class BankingCommitRequest(BaseModel):
    source_file: str
    file_hash: str
    institution: str
    columns: list[str]
    mapping: dict[str, int | None]
    rows: list[BankingPreviewRow]


class BankingCategoryUpdateRequest(BaseModel):
    category: str
    subcategory: str | None = None


class BankingBudgetRequest(BaseModel):
    category: str
    amount: float
    month: str | None = None


class ProfileUpdateRequest(BaseModel):
    age: int | None = None


class DebtRequest(BaseModel):
    name: str
    original_amount: float
    current_balance: float
    monthly_payment: float


class GoalCreateRequest(BaseModel):
    name: str


class GoalUpdateRequest(BaseModel):
    name: str


class GoalInputRequest(BaseModel):
    start_date: str | None = None
    duration_years: float | None = None
    sp500_return: float | None = None
    desired_monthly: float | None = None
    planned_monthly: float | None = None
    withdrawal_rate: float | None = None
    initial_investment: float | None = None
    inflation_rate: float | None = None
    portfolio_inflation_rate: float | None = None
    return_method: str | None = None
    simulation_current_age: float | None = None
    simulation_retirement_age: float | None = None
    simulation_annual_spending: float | None = None
    simulation_current_assets: float | None = None
    simulation_monthly_contribution: float | None = None
    simulation_return_rate: float | None = None
    simulation_inflation_rate: float | None = None
    simulation_swr: float | None = None


class GoalContributionRequest(BaseModel):
    contribution_date: str
    amount: float


class CategoryRemoveRequest(BaseModel):
    category: str
    clear_data: bool = False


class CategorySettingRequest(BaseModel):
    category: str
    is_investment: bool


class CategoryAddRequest(BaseModel):
    category: str


@contextmanager
def _db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _init_db() -> None:
    with _db_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY,
                salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                verified INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS verification_codes (
                email TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                expires_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS reset_codes (
                email TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                expires_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS portfolios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_email TEXT NOT NULL,
                name TEXT NOT NULL,
                currency TEXT NOT NULL,
                categories_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(owner_email, name)
            );
            CREATE TABLE IF NOT EXISTS user_profiles (
                email TEXT PRIMARY KEY,
                age INTEGER,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS debts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_email TEXT NOT NULL,
                name TEXT NOT NULL,
                original_amount REAL NOT NULL,
                current_balance REAL NOT NULL,
                monthly_payment REAL NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS goals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_email TEXT NOT NULL,
                name TEXT NOT NULL,
                is_default INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(owner_email, name)
            );
            CREATE TABLE IF NOT EXISTS goal_inputs (
                goal_id INTEGER PRIMARY KEY,
                start_date TEXT NOT NULL,
                duration_years REAL NOT NULL,
                sp500_return REAL NOT NULL,
                desired_monthly REAL NOT NULL,
                planned_monthly REAL NOT NULL,
                withdrawal_rate REAL NOT NULL,
                initial_investment REAL NOT NULL,
                inflation_rate REAL NOT NULL,
                portfolio_inflation_rate REAL,
                simulation_current_age REAL,
                simulation_retirement_age REAL,
                simulation_annual_spending REAL,
                simulation_current_assets REAL,
                simulation_monthly_contribution REAL,
                simulation_return_rate REAL,
                simulation_inflation_rate REAL,
                simulation_swr REAL,
                return_method TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS goal_contributions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                goal_id INTEGER NOT NULL,
                contribution_date TEXT NOT NULL,
                amount REAL NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS santander_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                imported_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS santander_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                section TEXT NOT NULL,
                account TEXT NOT NULL,
                description TEXT,
                balance REAL NOT NULL,
                category TEXT NOT NULL,
                invested REAL,
                gains REAL
            );
            CREATE TABLE IF NOT EXISTS santander_category_map (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                account_key TEXT NOT NULL,
                category TEXT NOT NULL,
                ignore INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, account_key)
            );
            CREATE TABLE IF NOT EXISTS trade_republic_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                available_cash REAL NOT NULL,
                interests_received REAL NOT NULL,
                invested REAL NOT NULL,
                value REAL NOT NULL,
                gains REAL NOT NULL,
                currency TEXT NOT NULL,
                category TEXT,
                source TEXT NOT NULL,
                source_file TEXT,
                file_hash TEXT,
                snapshot_date TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS save_ngrow_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                invested REAL NOT NULL,
                current_value REAL NOT NULL,
                profit_value REAL,
                profit_percent REAL,
                currency TEXT NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                snapshot_date TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS save_ngrow_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                invested_total REAL NOT NULL,
                current_value_total REAL NOT NULL,
                profit_value_total REAL,
                profit_percent_total REAL,
                currency TEXT NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                snapshot_date TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS save_ngrow_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                invested REAL,
                current_value REAL NOT NULL,
                profit_value REAL,
                profit_percent REAL,
                category TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS aforronet_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                invested_total REAL NOT NULL,
                current_value_total REAL NOT NULL,
                category TEXT NOT NULL,
                currency TEXT NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                snapshot_date TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS aforronet_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                invested REAL NOT NULL,
                current_value REAL NOT NULL,
                category TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS bancoinvest_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                snapshot_date TEXT,
                imported_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS bancoinvest_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                holder TEXT NOT NULL,
                invested REAL,
                current_value REAL NOT NULL,
                gains REAL,
                category TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS bancoinvest_category_map (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                holder_key TEXT NOT NULL,
                category TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, holder_key)
            );
            CREATE TABLE IF NOT EXISTS xtb_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                account_type TEXT NOT NULL,
                category TEXT NOT NULL,
                current_value REAL NOT NULL,
                cash_value REAL NOT NULL,
                invested REAL NOT NULL,
                profit_value REAL,
                profit_percent REAL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS holdings_prices (
                ticker TEXT PRIMARY KEY,
                price REAL NOT NULL,
                currency TEXT,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS ticker_metadata (
                ticker TEXT PRIMARY KEY,
                name TEXT,
                asset_class TEXT,
                sector TEXT,
                industry TEXT,
                country TEXT,
                region TEXT,
                currency TEXT,
                exchange TEXT,
                dividend_yield REAL,
                dividend_frequency TEXT,
                next_dividend_date TEXT,
                next_dividend_amount REAL,
                last_updated TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS holdings_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                institution TEXT NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                snapshot_date TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS holdings_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                ticker TEXT NOT NULL,
                name TEXT,
                shares REAL NOT NULL,
                avg_price REAL NOT NULL,
                cost_basis REAL NOT NULL,
                current_price REAL,
                current_value REAL,
                profit_value REAL,
                profit_percent REAL,
                category TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS holdings_metadata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                ticker TEXT NOT NULL,
                sector TEXT,
                industry TEXT,
                country TEXT,
                asset_type TEXT,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, ticker)
            );
            CREATE TABLE IF NOT EXISTS investment_tags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_email TEXT NOT NULL,
                name TEXT NOT NULL,
                name_key TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(owner_email, name_key)
            );
            CREATE TABLE IF NOT EXISTS holding_tags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                ticker TEXT NOT NULL,
                tag_name TEXT NOT NULL,
                tag_key TEXT NOT NULL,
                source TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, ticker, tag_key)
            );
            CREATE TABLE IF NOT EXISTS holding_tag_suppressed (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                ticker TEXT NOT NULL,
                tag_key TEXT NOT NULL,
                removed_at TEXT NOT NULL,
                UNIQUE(portfolio_id, ticker, tag_key)
            );
            CREATE TABLE IF NOT EXISTS holding_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                institution TEXT,
                ticker TEXT NOT NULL,
                name TEXT,
                operation TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                shares REAL NOT NULL,
                price REAL NOT NULL,
                fee REAL,
                note TEXT,
                category TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS holdings_operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER,
                portfolio_id INTEGER NOT NULL,
                source_file TEXT,
                operation_type TEXT NOT NULL,
                operation_kind TEXT,
                ticker TEXT,
                description TEXT,
                amount REAL,
                currency TEXT,
                trade_date TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS portfolio_category_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                is_investment INTEGER NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, category)
            );
            CREATE TABLE IF NOT EXISTS aggregated_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_email TEXT NOT NULL,
                snapshot_date TEXT NOT NULL,
                total_value REAL NOT NULL,
                total_invested REAL NOT NULL,
                total_profit REAL NOT NULL,
                profit_percent REAL NOT NULL,
                totals_by_category_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(owner_email, snapshot_date)
            );
            CREATE TABLE IF NOT EXISTS banking_institutions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, name)
            );
            CREATE TABLE IF NOT EXISTS banking_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                parent_id INTEGER,
                name TEXT NOT NULL,
                is_default INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                UNIQUE(portfolio_id, parent_id, name)
            );
            CREATE TABLE IF NOT EXISTS banking_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                institution TEXT NOT NULL,
                source_file TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                UNIQUE(portfolio_id, file_hash)
            );
            CREATE TABLE IF NOT EXISTS banking_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                portfolio_id INTEGER NOT NULL,
                institution TEXT NOT NULL,
                tx_date TEXT NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL,
                balance REAL,
                currency TEXT NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT NOT NULL,
                raw_json TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS banking_category_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                institution TEXT NOT NULL,
                match_type TEXT NOT NULL,
                match_value TEXT NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, institution, match_type, match_value)
            );
            CREATE TABLE IF NOT EXISTS banking_budgets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                month TEXT NOT NULL,
                amount REAL NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(portfolio_id, category, month)
            );
            """
        )
        columns = conn.execute("PRAGMA table_info(goal_inputs)").fetchall()
        column_names = {row["name"] for row in columns}
        if "portfolio_inflation_rate" not in column_names:
            conn.execute("ALTER TABLE goal_inputs ADD COLUMN portfolio_inflation_rate REAL")
        if "planned_monthly" not in column_names:
            conn.execute("ALTER TABLE goal_inputs ADD COLUMN planned_monthly REAL")
            conn.execute(
                "UPDATE goal_inputs SET planned_monthly = desired_monthly WHERE planned_monthly IS NULL"
            )
        columns = [
            row["name"]
            for row in conn.execute("PRAGMA table_info(sessions)").fetchall()
        ]
        if "expires_at" not in columns:
            conn.execute("ALTER TABLE sessions ADD COLUMN expires_at TEXT")
        item_columns = [
            row["name"]
            for row in conn.execute("PRAGMA table_info(santander_items)").fetchall()
        ]
        if "invested" not in item_columns:
            conn.execute("ALTER TABLE santander_items ADD COLUMN invested REAL")
        if "gains" not in item_columns:
            conn.execute("ALTER TABLE santander_items ADD COLUMN gains REAL")
        trade_columns = [
            row["name"]
            for row in conn.execute("PRAGMA table_info(trade_republic_entries)").fetchall()
        ]
        if "category" not in trade_columns:
            conn.execute("ALTER TABLE trade_republic_entries ADD COLUMN category TEXT")
            conn.execute(
                "UPDATE trade_republic_entries SET category = 'Cash' WHERE category IS NULL"
            )
        if "source_file" not in trade_columns:
            conn.execute("ALTER TABLE trade_republic_entries ADD COLUMN source_file TEXT")
        if "file_hash" not in trade_columns:
            conn.execute("ALTER TABLE trade_republic_entries ADD COLUMN file_hash TEXT")
        if "snapshot_date" not in trade_columns:
            conn.execute("ALTER TABLE trade_republic_entries ADD COLUMN snapshot_date TEXT")
        
        # Migration for ticker_metadata: add next_dividend_amount if not exists
        ticker_metadata_columns = [
            row["name"]
            for row in conn.execute("PRAGMA table_info(ticker_metadata)").fetchall()
        ]
        if ticker_metadata_columns and "next_dividend_amount" not in ticker_metadata_columns:
            conn.execute("ALTER TABLE ticker_metadata ADD COLUMN next_dividend_amount REAL")
        
        settings_columns = [
            row["name"]
            for row in conn.execute(
                "PRAGMA table_info(portfolio_category_settings)"
            ).fetchall()
        ]
        if "is_investment" in settings_columns:
            conn.execute(
                """
                UPDATE portfolio_category_settings
                SET is_investment = 0
                WHERE lower(category) = 'cash' AND is_investment != 0
                """
            )


def _get_user(email: str) -> sqlite3.Row | None:
    with _db_connection() as conn:
        return conn.execute(
            "SELECT email, salt, password_hash, verified, created_at FROM users WHERE email = ?",
            (email,),
        ).fetchone()


def _save_user(email: str, salt: str, password_hash: str, verified: bool, created_at: str) -> None:
    updated_at = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO users (email, salt, password_hash, verified, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (email, salt, password_hash, 1 if verified else 0, created_at, updated_at),
        )


def _set_verified(email: str, verified: bool) -> None:
    with _db_connection() as conn:
        conn.execute(
            "UPDATE users SET verified = ?, updated_at = ? WHERE email = ?",
            (1 if verified else 0, datetime.utcnow().isoformat(), email),
        )


def _set_password(email: str, salt: str, password_hash: str) -> None:
    with _db_connection() as conn:
        conn.execute(
            "UPDATE users SET salt = ?, password_hash = ?, updated_at = ? WHERE email = ?",
            (salt, password_hash, datetime.utcnow().isoformat(), email),
                )


def _get_profile_age(email: str) -> int | None:
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT age FROM user_profiles WHERE email = ?",
            (email,),
        ).fetchone()
    return row["age"] if row and row["age"] is not None else None


def _set_profile_age(email: str, age: int | None) -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO user_profiles (email, age, updated_at)
            VALUES (?, ?, ?)
            """,
            (email, age, datetime.utcnow().isoformat()),
        )


def _get_debt(email: str, debt_id: int) -> dict | None:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT id, name, original_amount, current_balance, monthly_payment, created_at, updated_at
            FROM debts
            WHERE id = ? AND owner_email = ?
            """,
            (debt_id, email),
        ).fetchone()
    if not row:
        return None
    age = _get_profile_age(email)
    original_amount = float(row["original_amount"] or 0)
    current_balance = float(row["current_balance"] or 0)
    monthly_payment = float(row["monthly_payment"] or 0)
    percent_paid = 0.0
    if original_amount > 0:
        percent_paid = (original_amount - current_balance) / original_amount * 100
    months_remaining = None
    if monthly_payment > 0:
        months_remaining = int((current_balance + monthly_payment - 1) // monthly_payment)
    payoff_age = None
    if age is not None and months_remaining is not None:
        payoff_age = age + months_remaining / 12
    return {
        "id": row["id"],
        "name": row["name"],
        "original_amount": original_amount,
        "current_balance": current_balance,
        "monthly_payment": monthly_payment,
        "percent_paid": percent_paid,
        "months_remaining": months_remaining,
        "payoff_age": payoff_age,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _list_debts(email: str) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, original_amount, current_balance, monthly_payment, created_at, updated_at
            FROM debts
            WHERE owner_email = ?
            ORDER BY created_at DESC
            """,
            (email,),
        ).fetchall()
    age = _get_profile_age(email)
    items: list[dict] = []
    for row in rows:
        original_amount = float(row["original_amount"] or 0)
        current_balance = float(row["current_balance"] or 0)
        monthly_payment = float(row["monthly_payment"] or 0)
        percent_paid = 0.0
        if original_amount > 0:
            percent_paid = (original_amount - current_balance) / original_amount * 100
        months_remaining = None
        if monthly_payment > 0:
            months_remaining = int((current_balance + monthly_payment - 1) // monthly_payment)
        payoff_age = None
        if age is not None and months_remaining is not None:
            payoff_age = age + months_remaining / 12
        items.append(
            {
                "id": row["id"],
                "name": row["name"],
                "original_amount": original_amount,
                "current_balance": current_balance,
                "monthly_payment": monthly_payment,
                "percent_paid": percent_paid,
                "months_remaining": months_remaining,
                "payoff_age": payoff_age,
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
            }
        )
    return items


def _create_debt(email: str, payload: DebtRequest) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO debts (owner_email, name, original_amount, current_balance, monthly_payment, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                email,
                payload.name,
                payload.original_amount,
                payload.current_balance,
                payload.monthly_payment,
                now,
                now,
            ),
        )
        debt_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return _get_debt(email, int(debt_id))


def _update_debt(email: str, debt_id: int, payload: DebtRequest) -> dict | None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT id FROM debts WHERE id = ? AND owner_email = ?",
            (debt_id, email),
        ).fetchone()
        if not row:
            return None
        conn.execute(
            """
            UPDATE debts
            SET name = ?, original_amount = ?, current_balance = ?, monthly_payment = ?, updated_at = ?
            WHERE id = ? AND owner_email = ?
            """,
            (
                payload.name,
                payload.original_amount,
                payload.current_balance,
                payload.monthly_payment,
                now,
                debt_id,
                email,
            ),
        )
    return _get_debt(email, debt_id)


def _delete_debt(email: str, debt_id: int) -> bool:
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT id FROM debts WHERE id = ? AND owner_email = ?",
            (debt_id, email),
        ).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM debts WHERE id = ?", (debt_id,))
    return True


def _goal_default_inputs() -> dict:
    default_ecb = _ecb_inflation_10y_avg() or 0.03
    return {
        "start_date": date.today().isoformat(),
        "duration_years": 30.0,
        "sp500_return": 0.1056,
        "desired_monthly": 1000.0,
        "planned_monthly": 1000.0,
        "withdrawal_rate": 0.04,
        "initial_investment": 0.0,
        "inflation_rate": 0.03,
        "portfolio_inflation_rate": default_ecb,
        "simulation_current_age": 30.0,
        "simulation_retirement_age": 67.0,
        "simulation_annual_spending": 30000.0,
        "simulation_current_assets": 100000.0,
        "simulation_monthly_contribution": 500.0,
        "simulation_return_rate": 0.07,
        "simulation_inflation_rate": 0.03,
        "simulation_swr": 0.04,
        "return_method": "cagr",
    }


def _normalize_rate(value: float | None) -> float | None:
    if value is None:
        return None
    if value > 1:
        return value / 100
    return value


def _ecb_inflation_10y_avg() -> float | None:
    raw = os.getenv("ECB_INFLATION_10Y", "").strip()
    if not raw:
        return None
    try:
        value = float(raw.replace(",", "."))
    except ValueError:
        return None
    if value > 1:
        value = value / 100
    if value < 0:
        return None
    return value


def _ensure_goal_input_columns(conn: sqlite3.Connection) -> None:
    columns = conn.execute("PRAGMA table_info(goal_inputs)").fetchall()
    column_names = {row["name"] for row in columns}
    required = [
        "portfolio_inflation_rate",
        "planned_monthly",
        "simulation_current_age",
        "simulation_retirement_age",
        "simulation_annual_spending",
        "simulation_current_assets",
        "simulation_monthly_contribution",
        "simulation_return_rate",
        "simulation_inflation_rate",
        "simulation_swr",
    ]
    for name in required:
        if name not in column_names:
            conn.execute(f"ALTER TABLE goal_inputs ADD COLUMN {name} REAL")


def _parse_iso_date(value: str) -> date:
    try:
        return datetime.fromisoformat(value).date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid date format.") from exc


def _ensure_default_goal(email: str) -> None:
    with _db_connection() as conn:
        _ensure_goal_input_columns(conn)
        row = conn.execute(
            "SELECT id FROM goals WHERE owner_email = ? AND is_default = 1",
            (email,),
        ).fetchone()
        if row:
            return
        now = datetime.utcnow().isoformat()
        conn.execute(
            """
            INSERT INTO goals (owner_email, name, is_default, created_at, updated_at)
            VALUES (?, ?, 1, ?, ?)
            """,
            (email, "FIRE", now, now),
        )
        goal_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        defaults = _goal_default_inputs()
        conn.execute(
            """
            INSERT INTO goal_inputs (
                goal_id,
                start_date,
                duration_years,
                sp500_return,
                desired_monthly,
                planned_monthly,
                withdrawal_rate,
                initial_investment,
                inflation_rate,
                portfolio_inflation_rate,
                simulation_current_age,
                simulation_retirement_age,
                simulation_annual_spending,
                simulation_current_assets,
                simulation_monthly_contribution,
                simulation_return_rate,
                simulation_inflation_rate,
                simulation_swr,
                return_method,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                goal_id,
                defaults["start_date"],
                defaults["duration_years"],
                defaults["sp500_return"],
                defaults["desired_monthly"],
                defaults["planned_monthly"],
                defaults["withdrawal_rate"],
                defaults["initial_investment"],
                defaults["inflation_rate"],
                defaults["portfolio_inflation_rate"],
                defaults["simulation_current_age"],
                defaults["simulation_retirement_age"],
                defaults["simulation_annual_spending"],
                defaults["simulation_current_assets"],
                defaults["simulation_monthly_contribution"],
                defaults["simulation_return_rate"],
                defaults["simulation_inflation_rate"],
                defaults["simulation_swr"],
                defaults["return_method"],
                now,
            ),
        )


def _get_goal(goal_id: int, email: str) -> sqlite3.Row | None:
    with _db_connection() as conn:
        return conn.execute(
            """
            SELECT id, name, is_default, created_at, updated_at
            FROM goals
            WHERE id = ? AND owner_email = ?
            """,
            (goal_id, email),
        ).fetchone()


def _list_goals(email: str) -> list[dict]:
    _ensure_default_goal(email)
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, is_default, created_at, updated_at
            FROM goals
            WHERE owner_email = ?
            ORDER BY is_default DESC, created_at ASC
            """,
            (email,),
        ).fetchall()
    return [
        {
            "id": row["id"],
            "name": row["name"],
            "is_default": bool(row["is_default"]),
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        }
        for row in rows
    ]


def _create_goal(email: str, name: str) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        exists = conn.execute(
            "SELECT id FROM goals WHERE owner_email = ? AND name = ?",
            (email, name),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="Goal name already exists.")
        _ensure_goal_input_columns(conn)
        conn.execute(
            """
            INSERT INTO goals (owner_email, name, is_default, created_at, updated_at)
            VALUES (?, ?, 0, ?, ?)
            """,
            (email, name, now, now),
        )
        goal_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        defaults = _goal_default_inputs()
        conn.execute(
            """
            INSERT INTO goal_inputs (
                goal_id,
                start_date,
                duration_years,
                sp500_return,
                desired_monthly,
                planned_monthly,
                withdrawal_rate,
                initial_investment,
                inflation_rate,
                portfolio_inflation_rate,
                simulation_current_age,
                simulation_retirement_age,
                simulation_annual_spending,
                simulation_current_assets,
                simulation_monthly_contribution,
                simulation_return_rate,
                simulation_inflation_rate,
                simulation_swr,
                return_method,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                goal_id,
                defaults["start_date"],
                defaults["duration_years"],
                defaults["sp500_return"],
                defaults["desired_monthly"],
                defaults["planned_monthly"],
                defaults["withdrawal_rate"],
                defaults["initial_investment"],
                defaults["inflation_rate"],
                defaults["portfolio_inflation_rate"],
                defaults["simulation_current_age"],
                defaults["simulation_retirement_age"],
                defaults["simulation_annual_spending"],
                defaults["simulation_current_assets"],
                defaults["simulation_monthly_contribution"],
                defaults["simulation_return_rate"],
                defaults["simulation_inflation_rate"],
                defaults["simulation_swr"],
                defaults["return_method"],
                now,
            ),
        )
    return _get_goal(int(goal_id), email)


def _update_goal_name(email: str, goal_id: int, name: str) -> dict | None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT id FROM goals WHERE id = ? AND owner_email = ?",
            (goal_id, email),
        ).fetchone()
        if not row:
            return None
        exists = conn.execute(
            "SELECT id FROM goals WHERE owner_email = ? AND name = ? AND id != ?",
            (email, name, goal_id),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="Goal name already exists.")
        conn.execute(
            "UPDATE goals SET name = ?, updated_at = ? WHERE id = ? AND owner_email = ?",
            (name, now, goal_id, email),
        )
    return _get_goal(goal_id, email)


def _delete_goal(email: str, goal_id: int) -> bool:
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT id, is_default FROM goals WHERE id = ? AND owner_email = ?",
            (goal_id, email),
        ).fetchone()
        if not row:
            return False
        if row["is_default"]:
            raise HTTPException(status_code=400, detail="Default goal cannot be deleted.")
        conn.execute("DELETE FROM goal_contributions WHERE goal_id = ?", (goal_id,))
        conn.execute("DELETE FROM goal_inputs WHERE goal_id = ?", (goal_id,))
        conn.execute("DELETE FROM goals WHERE id = ?", (goal_id,))
    return True


def _get_goal_inputs(goal_id: int) -> dict:
    defaults = _goal_default_inputs()
    with _db_connection() as conn:
        _ensure_goal_input_columns(conn)
        row = conn.execute(
            """
            SELECT start_date, duration_years, sp500_return, desired_monthly, planned_monthly,
                   withdrawal_rate, initial_investment, inflation_rate,
                   portfolio_inflation_rate,
                   simulation_current_age,
                   simulation_retirement_age,
                   simulation_annual_spending,
                   simulation_current_assets,
                   simulation_monthly_contribution,
                   simulation_return_rate,
                   simulation_inflation_rate,
                   simulation_swr,
                   return_method, updated_at
            FROM goal_inputs
            WHERE goal_id = ?
            """,
            (goal_id,),
        ).fetchone()
        if not row:
            now = datetime.utcnow().isoformat()
            conn.execute(
                """
                INSERT INTO goal_inputs (
                    goal_id,
                    start_date,
                    duration_years,
                    sp500_return,
                    desired_monthly,
                    planned_monthly,
                    withdrawal_rate,
                    initial_investment,
                    inflation_rate,
                    portfolio_inflation_rate,
                    simulation_current_age,
                    simulation_retirement_age,
                    simulation_annual_spending,
                    simulation_current_assets,
                    simulation_monthly_contribution,
                    simulation_return_rate,
                    simulation_inflation_rate,
                    simulation_swr,
                    return_method,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    goal_id,
                    defaults["start_date"],
                    defaults["duration_years"],
                    defaults["sp500_return"],
                    defaults["desired_monthly"],
                    defaults["planned_monthly"],
                    defaults["withdrawal_rate"],
                    defaults["initial_investment"],
                    defaults["inflation_rate"],
                    defaults["portfolio_inflation_rate"],
                    defaults["simulation_current_age"],
                    defaults["simulation_retirement_age"],
                    defaults["simulation_annual_spending"],
                    defaults["simulation_current_assets"],
                    defaults["simulation_monthly_contribution"],
                    defaults["simulation_return_rate"],
                    defaults["simulation_inflation_rate"],
                    defaults["simulation_swr"],
                    defaults["return_method"],
                    now,
                ),
            )
            return {**defaults, "updated_at": now}
    return {
        "start_date": row["start_date"],
        "duration_years": float(row["duration_years"]),
        "sp500_return": float(row["sp500_return"]),
        "desired_monthly": float(row["desired_monthly"]),
        "planned_monthly": (
            float(row["planned_monthly"])
            if row["planned_monthly"] is not None
            else defaults["planned_monthly"]
        ),
        "withdrawal_rate": float(row["withdrawal_rate"]),
        "initial_investment": float(row["initial_investment"]),
        "inflation_rate": float(row["inflation_rate"]),
        "portfolio_inflation_rate": (
            float(row["portfolio_inflation_rate"])
            if row["portfolio_inflation_rate"] is not None
            else defaults["portfolio_inflation_rate"]
        ),
        "simulation_current_age": (
            float(row["simulation_current_age"])
            if row["simulation_current_age"] is not None
            else defaults["simulation_current_age"]
        ),
        "simulation_retirement_age": (
            float(row["simulation_retirement_age"])
            if row["simulation_retirement_age"] is not None
            else defaults["simulation_retirement_age"]
        ),
        "simulation_annual_spending": (
            float(row["simulation_annual_spending"])
            if row["simulation_annual_spending"] is not None
            else defaults["simulation_annual_spending"]
        ),
        "simulation_current_assets": (
            float(row["simulation_current_assets"])
            if row["simulation_current_assets"] is not None
            else defaults["simulation_current_assets"]
        ),
        "simulation_monthly_contribution": (
            float(row["simulation_monthly_contribution"])
            if row["simulation_monthly_contribution"] is not None
            else defaults["simulation_monthly_contribution"]
        ),
        "simulation_return_rate": (
            float(row["simulation_return_rate"])
            if row["simulation_return_rate"] is not None
            else defaults["simulation_return_rate"]
        ),
        "simulation_inflation_rate": (
            float(row["simulation_inflation_rate"])
            if row["simulation_inflation_rate"] is not None
            else defaults["simulation_inflation_rate"]
        ),
        "simulation_swr": (
            float(row["simulation_swr"])
            if row["simulation_swr"] is not None
            else defaults["simulation_swr"]
        ),
        "return_method": row["return_method"],
        "updated_at": row["updated_at"],
    }


def _update_goal_inputs(goal_id: int, payload: GoalInputRequest) -> dict:
    current = _get_goal_inputs(goal_id)
    start_date = payload.start_date or current["start_date"]
    duration_years = payload.duration_years if payload.duration_years is not None else current["duration_years"]
    sp500_return = _normalize_rate(
        payload.sp500_return if payload.sp500_return is not None else current["sp500_return"]
    )
    desired_monthly = payload.desired_monthly if payload.desired_monthly is not None else current["desired_monthly"]
    planned_monthly = (
        payload.planned_monthly
        if payload.planned_monthly is not None
        else current["planned_monthly"]
    )
    withdrawal_rate = _normalize_rate(
        payload.withdrawal_rate if payload.withdrawal_rate is not None else current["withdrawal_rate"]
    )
    initial_investment = (
        payload.initial_investment
        if payload.initial_investment is not None
        else current["initial_investment"]
    )
    inflation_rate = _normalize_rate(
        payload.inflation_rate if payload.inflation_rate is not None else current["inflation_rate"]
    )
    portfolio_inflation_rate = _normalize_rate(
        payload.portfolio_inflation_rate
        if payload.portfolio_inflation_rate is not None
        else current["portfolio_inflation_rate"]
    )
    simulation_current_age = (
        payload.simulation_current_age
        if payload.simulation_current_age is not None
        else current["simulation_current_age"]
    )
    simulation_retirement_age = (
        payload.simulation_retirement_age
        if payload.simulation_retirement_age is not None
        else current["simulation_retirement_age"]
    )
    simulation_annual_spending = (
        payload.simulation_annual_spending
        if payload.simulation_annual_spending is not None
        else current["simulation_annual_spending"]
    )
    simulation_current_assets = (
        payload.simulation_current_assets
        if payload.simulation_current_assets is not None
        else current["simulation_current_assets"]
    )
    simulation_monthly_contribution = (
        payload.simulation_monthly_contribution
        if payload.simulation_monthly_contribution is not None
        else current["simulation_monthly_contribution"]
    )
    simulation_return_rate = _normalize_rate(
        payload.simulation_return_rate
        if payload.simulation_return_rate is not None
        else current["simulation_return_rate"]
    )
    simulation_inflation_rate = _normalize_rate(
        payload.simulation_inflation_rate
        if payload.simulation_inflation_rate is not None
        else current["simulation_inflation_rate"]
    )
    simulation_swr = _normalize_rate(
        payload.simulation_swr
        if payload.simulation_swr is not None
        else current["simulation_swr"]
    )
    return_method = (payload.return_method or current["return_method"]).lower()
    if return_method not in {"xirr", "cagr"}:
        return_method = "cagr"
    if duration_years <= 0:
        raise HTTPException(status_code=400, detail="Duration must be greater than 0.")
    if withdrawal_rate is not None and withdrawal_rate <= 0:
        raise HTTPException(status_code=400, detail="Withdrawal rate must be greater than 0.")
    if desired_monthly < 0:
        raise HTTPException(status_code=400, detail="Desired monthly must be zero or greater.")
    if planned_monthly < 0:
        raise HTTPException(
            status_code=400,
            detail="Planned monthly contribution must be zero or greater."
        )
    if initial_investment < 0:
        raise HTTPException(status_code=400, detail="Initial investment cannot be negative.")
    if simulation_current_age < 0:
        raise HTTPException(status_code=400, detail="Current age must be zero or greater.")
    if simulation_retirement_age <= simulation_current_age:
        raise HTTPException(
            status_code=400,
            detail="Retirement age must be greater than current age."
        )
    if simulation_annual_spending < 0:
        raise HTTPException(
            status_code=400, detail="Annual spending must be zero or greater."
        )
    if simulation_current_assets < 0:
        raise HTTPException(
            status_code=400, detail="Current assets cannot be negative."
        )
    if simulation_monthly_contribution < 0:
        raise HTTPException(
            status_code=400, detail="Monthly contribution must be zero or greater."
        )
    if sp500_return is None or sp500_return < -0.99:
        raise HTTPException(status_code=400, detail="Return rate must be greater than -99%.")
    if inflation_rate is None or inflation_rate < -0.99:
        raise HTTPException(status_code=400, detail="Inflation rate must be greater than -99%.")
    if portfolio_inflation_rate is None or portfolio_inflation_rate < -0.99:
        raise HTTPException(
            status_code=400, detail="Portfolio inflation rate must be greater than -99%."
        )
    if simulation_return_rate is None or simulation_return_rate < -0.99:
        raise HTTPException(
            status_code=400, detail="Simulation return rate must be greater than -99%."
        )
    if simulation_inflation_rate is None or simulation_inflation_rate < -0.99:
        raise HTTPException(
            status_code=400,
            detail="Simulation inflation rate must be greater than -99%.",
        )
    if simulation_swr is None or simulation_swr <= 0:
        raise HTTPException(
            status_code=400,
            detail="Simulation safe withdrawal rate must be greater than 0.",
        )
    _parse_iso_date(start_date)
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO goal_inputs (
                goal_id,
                start_date,
                duration_years,
                sp500_return,
                desired_monthly,
                planned_monthly,
                withdrawal_rate,
                initial_investment,
                inflation_rate,
                portfolio_inflation_rate,
                simulation_current_age,
                simulation_retirement_age,
                simulation_annual_spending,
                simulation_current_assets,
                simulation_monthly_contribution,
                simulation_return_rate,
                simulation_inflation_rate,
                simulation_swr,
                return_method,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                goal_id,
                start_date,
                duration_years,
                sp500_return,
                desired_monthly,
                planned_monthly,
                withdrawal_rate,
                initial_investment,
                inflation_rate,
                portfolio_inflation_rate,
                simulation_current_age,
                simulation_retirement_age,
                simulation_annual_spending,
                simulation_current_assets,
                simulation_monthly_contribution,
                simulation_return_rate,
                simulation_inflation_rate,
                simulation_swr,
                return_method,
                now,
            ),
        )
    return _get_goal_inputs(goal_id)


def _list_goal_contributions(goal_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, contribution_date, amount, created_at
            FROM goal_contributions
            WHERE goal_id = ?
            ORDER BY contribution_date ASC, id ASC
            """,
            (goal_id,),
        ).fetchall()
    return [
        {
            "id": row["id"],
            "contribution_date": row["contribution_date"],
            "amount": float(row["amount"]),
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def _add_goal_contribution(goal_id: int, payload: GoalContributionRequest) -> dict:
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0.")
    _parse_iso_date(payload.contribution_date)
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO goal_contributions (goal_id, contribution_date, amount, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (goal_id, payload.contribution_date, payload.amount, now),
        )
        contribution_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {
        "id": contribution_id,
        "contribution_date": payload.contribution_date,
        "amount": payload.amount,
        "created_at": now,
    }


def _delete_goal_contribution(goal_id: int, contribution_id: int) -> bool:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT id
            FROM goal_contributions
            WHERE id = ? AND goal_id = ?
            """,
            (contribution_id, goal_id),
        ).fetchone()
        if not row:
            return False
        conn.execute(
            "DELETE FROM goal_contributions WHERE id = ? AND goal_id = ?",
            (contribution_id, goal_id),
        )
    return True


def _goal_months_elapsed(start_date: date, now_date: date) -> int:
    if now_date <= start_date:
        return 0
    return max(0, int((now_date - start_date).days / 30.4375))


def _goal_years_elapsed(start_date: date, now_date: date) -> float:
    if now_date <= start_date:
        return 0.0
    return (now_date - start_date).days / 365.25


def _xirr(flows: list[tuple[date, float]], guess: float = 0.1) -> float | None:
    if not flows:
        return None
    flows = sorted(flows, key=lambda item: item[0])
    start_date = flows[0][0]
    if all(amount >= 0 for _, amount in flows) or all(amount <= 0 for _, amount in flows):
        return None

    def npv(rate: float) -> float:
        total = 0.0
        for dt_value, amount in flows:
            days = (dt_value - start_date).days / 365.25
            total += amount / ((1 + rate) ** days)
        return total

    def d_npv(rate: float) -> float:
        total = 0.0
        for dt_value, amount in flows:
            days = (dt_value - start_date).days / 365.25
            total -= (days * amount) / ((1 + rate) ** (days + 1))
        return total

    rate = guess
    for _ in range(50):
        value = npv(rate)
        deriv = d_npv(rate)
        if abs(value) < 1e-6:
            return rate
        if deriv == 0:
            break
        rate_next = rate - value / deriv
        if rate_next <= -0.9999 or math.isnan(rate_next) or math.isinf(rate_next):
            break
        rate = rate_next
    return None


def _nper(rate: float, payment: float, present: float, future: float) -> float | None:
    if payment == 0 and rate == 0:
        return None
    if rate == 0:
        return -((present + future) / payment)
    top = payment - future * rate
    bottom = payment + present * rate
    if bottom == 0:
        return None
    ratio = top / bottom
    if ratio <= 0:
        return None
    return math.log(ratio) / math.log(1 + rate)


def _goal_projection_series(
    invested_total: float,
    avg_monthly: float,
    annual_return: float,
    duration_years: float,
    coast_years: float | None,
    fire_target: float | None,
    discount_rate: float | None,
) -> list[dict]:
    if duration_years <= 0:
        return []
    annual_return = min(max(annual_return, -0.99), 1.0)
    rm = (1 + annual_return) ** (1 / 12) - 1 if annual_return != 0 else 0
    coast_rate = discount_rate if discount_rate is not None else annual_return
    coast_rate = min(max(coast_rate, -0.99), 1.0)
    points: list[dict] = []
    coast_months = int(coast_years * 12) if coast_years is not None else None
    value_at_coast = None
    if coast_months is not None:
        if rm == 0:
            value_at_coast = invested_total + avg_monthly * coast_months
        else:
            value_at_coast = invested_total * (1 + rm) ** coast_months + avg_monthly * (
                ((1 + rm) ** coast_months - 1) / rm
            )
    for year in range(int(duration_years) + 1):
        months = year * 12
        if rm == 0:
            with_contrib = invested_total + avg_monthly * months
        else:
            with_contrib = invested_total * (1 + rm) ** months + avg_monthly * (
                ((1 + rm) ** months - 1) / rm
            )
        without_contrib = None
        if coast_months is not None and value_at_coast is not None:
            if months <= coast_months:
                without_contrib = with_contrib
            else:
                without_contrib = value_at_coast * (1 + rm) ** (months - coast_months)
        coast_target = None
        if fire_target is not None:
            years_left = duration_years - year
            if years_left < 0:
                years_left = 0
            if coast_rate <= -0.99:
                coast_target = None
            elif coast_rate == 0:
                coast_target = fire_target
            else:
                coast_target = fire_target / ((1 + coast_rate) ** years_left)
        points.append(
            {
                "year": year,
                "with_contrib": with_contrib,
                "without_contrib": without_contrib,
                "coast_target": coast_target,
            }
        )
    return points


def _goal_summary(
    email: str, goal_id: int, portfolio_id: int | None = None
) -> dict:
    inputs = _get_goal_inputs(goal_id)
    start_date = _parse_iso_date(inputs["start_date"])
    now_date = datetime.utcnow().date()
    duration_years = float(inputs["duration_years"])
    expected_return = float(inputs["sp500_return"])
    desired_monthly = float(inputs["desired_monthly"])
    planned_monthly_raw = inputs.get("planned_monthly")
    planned_monthly = (
        float(planned_monthly_raw)
        if planned_monthly_raw is not None
        else desired_monthly
    )
    withdrawal_rate = float(inputs["withdrawal_rate"])
    initial_investment = float(inputs["initial_investment"])
    inflation_rate = float(inputs["inflation_rate"])
    portfolio_inflation_rate = float(
        inputs.get("portfolio_inflation_rate") or inflation_rate
    )
    simulation_current_age = float(inputs.get("simulation_current_age"))
    simulation_retirement_age = float(inputs.get("simulation_retirement_age"))
    simulation_annual_spending = float(inputs.get("simulation_annual_spending"))
    simulation_current_assets = float(inputs.get("simulation_current_assets"))
    simulation_monthly_contribution = float(inputs.get("simulation_monthly_contribution"))
    simulation_return_rate = float(inputs.get("simulation_return_rate"))
    simulation_inflation_rate = float(inputs.get("simulation_inflation_rate"))
    simulation_swr = float(inputs.get("simulation_swr"))
    contributions = _list_goal_contributions(goal_id)
    if contributions:
        latest_contribution = max(
            _parse_iso_date(item["contribution_date"]) for item in contributions
        )
        if latest_contribution <= now_date:
            now_date = latest_contribution
    total_contrib = sum(item["amount"] for item in contributions)
    contrib_count = len(contributions)
    invested_total = initial_investment + total_contrib
    avg_contribution = total_contrib / contrib_count if contrib_count else 0.0
    past_contributions = []
    for item in contributions:
        item_date = _parse_iso_date(item["contribution_date"])
        if item_date <= now_date:
            past_contributions.append(item)
    past_total = sum(item["amount"] for item in past_contributions)
    invested_total_to_date = initial_investment + past_total

    current_value = invested_total
    portfolio_summary = None
    if portfolio_id is not None:
        portfolio = _get_portfolio(portfolio_id, email)
        if portfolio:
            categories = _filter_categories(json.loads(portfolio["categories_json"]))
            _ensure_category_settings(portfolio_id, categories)
            settings = _get_category_settings(portfolio_id)
            settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
            totals, _, _, investment_current_total, _ = _aggregate_latest_totals(
                portfolio_id, settings_lookup
            )
            if totals:
                portfolio_total = sum(totals.values())
                if investment_current_total is not None and investment_current_total > 0:
                    current_value = investment_current_total
                else:
                    current_value = portfolio_total
                portfolio_summary = {
                    "portfolio_total": portfolio_total,
                    "investment_total": investment_current_total,
                }

    xirr_value = current_value
    if initial_investment > 0:
        xirr_value = initial_investment

    years_elapsed = _goal_years_elapsed(start_date, now_date)
    years_remaining = max(0.0, duration_years - years_elapsed)

    rate_method = inputs.get("return_method", "cagr").lower()
    return_rate = None
    if rate_method == "xirr":
        flows: list[tuple[date, float]] = []
        if initial_investment > 0:
            flows.append((start_date, -initial_investment))
        for item in past_contributions:
            flows.append((_parse_iso_date(item["contribution_date"]), -item["amount"]))
        if flows:
            flows.append((now_date, xirr_value))
            result = _xirr(flows)
            if result is not None:
                return_rate = result * 100
    else:
        if invested_total_to_date > 0 and current_value > 0 and years_elapsed > 0:
            return_rate = (
                (current_value / invested_total_to_date) ** (1 / years_elapsed) - 1
            ) * 100

    calc_return = expected_return
    if return_rate is not None and math.isfinite(return_rate):
        calc_return = return_rate / 100
    if calc_return is None or not math.isfinite(calc_return) or calc_return <= -0.99:
        calc_return = expected_return
    if calc_return is None:
        calc_return = expected_return
    calc_return = min(max(calc_return, -0.99), 1.0)
    expected_return = min(max(expected_return, -0.99), 1.0)
    assumption_return = expected_return * 100

    def build_section(
        invested_base: float,
        avg_monthly: float,
        inflation_value: float,
        discount_rate: float,
        desired_monthly_value: float,
        current_value_override: float,
    ) -> dict:
        discount_rate = min(max(discount_rate, -0.99), 1.0)
        future_value_1000 = 1000 * ((1 + inflation_value) ** duration_years)
        desired_future = desired_monthly_value * ((1 + inflation_value) ** duration_years)
        fire_target = (
            (desired_future * 12 / withdrawal_rate) if withdrawal_rate > 0 else None
        )

        coast_years = None
        coast_status = "ok"
        if fire_target is None:
            coast_status = "missing"
        else:
            if invested_base >= fire_target:
                coast_years = 0.0
                coast_status = "achieved"
            else:
                rm = (1 + discount_rate) ** (1 / 12) - 1 if discount_rate != 0 else 0
                t_months = int(duration_years * 12)
                if t_months <= 0:
                    coast_status = "imp"
                else:
                    found = None
                    if rm == 0:
                        for month in range(1, t_months + 1):
                            balance = invested_base + avg_monthly * month
                            value_future = balance
                            if value_future >= fire_target:
                                found = month
                                break
                    else:
                        for month in range(1, t_months + 1):
                            balance = invested_base * (1 + rm) ** month + avg_monthly * (
                                ((1 + rm) ** month - 1) / rm
                            )
                            value_future = balance * (1 + rm) ** (t_months - month)
                            if value_future >= fire_target:
                                found = month
                                break
                    if found is not None:
                        coast_years = found / 12
                    else:
                        coast_status = "imp"
        if coast_years is not None and years_elapsed >= coast_years:
            coast_status = "achieved"

        fire_years = None
        fire_months = None
        fire_status = "ok"
        if fire_target is None:
            fire_status = "missing"
        else:
          rm = (1 + discount_rate) ** (1 / 12) - 1 if discount_rate != 0 else 0
          months_to_fire = _nper(rm, -avg_monthly, -invested_base, fire_target)
          if months_to_fire is None:
              fire_status = "imp"
          else:
                fire_years = months_to_fire / 12
                if fire_years > duration_years:
                    fire_status = "imp"
                else:
                    years_int = int(fire_years)
                    months_int = int(round((fire_years - years_int) * 12))
                    fire_years = float(years_int)
                    fire_months = months_int

        projection = _goal_projection_series(
            invested_base,
            avg_monthly,
            expected_return,
            duration_years,
            coast_years,
            fire_target,
            discount_rate,
        )
        coast_target_value = None
        if projection:
            coast_target_value = projection[0].get("coast_target")

        return {
            "metrics": {
                "years_elapsed": years_elapsed,
                "years_remaining": years_remaining,
                "avg_monthly": avg_monthly,
                "invested_total": invested_base,
                "current_value": current_value_override,
                "return_rate": return_rate,
                "return_method": rate_method,
                "assumption_return": assumption_return,
                "future_value_1000": future_value_1000,
                "fire_target": fire_target,
                "coast_target": coast_target_value,
                "coast_years": coast_years,
                "coast_status": coast_status,
                "fire_years": fire_years,
                "fire_months": fire_months,
                "fire_status": fire_status,
                "inflation_rate": inflation_value,
            },
            "projection": projection,
        }

    def build_walletburst_section() -> dict:
        current_age = max(0.0, simulation_current_age)
        retirement_age = max(current_age, simulation_retirement_age)
        years_to_retire = max(0.0, retirement_age - current_age)
        annual_spending = max(0.0, simulation_annual_spending)
        current_assets = max(0.0, simulation_current_assets)
        monthly_contribution = max(0.0, simulation_monthly_contribution)
        invest_rate = min(max(simulation_return_rate, -0.99), 1.0)
        infl_rate = min(max(simulation_inflation_rate, -0.99), 1.0)
        swr_rate = min(max(simulation_swr, 0.0), 1.0)
        adjusted_return = min(max(invest_rate - infl_rate, -0.99), 1.0)

        fire_target = None
        if swr_rate > 0:
            fire_target = annual_spending / swr_rate

        coast_target = None
        if fire_target is not None:
            if years_to_retire <= 0 or adjusted_return == 0:
                coast_target = fire_target
            else:
                coast_target = fire_target / ((1 + adjusted_return) ** years_to_retire)

        rm = 0.0
        if adjusted_return != 0:
            rm = (1 + adjusted_return) ** (1 / 12) - 1
        t_months = int(round(years_to_retire * 12))

        coast_months = None
        coast_status = "ok"
        value_at_coast = None
        if fire_target is None:
            coast_status = "missing"
        else:
            if current_assets >= fire_target:
                coast_months = 0
                coast_status = "achieved"
            elif t_months <= 0:
                coast_status = "imp"
            else:
                found = None
                if rm == 0:
                    for month in range(1, t_months + 1):
                        balance = current_assets + monthly_contribution * month
                        if balance >= fire_target:
                            found = month
                            break
                else:
                    for month in range(1, t_months + 1):
                        balance = current_assets * (1 + rm) ** month + monthly_contribution * (
                            ((1 + rm) ** month - 1) / rm
                        )
                        value_future = balance * (1 + rm) ** (t_months - month)
                        if value_future >= fire_target:
                            found = month
                            break
                if found is not None:
                    coast_months = found
                else:
                    coast_status = "imp"
        if coast_months is not None:
            if rm == 0:
                value_at_coast = current_assets + monthly_contribution * coast_months
            else:
                value_at_coast = current_assets * (1 + rm) ** coast_months + monthly_contribution * (
                    ((1 + rm) ** coast_months - 1) / rm
                )

        projection = []
        years_steps = int(math.floor(years_to_retire))
        for year in range(years_steps + 1):
            months = year * 12
            if rm == 0:
                with_contrib = current_assets + monthly_contribution * months
            else:
                with_contrib = current_assets * (1 + rm) ** months + monthly_contribution * (
                    ((1 + rm) ** months - 1) / rm
                )
            without_contrib = None
            if coast_months is not None and value_at_coast is not None:
                if months <= coast_months:
                    without_contrib = with_contrib
                else:
                    without_contrib = value_at_coast * (1 + rm) ** (months - coast_months)
            coast_target_point = None
            if fire_target is not None:
                years_left = years_to_retire - year
                if years_left < 0:
                    years_left = 0
                if adjusted_return == 0:
                    coast_target_point = fire_target
                else:
                    coast_target_point = fire_target / ((1 + adjusted_return) ** years_left)
            projection.append(
                {
                    "year": year,
                    "age": current_age + year,
                    "with_contrib": with_contrib,
                    "without_contrib": without_contrib,
                    "coast_target": coast_target_point,
                }
            )

        return {
            "metrics": {
                "current_age": current_age,
                "retirement_age": retirement_age,
                "years_to_retire": years_to_retire,
                "annual_spending": annual_spending,
                "current_assets": current_assets,
                "monthly_contribution": monthly_contribution,
                "investment_return": invest_rate * 100,
                "inflation_rate": infl_rate,
                "swr": swr_rate,
                "adjusted_return": adjusted_return * 100,
                "fire_target": fire_target,
                "coast_target": coast_target,
                "coast_years": coast_months / 12 if coast_months is not None else None,
                "coast_status": coast_status,
            },
            "projection": projection,
        }

    inflation_portfolio = portfolio_inflation_rate or _ecb_inflation_10y_avg() or inflation_rate
    portfolio_section = build_section(
        invested_total,
        avg_contribution,
        inflation_portfolio,
        calc_return,
        desired_monthly,
        current_value,
    )
    simulation_section = build_walletburst_section()

    return {
        "inputs": inputs,
        "contributions": contributions,
        "portfolio": portfolio_summary,
        "portfolio_fire": portfolio_section,
        "simulation_fire": simulation_section,
    }


_init_db()


def _store_code(table: str, email: str, code: str, expires_at: str) -> None:
    with _db_connection() as conn:
        conn.execute(
            f"INSERT OR REPLACE INTO {table} (email, code, expires_at) VALUES (?, ?, ?)",
            (email, code, expires_at),
        )


def _get_code(table: str, email: str) -> sqlite3.Row | None:
    with _db_connection() as conn:
        return conn.execute(
            f"SELECT code, expires_at FROM {table} WHERE email = ?",
            (email,),
        ).fetchone()


def _delete_code(table: str, email: str) -> None:
    with _db_connection() as conn:
        conn.execute(f"DELETE FROM {table} WHERE email = ?", (email,))


def _store_session(token: str, email: str) -> None:
    expires_at = datetime.utcnow() + timedelta(hours=SESSION_TTL_HOURS)
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO sessions (token, email, created_at, expires_at)
            VALUES (?, ?, ?, ?)
            """,
            (token, email, datetime.utcnow().isoformat(), expires_at.isoformat()),
        )


def _get_session(token: str) -> sqlite3.Row | None:
    with _db_connection() as conn:
        return conn.execute(
            "SELECT token, email, created_at, expires_at FROM sessions WHERE token = ?",
            (token,),
        ).fetchone()


def _delete_session(token: str) -> None:
    with _db_connection() as conn:
        conn.execute("DELETE FROM sessions WHERE token = ?", (token,))


def _delete_sessions_for_email(email: str) -> None:
    with _db_connection() as conn:
        conn.execute("DELETE FROM sessions WHERE email = ?", (email,))


def _cleanup_expired_sessions() -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute("DELETE FROM sessions WHERE expires_at <= ?", (now,))


def _cleanup_expired_codes() -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute("DELETE FROM verification_codes WHERE expires_at <= ?", (now,))
        conn.execute("DELETE FROM reset_codes WHERE expires_at <= ?", (now,))


def _validate_email(email: str) -> None:
    if not EMAIL_REGEX.match(email):
        raise HTTPException(status_code=400, detail="Invalid email format.")


def _hash_password(password: str, salt: str) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), 100_000
    ).hex()


def _issue_code(email: str) -> str:
    code = f"{secrets.randbelow(1_000_000):06d}"
    expires_at = datetime.utcnow() + timedelta(minutes=CODE_TTL_MINUTES)
    _store_code("verification_codes", email, code, expires_at.isoformat())
    return code


def _issue_reset_code(email: str) -> str:
    code = f"{secrets.randbelow(1_000_000):06d}"
    expires_at = datetime.utcnow() + timedelta(minutes=CODE_TTL_MINUTES)
    _store_code("reset_codes", email, code, expires_at.isoformat())
    return code


def _verification_email(code: str) -> tuple[str, str]:
    text = (
        f"Your verification code is {code}. "
        f"It expires in {CODE_TTL_MINUTES} minutes."
    )
    html = (
        f"<p>Your verification code is <strong>{code}</strong>. "
        f"It expires in {CODE_TTL_MINUTES} minutes.</p>"
    )
    return text, html


def _reset_email(code: str) -> tuple[str, str]:
    text = (
        f"Your password reset code is {code}. "
        f"It expires in {CODE_TTL_MINUTES} minutes."
    )
    html = (
        f"<p>Your password reset code is <strong>{code}</strong>. "
        f"It expires in {CODE_TTL_MINUTES} minutes.</p>"
    )
    return text, html


def _send_email(to_email: str, subject: str, text: str, html: str | None = None) -> None:
    host = os.getenv("SMTP_HOST")
    port = int(os.getenv("SMTP_PORT", "587"))
    username = os.getenv("SMTP_USERNAME")
    password = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("SMTP_FROM") or username
    use_tls = os.getenv("SMTP_TLS", "true").lower() == "true"
    use_ssl = os.getenv("SMTP_SSL", "false").lower() == "true"
    if not host or not username or not password or not from_email:
        raise HTTPException(status_code=500, detail="Email service not configured.")

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = from_email
    message["To"] = to_email
    message.set_content(text)
    if html:
        message.add_alternative(html, subtype="html")

    if use_ssl:
        with smtplib.SMTP_SSL(host, port) as server:
            server.login(username, password)
            server.send_message(message)
        return

    with smtplib.SMTP(host, port) as server:
        server.ehlo()
        if use_tls:
            server.starttls()
            server.ehlo()
        server.login(username, password)
        server.send_message(message)


def _send_email_async(
    to_email: str, subject: str, text: str, html: str | None = None
) -> None:
    thread = threading.Thread(
        target=_send_email, args=(to_email, subject, text, html), daemon=True
    )
    thread.start()


def _verify_code(email: str, code: str) -> None:
    _cleanup_expired_codes()
    record = _get_code("verification_codes", email)
    if not record:
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    if record["code"] != code:
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    expires = datetime.fromisoformat(record["expires_at"])
    if expires < datetime.utcnow():
        _delete_code("verification_codes", email)
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    _delete_code("verification_codes", email)


def _verify_reset_code(email: str, code: str) -> None:
    _cleanup_expired_codes()
    record = _get_code("reset_codes", email)
    if not record:
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    if record["code"] != code:
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    expires = datetime.fromisoformat(record["expires_at"])
    if expires < datetime.utcnow():
        _delete_code("reset_codes", email)
        raise HTTPException(status_code=400, detail="Invalid or expired code.")
    _delete_code("reset_codes", email)


def _issue_session(email: str) -> str:
    _cleanup_expired_sessions()
    token = secrets.token_hex(24)
    _store_session(token, email)
    return token


def _require_token(authorization: str | None) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing authorization.")
    if not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Invalid authorization.")
    return authorization.replace("Bearer ", "").strip()


def _require_session(authorization: str | None) -> sqlite3.Row:
    token = _require_token(authorization)
    _cleanup_expired_sessions()
    session = _get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session.")
    if session["expires_at"] and datetime.fromisoformat(session["expires_at"]) < datetime.utcnow():
        _delete_session(token)
        raise HTTPException(status_code=401, detail="Session expired.")
    return session


def _require_admin(authorization: str | None) -> sqlite3.Row:
    """Verifica se o usuário é administrador."""
    session = _require_session(authorization)
    if session["email"] != ADMIN_USERNAME:
        raise HTTPException(status_code=403, detail="Admin access required.")
    return session


def _fetch_ticker_metadata_yfinance(ticker: str) -> dict | None:
    """Busca metadados de um ticker usando yfinance (Yahoo Finance).
    
    Usa múltiplas estratégias para evitar rate limiting:
    1. Tenta .info com delays
    2. Fallback para download() que é menos bloqueado
    3. Retorna dados básicos se tudo falhar
    """
    try:
        import yfinance as yf
        import time
        import random
        
        # Add random delay to avoid detection (3-6 seconds)
        delay = random.uniform(3.0, 6.0)
        print(f"Waiting {delay:.1f}s before fetching {ticker}...")
        time.sleep(delay)
        
        stock = yf.Ticker(ticker)
        
        # Strategy 1: Try .info (often blocked)
        info = None
        try:
            info = stock.info
            if info and "symbol" in info:
                print(f"✓ Got full metadata for {ticker} via .info")
        except Exception as info_error:
            print(f"⚠️ .info failed for {ticker}: {str(info_error)[:50]}")
        
        # Strategy 2: Try download() method (less blocked)
        if not info:
            try:
                print(f"Trying alternate method for {ticker}...")
                import pandas as pd
                # Download uses different endpoint, often works when .info fails
                hist = yf.download(ticker, period="5d", progress=False)
                if not hist.empty:
                    # Get basic info from history
                    latest_price = hist["Close"].iloc[-1]
                    print(f"✓ Got historical data for {ticker}, creating basic metadata")
                    return {
                        "ticker": ticker.upper(),
                        "name": ticker.upper(),
                        "asset_class": "Stock",
                        "sector": None,
                        "industry": None,
                        "country": None,
                        "region": None,
                        "currency": "USD",
                        "exchange": None,
                        "dividend_yield": None,
                        "dividend_frequency": None,
                        "next_dividend_date": None
                    }
            except Exception as download_error:
                print(f"⚠️ download() also failed for {ticker}: {str(download_error)[:50]}")
        
        # If we got info data, parse it
        if info and "symbol" in info:
            # Determinar asset class
            asset_class = "Stock"
            quote_type = info.get("quoteType", "").upper()
            if quote_type == "ETF":
                asset_class = "ETF"
            elif "REIT" in info.get("longBusinessSummary", "").upper() or info.get("industry") == "REIT":
                asset_class = "REIT"
            
            # Calcular próxima data de dividendo (se aplicável)
            next_div_date = None
            if info.get("exDividendDate"):
                try:
                    import datetime as dt
                    next_div_date = dt.datetime.fromtimestamp(info["exDividendDate"]).strftime("%Y-%m-%d")
                except:
                    pass
            
            return {
                "ticker": ticker.upper(),
                "name": info.get("longName") or info.get("shortName"),
                "asset_class": asset_class,
                "sector": info.get("sector"),
                "industry": info.get("industry"),
                "country": info.get("country"),
                "region": info.get("region"),
                "currency": info.get("currency"),
                "exchange": info.get("exchange"),
                "dividend_yield": info.get("dividendYield"),
                "dividend_frequency": "Quarterly" if info.get("dividendYield") else None,
                "next_dividend_date": next_div_date
            }
        
        print(f"✗ All methods failed for {ticker}")
        return None
        
    except Exception as e:
        print(f"✗ Error fetching metadata for {ticker}: {e}")
        return None


def _fetch_ticker_price_yfinance(ticker: str) -> dict | None:
    """Busca preço atual de um ticker usando yfinance."""
    try:
        import yfinance as yf
        import time
        import random
        
        # Add random delay (2-3 seconds)
        time.sleep(random.uniform(2.0, 3.0))
        
        # Try download() first (less blocked)
        try:
            hist = yf.download(ticker, period="1d", progress=False)
            if not hist.empty:
                latest_price = hist["Close"].iloc[-1]
                return {
                    "ticker": ticker.upper(),
                    "price": float(latest_price),
                    "currency": "EUR" if ticker.endswith(".DE") else "USD"
                }
        except Exception as e:
            print(f"yf.download failed for {ticker}: {str(e)[:100]}")
        
        # Fallback to stock.history()
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(period="1d")
            
            if hist.empty:
                print(f"No history data for {ticker}")
                return None
            
            latest_price = hist["Close"].iloc[-1]
            
            # Try to get currency, fallback based on ticker suffix
            currency = "USD"
            if ticker.endswith(".DE"):
                currency = "EUR"
            elif ticker.endswith(".L"):
                currency = "GBP"
            
            try:
                info = stock.info
                if info and "currency" in info:
                    currency = info.get("currency", currency)
            except Exception as e:
                print(f"Could not fetch info for {ticker}, using {currency}: {str(e)[:100]}")
            
            return {
                "ticker": ticker.upper(),
                "price": float(latest_price),
                "currency": currency
            }
        except Exception as e:
            print(f"yf.Ticker.history failed for {ticker}: {str(e)[:100]}")
            return None
        
    except Exception as e:
        print(f"Error fetching price for {ticker}: {e}")
        return None


def _save_ticker_metadata(metadata: dict) -> None:
    """Salva ou atualiza metadados de um ticker."""
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO ticker_metadata (
                ticker, name, asset_class, sector, industry, country, region,
                currency, exchange, dividend_yield, dividend_frequency,
                next_dividend_date, next_dividend_amount, last_updated
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(ticker) DO UPDATE SET
                name = excluded.name,
                asset_class = excluded.asset_class,
                sector = excluded.sector,
                industry = excluded.industry,
                country = excluded.country,
                region = excluded.region,
                currency = excluded.currency,
                exchange = excluded.exchange,
                dividend_yield = excluded.dividend_yield,
                dividend_frequency = excluded.dividend_frequency,
                next_dividend_date = excluded.next_dividend_date,
                next_dividend_amount = excluded.next_dividend_amount,
                last_updated = excluded.last_updated
            """,
            (
                metadata.get("ticker"),
                metadata.get("name"),
                metadata.get("asset_class"),
                metadata.get("sector"),
                metadata.get("industry"),
                metadata.get("country"),
                metadata.get("region"),
                metadata.get("currency"),
                metadata.get("exchange"),
                metadata.get("dividend_yield"),
                metadata.get("dividend_frequency"),
                metadata.get("next_dividend_date"),
                metadata.get("next_dividend_amount"),
                now
            ),
        )


def _get_ticker_metadata(ticker: str) -> dict | None:
    """Obtém metadados de um ticker da base de dados."""
    with _db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM ticker_metadata WHERE ticker = ?",
            (ticker.upper(),)
        ).fetchone()
        if not row:
            return None
        return dict(row)


def _list_portfolios(owner_email: str) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   name,
                   currency,
                   categories_json,
                   created_at,
                   (
                       EXISTS(
                           SELECT 1 FROM santander_imports WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM trade_republic_entries WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM save_ngrow_imports WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM save_ngrow_entries WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM aforronet_imports WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM xtb_imports WHERE portfolio_id = portfolios.id
                       )
                       OR EXISTS(
                           SELECT 1 FROM bancoinvest_imports WHERE portfolio_id = portfolios.id
                       )
                   ) AS has_data
            FROM portfolios
            WHERE owner_email = ?
            ORDER BY created_at ASC
            """,
            (owner_email,),
        ).fetchall()
    portfolios = []
    for row in rows:
        portfolios.append(
            {
                "id": row["id"],
                "name": row["name"],
                "currency": row["currency"],
                "categories": _filter_categories(json.loads(row["categories_json"])),
                "created_at": row["created_at"],
                "has_data": bool(row["has_data"]),
            }
        )
    return portfolios


def _create_portfolio(owner_email: str, name: str, currency: str, categories: list[str]) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO portfolios (owner_email, name, currency, categories_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (owner_email, name, currency, json.dumps(categories), now, now),
        )
        portfolio_id = cursor.lastrowid
    _ensure_category_settings(portfolio_id, categories)
    _ensure_banking_categories(portfolio_id)
    return {
        "id": portfolio_id,
        "name": name,
        "currency": currency,
        "categories": categories,
        "created_at": now,
        "has_data": False,
    }


def _get_portfolio(portfolio_id: int, owner_email: str) -> sqlite3.Row | None:
    with _db_connection() as conn:
        return conn.execute(
            """
            SELECT id, name, currency, categories_json
            FROM portfolios
            WHERE id = ? AND owner_email = ?
            """,
            (portfolio_id, owner_email),
        ).fetchone()




def _normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    normalized = unicodedata.normalize("NFKD", value)
    cleaned = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return cleaned.strip().lower()


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_tag_name(value: str | None) -> tuple[str, str]:
    if value is None:
        return ("", "")
    cleaned = " ".join(str(value).strip().split())
    if not cleaned:
        return ("", "")
    return cleaned, _normalize_text(cleaned)


def _list_custom_investment_tags(owner_email: str) -> list[str]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT name
            FROM investment_tags
            WHERE owner_email = ?
            ORDER BY name_key
            """,
            (owner_email,),
        ).fetchall()
    return [row["name"] for row in rows]


def _list_investment_tags(owner_email: str) -> dict:
    custom_tags = _list_custom_investment_tags(owner_email)
    combined = {tag.lower(): tag for tag in DEFAULT_INVESTMENT_TAGS}
    for tag in custom_tags:
        combined[_normalize_text(tag)] = tag
    ordered = [combined[key] for key in sorted(combined.keys())]
    return {"items": ordered, "custom": custom_tags}


def _save_investment_tag(owner_email: str, tag_name: str) -> dict:
    cleaned, tag_key = _normalize_tag_name(tag_name)
    if not cleaned:
        raise HTTPException(status_code=400, detail="Tag name is required.")
    if len(cleaned) > MAX_TAG_LENGTH:
        raise HTTPException(status_code=400, detail="Tag name is too long.")
    if tag_key in {_normalize_text(tag) for tag in DEFAULT_INVESTMENT_TAGS}:
        raise HTTPException(status_code=400, detail="Tag already exists.")
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        try:
            conn.execute(
                """
                INSERT INTO investment_tags (owner_email, name, name_key, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (owner_email, cleaned, tag_key, now, now),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Tag already exists.")
    return {"name": cleaned, "created_at": now}


def _delete_investment_tag(owner_email: str, tag_name: str) -> None:
    cleaned, tag_key = _normalize_tag_name(tag_name)
    if not cleaned:
        raise HTTPException(status_code=400, detail="Tag name is required.")
    if tag_key in {_normalize_text(tag) for tag in DEFAULT_INVESTMENT_TAGS}:
        raise HTTPException(status_code=400, detail="Default tags cannot be removed.")
    with _db_connection() as conn:
        conn.execute(
            "DELETE FROM investment_tags WHERE owner_email = ? AND name_key = ?",
            (owner_email, tag_key),
        )
        conn.execute(
            "DELETE FROM holding_tags WHERE portfolio_id IN (SELECT id FROM portfolios WHERE owner_email = ?) AND tag_key = ?",
            (owner_email, tag_key),
        )


def _list_holding_tags(portfolio_id: int) -> dict[str, list[str]]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT ticker, tag_name
            FROM holding_tags
            WHERE portfolio_id = ?
            ORDER BY tag_key
            """,
            (portfolio_id,),
        ).fetchall()
    tags: dict[str, list[str]] = {}
    for row in rows:
        tags.setdefault(row["ticker"].upper(), []).append(row["tag_name"])
    return tags


def _list_suppressed_tags(portfolio_id: int) -> dict[str, set[str]]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT ticker, tag_key
            FROM holding_tag_suppressed
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
    suppressed: dict[str, set[str]] = {}
    for row in rows:
        suppressed.setdefault(row["ticker"].upper(), set()).add(row["tag_key"])
    return suppressed


def _set_holding_tags(portfolio_id: int, ticker: str, tags: list[str]) -> list[str]:
    ticker_key = ticker.strip().upper()
    desired: dict[str, str] = {}
    for tag in tags:
        cleaned, tag_key = _normalize_tag_name(tag)
        if not cleaned:
            continue
        desired[tag_key] = cleaned
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        existing = conn.execute(
            """
            SELECT tag_name, tag_key, source
            FROM holding_tags
            WHERE portfolio_id = ? AND ticker = ?
            """,
            (portfolio_id, ticker_key),
        ).fetchall()
        existing_keys = {row["tag_key"] for row in existing}
        to_remove = [row for row in existing if row["tag_key"] not in desired]
        if to_remove:
            conn.executemany(
                """
                DELETE FROM holding_tags
                WHERE portfolio_id = ? AND ticker = ? AND tag_key = ?
                """,
                [(portfolio_id, ticker_key, row["tag_key"]) for row in to_remove],
            )
            suppressed = [
                (portfolio_id, ticker_key, row["tag_key"], now)
                for row in to_remove
                if row["source"] == "auto"
            ]
            if suppressed:
                conn.executemany(
                    """
                    INSERT INTO holding_tag_suppressed (portfolio_id, ticker, tag_key, removed_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(portfolio_id, ticker, tag_key)
                    DO UPDATE SET removed_at = excluded.removed_at
                    """,
                    suppressed,
                )
        new_entries = [
            (portfolio_id, ticker_key, name, key, "manual", now, now)
            for key, name in desired.items()
            if key not in existing_keys
        ]
        if new_entries:
            conn.executemany(
                """
                INSERT INTO holding_tags (
                    portfolio_id,
                    ticker,
                    tag_name,
                    tag_key,
                    source,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                new_entries,
            )
        if desired:
            conn.execute(
                """
                DELETE FROM holding_tag_suppressed
                WHERE portfolio_id = ? AND ticker = ? AND tag_key IN ({})
                """.format(",".join(["?"] * len(desired))),
                (portfolio_id, ticker_key, *desired.keys()),
            )
    return [desired[key] for key in sorted(desired.keys())]


def _ensure_auto_tags(
    portfolio_id: int,
    ticker: str,
    auto_tags: list[str],
    current_tags: list[str],
    suppressed: set[str],
) -> list[str]:
    if not auto_tags:
        return sorted(current_tags, key=str.lower)
    now = datetime.utcnow().isoformat()
    ticker_key = ticker.strip().upper()
    tag_map = {_normalize_text(tag): tag for tag in current_tags}
    to_insert = []
    for tag in auto_tags:
        cleaned, tag_key = _normalize_tag_name(tag)
        if not cleaned or tag_key in tag_map or tag_key in suppressed:
            continue
        to_insert.append((portfolio_id, ticker_key, cleaned, tag_key, "auto", now, now))
        tag_map[tag_key] = cleaned
    if to_insert:
        with _db_connection() as conn:
            conn.executemany(
                """
                INSERT INTO holding_tags (
                    portfolio_id,
                    ticker,
                    tag_name,
                    tag_key,
                    source,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(portfolio_id, ticker, tag_key)
                DO NOTHING
                """,
                to_insert,
            )
        logger.info("Auto-tagged %s with %s", ticker_key, [row[2] for row in to_insert])
    return [tag_map[key] for key in sorted(tag_map.keys())]


def _auto_tags_from_entry(entry: dict) -> list[str]:
    asset_type = _normalize_text(entry.get("asset_type"))
    name = _normalize_text(entry.get("name"))
    ticker = _normalize_text(entry.get("ticker"))
    ticker_key = str(entry.get("ticker") or "").strip().upper()
    tags: list[str] = []
    if ticker_key in KNOWN_ETF_TICKERS:
        tags.append("ETF")
    if ticker_key in KNOWN_REIT_TICKERS:
        tags.append("REITs")
    if "etf" in asset_type or "etf" in name or ticker.endswith("etf"):
        tags.append("ETF")
    if "reit" in asset_type or "reit" in name or "reit" in ticker:
        tags.append("REITs")
    return tags


def _normalize_categories(categories: list[str] | None) -> list[str]:
    if not categories:
        return []
    seen = set()
    result: list[str] = []
    for entry in categories:
        if not entry:
            continue
        label = str(entry).strip()
        if not label:
            continue
        key = _normalize_text(label)
        if key in seen:
            continue
        seen.add(key)
        result.append(label)
    return result


def _filter_categories(categories: list[str] | None) -> list[str]:
    base = list(DEFAULT_CATEGORIES)
    custom = _normalize_categories(categories)
    for entry in custom:
        if _normalize_text(entry) in {_normalize_text(item) for item in base}:
            continue
        base.append(entry)
    return base


def _parse_number(value: str | float | int | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    text = text.replace("\u20ac", "").replace("$", "")
    text = text.replace("EUR", "").replace("USD", "").replace("GBP", "")
    text = text.replace("%", "")
    text = text.replace("\u00a0", " ").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if text.count(",") > 0 and text.count(".") > 0:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") > 0:
        text = text.replace(",", ".")
    if text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        parsed = float(text)
    except ValueError:
        return None
    return -parsed if negative else parsed


def _normalize_tickers(tickers: list[str] | None) -> list[str]:
    if not tickers:
        return []
    seen: set[str] = set()
    normalized: list[str] = []
    for ticker in tickers:
        value = (ticker or "").strip().upper()
        if not value or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def _fix_mojibake(value: str | None) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        return str(value)
    if "Ã" in value or "Â" in value:
        try:
            return value.encode("latin-1").decode("utf-8")
        except UnicodeError:
            return value
    return value


def _parse_transaction_date(value: str | datetime | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    if " " in text:
        try:
            return datetime.fromisoformat(text).date().isoformat()
        except ValueError:
            pass
    return _date_key(text)


def _detect_delimiter(sample: str) -> str:
    for delimiter in (";", ",", "\t", "|"):
        if delimiter in sample:
            return delimiter
    return ","


def _load_rows_from_text(text: str) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    sample = text.splitlines()[0] if text.splitlines() else ""
    delimiter = _detect_delimiter(sample)
    for row in csv.reader(text.splitlines(), delimiter=delimiter):
        rows.append([_fix_mojibake(cell) for cell in row])
    return rows


def _load_rows_from_excel(file_bytes: bytes, filename: str) -> list[list[str | float | int | None]]:
    rows: list[list[str | float | int | None]] = []
    if filename.lower().endswith(".xls"):
        book = xlrd.open_workbook(file_contents=file_bytes)
        sheet = book.sheet_by_index(0)
        for row_idx in range(sheet.nrows):
            rows.append([_fix_mojibake(cell) for cell in sheet.row_values(row_idx)])
        return rows
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        rows.append([_fix_mojibake(cell) for cell in row])
    return rows


def _split_csv_column_rows(
    rows: list[list[str | float | int | None]],
) -> list[list[str | float | int | None]]:
    if not rows:
        return rows
    if not all(len(row) == 1 and isinstance(row[0], str) for row in rows):
        return rows
    if not rows[0] or "," not in str(rows[0][0]):
        return rows
    parsed: list[list[str | None]] = []
    for row in rows:
        value = str(row[0]) if row and row[0] is not None else ""
        for csv_row in csv.reader([value], delimiter=_detect_delimiter(value)):
            parsed.append([_fix_mojibake(cell) for cell in csv_row])
    return parsed


def _find_header_row(rows: list[list[str | float | int | None]]) -> int | None:
    header_keywords = {
        "date": {
            "data",
            "date",
            "data operacao",
            "data lanc",
            "data lanc.",
            "data valor",
            "data de inicio",
            "data de conclusao",
        },
        "description": {"descricao", "descr", "description"},
        "amount": {"montante", "valor", "amount"},
        "balance": {"saldo", "balance"},
        "currency": {"moeda", "currency"},
        "debit": {"debito"},
        "credit": {"credito"},
    }
    for idx, row in enumerate(rows[:40]):
        hits = 0
        for cell in row:
            if not cell:
                continue
            key = _normalize_text(str(cell))
            for words in header_keywords.values():
                if any(word in key for word in words):
                    hits += 1
                    break
        if hits >= 2:
            return idx
    return None


def _suggest_banking_mapping(columns: list[str]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {
        "date": None,
        "description": None,
        "amount": None,
        "balance": None,
        "currency": None,
        "debit": None,
        "credit": None,
    }
    for idx, label in enumerate(columns):
        key = _normalize_text(label)
        if mapping["date"] is None and "data" in key:
            mapping["date"] = idx
            continue
        if mapping["description"] is None and ("descr" in key or "description" in key):
            mapping["description"] = idx
            continue
        if mapping["balance"] is None and "saldo" in key:
            mapping["balance"] = idx
            continue
        if mapping["currency"] is None and ("moeda" in key or "curr" in key):
            mapping["currency"] = idx
            continue
        if mapping["debit"] is None and "deb" in key:
            mapping["debit"] = idx
            continue
        if mapping["credit"] is None and "cred" in key:
            mapping["credit"] = idx
            continue
        if "montante" in key or "amount" in key:
            mapping["amount"] = idx
            continue
        if mapping["amount"] is None and "valor" in key and "data" not in key:
            mapping["amount"] = idx
    return mapping


def _normalize_mapping(mapping: dict[str, int | None]) -> dict[str, int | None]:
    normalized: dict[str, int | None] = {}
    for key, value in mapping.items():
        normalized[key] = value if isinstance(value, int) else None
    return normalized


def _build_banking_items(
    rows: list[list[str | float | int | None]],
    columns: list[str],
    mapping: dict[str, int | None],
    currency_fallback: str,
) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    items: list[dict] = []
    mapping = _normalize_mapping(mapping)
    for idx, row in enumerate(rows):
        cells = [cell if cell is not None else "" for cell in row]

        def cell_at(key: str) -> str | float | int | None:
            col_idx = mapping.get(key)
            if col_idx is None or col_idx >= len(cells):
                return None
            return cells[col_idx]

        date_value = _parse_transaction_date(cell_at("date"))
        description_value = _normalize_optional_text(_fix_mojibake(cell_at("description")))
        debit_value = _parse_number(cell_at("debit")) if mapping.get("debit") is not None else None
        credit_value = (
            _parse_number(cell_at("credit")) if mapping.get("credit") is not None else None
        )
        if debit_value is not None or credit_value is not None:
            debit_value = debit_value or 0.0
            credit_value = credit_value or 0.0
            amount_value = credit_value - abs(debit_value)
        else:
            amount_value = _parse_number(cell_at("amount"))
        balance_value = _parse_number(cell_at("balance"))
        currency_value = (
            _normalize_optional_text(_fix_mojibake(cell_at("currency")))
            or currency_fallback
        )
        if not date_value or not description_value or amount_value is None:
            warnings.append(f"Row {idx + 1} skipped: missing required values.")
            continue
        items.append(
            {
                "tx_date": date_value,
                "description": description_value,
                "amount": float(amount_value),
                "balance": float(balance_value) if balance_value is not None else None,
                "currency": currency_value,
                "category": BANKING_DEFAULT_CATEGORY,
                "subcategory": BANKING_DEFAULT_SUBCATEGORY,
                "raw": dict(zip(columns, cells)),
            }
        )
    return items, warnings


def _ensure_banking_categories(portfolio_id: int) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        existing = conn.execute(
            "SELECT COUNT(*) AS total FROM banking_categories WHERE portfolio_id = ?",
            (portfolio_id,),
        ).fetchone()
        if existing and existing["total"] > 0:
            return
        for parent_name, children in BANKING_CATEGORY_TREE.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO banking_categories (portfolio_id, parent_id, name, is_default, created_at)
                VALUES (?, NULL, ?, 1, ?)
                """,
                (portfolio_id, parent_name, now),
            )
            parent_id = conn.execute(
                """
                SELECT id FROM banking_categories
                WHERE portfolio_id = ? AND parent_id IS NULL AND name = ?
                """,
                (portfolio_id, parent_name),
            ).fetchone()
            if not parent_id:
                continue
            for child_name in children:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO banking_categories (portfolio_id, parent_id, name, is_default, created_at)
                    VALUES (?, ?, ?, 1, ?)
                    """,
                    (portfolio_id, parent_id["id"], child_name, now),
                )


def _list_banking_categories(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, parent_id, name
            FROM banking_categories
            WHERE portfolio_id = ?
            ORDER BY parent_id NULLS FIRST, name
            """,
            (portfolio_id,),
        ).fetchall()
    by_parent: dict[int | None, list[dict]] = {}
    for row in rows:
        by_parent.setdefault(row["parent_id"], []).append(
            {"id": row["id"], "name": row["name"]}
        )
    result: list[dict] = []
    for parent in by_parent.get(None, []):
        children = by_parent.get(parent["id"], [])
        result.append(
            {
                "name": parent["name"],
                "subcategories": [child["name"] for child in children],
            }
        )
    return result


def _normalize_banking_description(value: str | None) -> str | None:
    if not value:
        return None
    return _normalize_text(value)


def _upsert_banking_category(
    portfolio_id: int, category: str | None, subcategory: str | None = None
) -> None:
    category_name = (category or "").strip()
    if not category_name:
        return
    subcategory_name = (subcategory or "").strip()
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        parent_row = conn.execute(
            """
            SELECT id
            FROM banking_categories
            WHERE portfolio_id = ? AND parent_id IS NULL AND lower(name) = ?
            """,
            (portfolio_id, _normalize_text(category_name)),
        ).fetchone()
        if parent_row:
            parent_id = parent_row["id"]
        else:
            conn.execute(
                """
                INSERT OR IGNORE INTO banking_categories (portfolio_id, parent_id, name, is_default, created_at)
                VALUES (?, NULL, ?, 0, ?)
                """,
                (portfolio_id, category_name, now),
            )
            parent_id = conn.execute(
                """
                SELECT id
                FROM banking_categories
                WHERE portfolio_id = ? AND parent_id IS NULL AND lower(name) = ?
                """,
                (portfolio_id, _normalize_text(category_name)),
            ).fetchone()["id"]
        if subcategory_name:
            conn.execute(
                """
                INSERT OR IGNORE INTO banking_categories (portfolio_id, parent_id, name, is_default, created_at)
                VALUES (?, ?, ?, 0, ?)
                """,
                (portfolio_id, parent_id, subcategory_name, now),
            )


def _find_banking_rule(
    portfolio_id: int, institution: str, description: str | None
) -> dict | None:
    normalized_desc = _normalize_banking_description(description)
    if not normalized_desc:
        return None
    normalized_inst = _normalize_text(institution or "")
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT category, subcategory
            FROM banking_category_rules
            WHERE portfolio_id = ? AND institution = ? AND match_type = 'exact' AND match_value = ?
            """,
            (portfolio_id, normalized_inst, normalized_desc),
        ).fetchone()
    if not row:
        return None
    return {"category": row["category"], "subcategory": row["subcategory"]}


def _learn_banking_rule(
    portfolio_id: int,
    institution: str,
    description: str | None,
    category: str | None,
    subcategory: str | None,
) -> None:
    normalized_desc = _normalize_banking_description(description)
    if not normalized_desc:
        return
    normalized_inst = _normalize_text(institution or "")
    category_name = (category or BANKING_DEFAULT_CATEGORY).strip()
    subcategory_name = (subcategory or BANKING_DEFAULT_SUBCATEGORY).strip()
    _upsert_banking_category(portfolio_id, category_name, subcategory_name)
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO banking_category_rules (
                portfolio_id,
                institution,
                match_type,
                match_value,
                category,
                subcategory,
                updated_at
            )
            VALUES (?, ?, 'exact', ?, ?, ?, ?)
            ON CONFLICT(portfolio_id, institution, match_type, match_value)
            DO UPDATE SET
                category = excluded.category,
                subcategory = excluded.subcategory,
                updated_at = excluded.updated_at
            """,
            (
                portfolio_id,
                normalized_inst,
                normalized_desc,
                category_name,
                subcategory_name,
                now,
            ),
        )


def _suggest_banking_category(description: str | None) -> tuple[str, str] | None:
    normalized_desc = _normalize_banking_description(description)
    if not normalized_desc:
        return None
    for category, subcategory, keywords in BANKING_CATEGORY_KEYWORDS:
        for keyword in keywords:
            if keyword in normalized_desc:
                return category, subcategory
    return None


def _apply_banking_category_rules(
    portfolio_id: int, institution: str, items: list[dict]
) -> None:
    for item in items:
        category = item.get("category") or BANKING_DEFAULT_CATEGORY
        subcategory = item.get("subcategory") or BANKING_DEFAULT_SUBCATEGORY
        if _normalize_text(category) == _normalize_text(BANKING_DEFAULT_CATEGORY):
            rule = _find_banking_rule(portfolio_id, institution, item.get("description"))
            if rule:
                category = rule["category"]
                subcategory = rule.get("subcategory") or BANKING_DEFAULT_SUBCATEGORY
                item["category"] = category
                item["subcategory"] = subcategory
            else:
                suggestion = _suggest_banking_category(item.get("description"))
                if suggestion:
                    category, subcategory = suggestion
                    item["category"] = category
                    item["subcategory"] = subcategory
                else:
                    item["category"] = category
                    item["subcategory"] = subcategory
        else:
            item["subcategory"] = subcategory
        _upsert_banking_category(portfolio_id, category, subcategory)


def _upsert_banking_institution(portfolio_id: int, name: str) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO banking_institutions (portfolio_id, name, created_at)
            VALUES (?, ?, ?)
            """,
            (portfolio_id, name, now),
        )


def _list_banking_institutions(portfolio_id: int) -> list[str]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT name
            FROM banking_institutions
            WHERE portfolio_id = ?
            ORDER BY name
            """,
            (portfolio_id,),
        ).fetchall()
    return [row["name"] for row in rows]


def _load_banking_rows(
    file_bytes: bytes | None, filename: str | None, text: str | None
) -> list[list[str | float | int | None]]:
    if text:
        return _load_rows_from_text(text)
    if not file_bytes or not filename:
        return []
    if file_bytes[:2] == b"PK":
        rows = _load_rows_from_excel(file_bytes, f"{filename}.xlsx")
    elif filename.lower().endswith((".xls", ".xlsx")):
        rows = _load_rows_from_excel(file_bytes, filename)
    else:
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                decoded = file_bytes.decode(encoding)
                rows = _load_rows_from_text(decoded)
                break
            except UnicodeError:
                rows = []
        else:
            rows = []
    rows = _split_csv_column_rows(rows)
    return rows


def _trim_empty_rows(rows: list[list[str | float | int | None]]) -> list[list]:
    trimmed: list[list] = []
    for row in rows:
        if any(str(cell).strip() if cell is not None else "" for cell in row):
            trimmed.append(row)
    return trimmed


def _build_banking_preview(
    rows: list[list[str | float | int | None]],
    currency_fallback: str,
) -> tuple[list[str], list[dict], dict[str, int | None], list[str]]:
    rows = _trim_empty_rows(rows)
    header_row = _find_header_row(rows)
    if header_row is None:
        return [], [], _suggest_banking_mapping([]), ["No header row detected."]
    columns = [
        _normalize_optional_text(_fix_mojibake(cell)) or ""
        for cell in rows[header_row]
    ]
    data_rows = rows[header_row + 1 :]
    mapping = _suggest_banking_mapping(columns)
    items, warnings = _build_banking_items(data_rows, columns, mapping, currency_fallback)
    preview_rows: list[dict] = []
    for row in data_rows:
        preview_rows.append(
            {
                "cells": [cell if cell is not None else "" for cell in row],
                "include": True,
            }
        )
    return columns, preview_rows, mapping, warnings


def _map_santander_category(account: str) -> str:
    mapping = {
        "conta ordenado": "Cash",
        "backup imi seguros": "Cash",
        "aforro": "Emergency Funds",
        "benedita": "Beni",
        "poupanca 2": "Beni",
        "margarida": "Magui",
        "000360966967020": "Magui",
        "000360973013020": "Magui",
        "ferias": "Cash",
        "backupgeral": "Cash",
        "ppr": "Retirement Plans",
        "xxxxxxxxxxxx6149": "Cash",
        "xxxxxxxxxxxx3232": "Cash",
        "ppr+ equilibrado": "Retirement Plans",
    }
    key = _normalize_text(account)
    return mapping.get(key, "Unknown")



def _iter_santander_rows(file_bytes: bytes, filename: str) -> list[list[object]]:
    if filename.lower().endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        return [
            sheet.row_values(row_index, start_colx=0, end_colx=4)
            for row_index in range(sheet.nrows)
        ]
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        return [
            [cell.value for cell in row]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=4)
        ]
    finally:
        workbook.close()



def _parse_santander_sheet(
    file_bytes: bytes, filename: str
) -> tuple[list[SantanderItem], list[str]]:
    rows = _iter_santander_rows(file_bytes, filename)
    items: list[SantanderItem] = []
    warnings: list[str] = []
    current_section = ""

    section_map = {
        "contas": "Contas",
        "poupancas": "Poupancas",
        "cartoes de debito": "Cartoes de debito",
        "cartoes pre-pagos / refeicao": "Cartoes pre-pagos / Refeicao",
        "cartoes pre-pagos/ refeicao": "Cartoes pre-pagos / Refeicao",
        "planos poupanca reforma": "Planos poupanca reforma",
    }

    for row in rows:
        cell_a = row[0] if len(row) > 0 else None
        if isinstance(cell_a, str):
            normalized = _normalize_text(cell_a)
            if normalized in section_map:
                current_section = section_map[normalized]
                continue

        if not current_section:
            continue

        account = str(cell_a).strip() if cell_a else ""
        if not account:
            continue

        normalized_account = _normalize_text(account)
        if normalized_account in {"conta", "nome", "cartao", "total"}:
            continue

        description = row[1] if len(row) > 1 else None
        description_text = str(description).strip() if description else None

        value_c = _parse_number(row[2] if len(row) > 2 else None)
        value_d = _parse_number(row[3] if len(row) > 3 else None)

        if current_section == "Contas":
            balance = value_c
        elif current_section == "Poupancas":
            balance = value_c
        elif current_section == "Cartoes de debito":
            balance = value_d if value_d is not None else value_c
        elif current_section == "Cartoes pre-pagos / Refeicao":
            balance = value_d if value_d is not None else value_c
        elif current_section == "Planos poupanca reforma":
            balance = value_d if value_d is not None else value_c
        else:
            balance = value_c

        if balance is None:
            continue

        category = _map_santander_category(account)
        items.append(
            SantanderItem(
                section=current_section,
                account=account,
                description=description_text,
                balance=balance,
                category=category,
            )
        )

    if not items:
        warnings.append("No Santander rows detected.")

    return items, warnings


def _account_key(section: str, account: str) -> str:
    return f"{_normalize_text(section)}::{_normalize_text(account)}"


def _load_category_map(portfolio_id: int) -> dict[str, dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT account_key, category, ignore
            FROM santander_category_map
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
    return {
        row["account_key"]: {
            "category": row["category"],
            "ignore": bool(row["ignore"]),
        }
        for row in rows
    }


def _upsert_category_map(portfolio_id: int, items: list[SantanderItem]) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        for item in items:
            conn.execute(
                """
                INSERT INTO santander_category_map (
                    portfolio_id,
                    account_key,
                    category,
                    ignore,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(portfolio_id, account_key)
                DO UPDATE SET category = excluded.category,
                              ignore = excluded.ignore,
                              updated_at = excluded.updated_at
                """,
                (
                    portfolio_id,
                    _account_key(item.section, item.account),
                    item.category,
                    1 if item.ignore else 0,
                    now,
                ),
            )


def _delete_category_map_by_category(portfolio_id: int, category: str) -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            DELETE FROM santander_category_map
            WHERE portfolio_id = ? AND lower(category) = lower(?)
            """,
            (portfolio_id, category),
        )


def _count_santander_category(portfolio_id: int, category: str) -> int:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT COUNT(1) AS total
            FROM santander_items
            WHERE import_id IN (
                SELECT id FROM santander_imports WHERE portfolio_id = ?
            )
            AND lower(category) = lower(?)
            """,
            (portfolio_id, category),
        ).fetchone()
    return int(row["total"] or 0)


def _delete_santander_category(portfolio_id: int, category: str) -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            DELETE FROM santander_items
            WHERE import_id IN (
                SELECT id FROM santander_imports WHERE portfolio_id = ?
            )
            AND lower(category) = lower(?)
            """,
            (portfolio_id, category),
        )


def _update_portfolio_categories(portfolio_id: int, categories: list[str]) -> None:
    categories = _filter_categories(categories)
    with _db_connection() as conn:
        conn.execute(
            "UPDATE portfolios SET categories_json = ?, updated_at = ? WHERE id = ?",
            (json.dumps(categories), datetime.utcnow().isoformat(), portfolio_id),
        )
    _ensure_category_settings(portfolio_id, categories)


def _default_is_investment(category: str) -> bool:
    return _normalize_text(category) != "cash"


def _ensure_category_settings(portfolio_id: int, categories: list[str]) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        for category in categories:
            conn.execute(
                """
                INSERT OR IGNORE INTO portfolio_category_settings (
                    portfolio_id,
                    category,
                    is_investment,
                    updated_at
                )
                VALUES (?, ?, ?, ?)
                """,
                (portfolio_id, category, 1 if _default_is_investment(category) else 0, now),
            )


def _get_category_settings(portfolio_id: int) -> dict[str, bool]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT category, is_investment
            FROM portfolio_category_settings
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
    return {row["category"]: bool(row["is_investment"]) for row in rows}


def _set_category_setting(portfolio_id: int, category: str, is_investment: bool) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO portfolio_category_settings (
                portfolio_id,
                category,
                is_investment,
                updated_at
            )
            VALUES (?, ?, ?, ?)
            ON CONFLICT(portfolio_id, category)
            DO UPDATE SET is_investment = excluded.is_investment,
                          updated_at = excluded.updated_at
            """,
            (portfolio_id, category, 1 if is_investment else 0, now),
        )


def _delete_category_setting(portfolio_id: int, category: str) -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            DELETE FROM portfolio_category_settings
            WHERE portfolio_id = ? AND lower(category) = lower(?)
            """,
            (portfolio_id, category),
        )


def _save_santander_import(portfolio_id: int, filename: str, items: list[SantanderItem]) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO santander_imports (portfolio_id, filename, imported_at)
            VALUES (?, ?, ?)
            """,
            (portfolio_id, filename, now),
        )
        import_id = cursor.lastrowid
        for item in items:
            conn.execute(
                """
                INSERT INTO santander_items (
                    import_id, section, account, description, balance, category, invested, gains
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    item.section,
                    item.account,
                    item.description,
                    item.balance,
                    item.category,
                    item.invested,
                    item.gains,
                ),
            )
    return {"import_id": import_id, "imported_at": now}


def _save_trade_republic_entry(portfolio_id: int, entry: dict) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO trade_republic_entries (
                portfolio_id,
                available_cash,
                interests_received,
                invested,
                value,
                gains,
                currency,
                category,
                source,
                source_file,
                file_hash,
                snapshot_date,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                portfolio_id,
                entry["available_cash"],
                entry["interests_received"],
                entry["invested"],
                entry["value"],
                entry["gains"],
                entry["currency"],
                entry.get("category"),
                entry["source"],
                entry.get("source_file"),
                entry.get("file_hash"),
                entry.get("snapshot_date"),
                now,
            ),
        )
        entry_id = cursor.lastrowid
    return {"id": entry_id, "created_at": now}


def _list_trade_republic_entries(portfolio_id: int, source: str | None = None) -> list[dict]:
    with _db_connection() as conn:
        if source:
            rows = conn.execute(
                """
                SELECT id,
                       available_cash,
                       interests_received,
                       invested,
                       value,
                       gains,
                       currency,
                       category,
                       source,
                       source_file,
                       file_hash,
                       snapshot_date,
                       created_at
                FROM trade_republic_entries
                WHERE portfolio_id = ? AND source = ?
                ORDER BY created_at DESC
                """,
                (portfolio_id, source),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id,
                       available_cash,
                       interests_received,
                       invested,
                       value,
                       gains,
                       currency,
                       category,
                       source,
                       source_file,
                       file_hash,
                       snapshot_date,
                       created_at
                FROM trade_republic_entries
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                """,
                (portfolio_id,),
            ).fetchall()
    return [
        {
            "id": row["id"],
            "available_cash": float(row["available_cash"] or 0),
            "interests_received": float(row["interests_received"] or 0),
            "invested": float(row["invested"] or 0),
            "value": float(row["value"] or 0),
            "gains": float(row["gains"] or 0),
            "currency": row["currency"],
            "category": row["category"],
            "source": row["source"],
            "source_file": row["source_file"],
            "file_hash": row["file_hash"],
            "snapshot_date": row["snapshot_date"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def _latest_trade_republic_category(portfolio_id: int) -> str | None:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT category
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
    return row["category"] if row and row["category"] else None


def _trade_republic_file_exists(portfolio_id: int, file_hash: str) -> bool:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT 1
            FROM trade_republic_entries
            WHERE portfolio_id = ? AND file_hash = ?
            LIMIT 1
            """,
            (portfolio_id, file_hash),
        ).fetchone()
    return row is not None


def _build_trade_republic_entry(
    available_cash_value: str | float,
    interests_received_value: str | float,
    currency: str,
    category: str = "Cash",
    source: str = "manual",
    source_file: str | None = None,
    file_hash: str | None = None,
    snapshot_date: str | None = None,
) -> dict:
    available_cash = _parse_number(available_cash_value)
    interests_received = _parse_number(interests_received_value)
    if available_cash is None or interests_received is None:
        raise HTTPException(
            status_code=400,
            detail="Trade Republic requires Available Cash and Interests received.",
        )
    invested = float(available_cash) - float(interests_received)
    if invested < 0:
        raise HTTPException(status_code=400, detail="Juros nao podem exceder o saldo.")
    return {
        "available_cash": float(available_cash),
        "interests_received": float(interests_received),
        "invested": invested,
        "value": float(available_cash),
        "gains": float(interests_received),
        "currency": currency,
        "category": category,
        "source": source,
        "source_file": source_file,
        "file_hash": file_hash,
        "snapshot_date": snapshot_date,
    }


def _parse_percent(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        percent = float(value)
        return percent * 100 if 0 <= percent <= 1 else percent
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if "%" in stripped:
            matches = re.findall(r"[-+]?\d+(?:[.,]\d+)?", stripped)
            if matches:
                percent = _parse_number(matches[-1])
            else:
                percent = None
        else:
            percent = _parse_number(stripped)
        if percent is None:
            return None
        if "%" in stripped and 0 <= percent <= 1:
            return percent * 100
        return percent * 100 if 0 <= percent <= 1 else percent
    return None


def _parse_profit_cell(value: object) -> tuple[float | None, float | None]:
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None
    text = str(value).strip()
    if not text:
        return None, None
    matches = re.findall(r"[-+]?\d+(?:[.,]\d+)?", text)
    if not matches:
        return None, None
    if "%" in text:
        profit_value = _parse_number(matches[0])
        profit_percent = _parse_percent(f"{matches[-1]}%")
        return profit_value, profit_percent
    return _parse_number(matches[0]), None


def _normalize_profit_value(
    current_value: float | None,
    invested: float | None,
    profit_value: float | None,
) -> float | None:
    if profit_value is None:
        if invested is None or current_value is None:
            return None
        return float(current_value) - float(invested)
    if current_value is None or invested is None:
        return profit_value
    if abs(float(profit_value)) > abs(float(current_value)) * 10:
        return float(current_value) - float(invested)
    return profit_value


def _parse_date_value(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(stripped, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _read_savengrow_cells(file_bytes: bytes, filename: str) -> dict:
    if filename.lower().endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        rows = [
            sheet.row_values(row_index, start_colx=0, end_colx=8)
            for row_index in range(sheet.nrows)
        ]
    else:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows = [
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=8)
            ]
        finally:
            workbook.close()

    total_row_idx = None
    for idx, row in enumerate(rows):
        cell_a = row[0] if len(row) > 0 else None
        if isinstance(cell_a, str) and _normalize_text(cell_a) == "total":
            total_row_idx = idx
            break

    if total_row_idx is None:
        raise HTTPException(status_code=400, detail="Save N Grow missing TOTAL row.")

    items: list[dict] = []
    snapshot_date = None
    for row in rows[total_row_idx + 1 :]:
        name_raw = row[0] if len(row) > 0 else None
        name = str(name_raw).strip() if name_raw else ""
        if not name:
            if not any(cell not in (None, "") for cell in row):
                break
            continue
        invested = _parse_number(row[3] if len(row) > 3 else None)
        current_value = _parse_number(row[4] if len(row) > 4 else None)
        if current_value is None:
            continue
        raw_profit = row[5] if len(row) > 5 else None
        profit_value, profit_percent = _parse_profit_cell(raw_profit)
        if profit_percent is None:
            profit_percent = _parse_percent(row[6] if len(row) > 6 else None)
            if profit_percent is None:
                profit_percent = _parse_percent(raw_profit)
        if profit_value is None and invested is not None:
            profit_value = current_value - invested
        if profit_percent is None and profit_value is not None and current_value:
            profit_percent = (profit_value / current_value) * 100
        date_raw = row[7] if len(row) > 7 else None
        parsed_date = _parse_date_value(date_raw)
        if snapshot_date is None and parsed_date:
            snapshot_date = parsed_date
        items.append(
            {
                "name": name,
                "invested": invested,
                "current_value": current_value,
                "profit_value": profit_value,
                "profit_percent": profit_percent,
                "category": "Retirement Plans",
            }
        )

    if not items:
        raise HTTPException(status_code=400, detail="Save N Grow has no rows after TOTAL.")

    return {"items": items, "snapshot_date": snapshot_date}


def _extract_date_from_filename(name: str) -> str | None:
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", name)
    if not match:
        return None
    day, month, year = (int(value) for value in match.groups())
    try:
        return datetime(year, month, day).date().isoformat()
    except ValueError:
        return None


def _parse_aforronet_units(value: str) -> float | None:
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", text):
        return float(text.replace(".", ""))
    return _parse_number(text)


def _parse_aforronet_pdf(file_bytes: bytes, filename: str) -> dict:
    total_units = 0.0
    total_value = 0.0
    rows_found = 0
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                if not re.match(r"\d{2}-\d{2}-\d{4}", line):
                    continue
                parts = re.split(r"\s+", line)
                if len(parts) < 5:
                    continue
                units = _parse_aforronet_units(parts[-2])
                value = _parse_number(parts[-1])
                if units is None or value is None:
                    continue
                total_units += units
                total_value += value
                rows_found += 1
    if rows_found == 0:
        raise HTTPException(
            status_code=400,
            detail="AforroNet PDF structure not recognized. Check the monthly statement.",
        )
    return {
        "invested": total_units,
        "current_value": total_value,
        "snapshot_date": _extract_date_from_filename(filename),
    }


def _extract_currency_values(text: str) -> list[float]:
    values: list[float] = []
    for token in re.findall(r"€\s*[-+]?\d[\d.,]*", text):
        parsed = _parse_number(token)
        if parsed is not None:
            values.append(parsed)
    return values


def _parse_trade_republic_pdf(file_bytes: bytes, filename: str) -> dict:
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    available_cash = None
    for line in lines:
        if "Checking Account" in line:
            values = _extract_currency_values(line)
            if values:
                available_cash = values[-1]
                break
    if available_cash is None:
        last_value = None
        for line in lines:
            values = _extract_currency_values(line)
            if values:
                last_value = values[-1]
        available_cash = last_value
    if available_cash is None:
        raise HTTPException(
            status_code=400,
            detail="Trade Republic PDF structure not recognized (missing balance).",
        )
    interests_total = 0.0
    for line in lines:
        if re.search(r"\bInterest\b", line, re.IGNORECASE):
            if "tax" in line.lower():
                continue
            values = _extract_currency_values(line)
            if values:
                interests_total += values[0]
    return {
        "available_cash": available_cash,
        "interests_received": interests_total,
        "snapshot_date": _extract_date_from_filename(filename),
    }


def _read_bancoinvest_cells(file_bytes: bytes, filename: str) -> dict:
    if filename.lower().endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        rows = [
            sheet.row_values(row_index, start_colx=0, end_colx=12)
            for row_index in range(sheet.nrows)
        ]
    else:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows = [
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=12)
            ]
        finally:
            workbook.close()

    header_idx = None
    holder_idx = None
    var_idx = None
    value_idx = None
    for idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            label = _normalize_text(str(cell))
            if label == "titular":
                header_idx = idx
            if header_idx is not None:
                if label == "titular":
                    holder_idx = col_idx
                elif label in {"var moeda", "varmoeda"}:
                    var_idx = col_idx
                elif label == "valor":
                    value_idx = col_idx
        if header_idx is not None:
            break

    if header_idx is None or holder_idx is None or value_idx is None:
        raise HTTPException(status_code=400, detail="BancoInvest header not found.")

    snapshot_date = None
    for row in rows[:10]:
        for cell in row:
            parsed = _parse_date_value(cell)
            if parsed:
                snapshot_date = parsed
                break
        if snapshot_date:
            break

    items: list[dict] = []
    for row in rows[header_idx + 1 :]:
        holder_raw = row[holder_idx] if len(row) > holder_idx else None
        holder = str(holder_raw).strip() if holder_raw else ""
        if not holder:
            if not any(cell not in (None, "") for cell in row):
                break
            continue
        current_value = _parse_number(row[value_idx] if len(row) > value_idx else None)
        if current_value is None:
            continue
        var_value = (
            _parse_number(row[var_idx] if var_idx is not None and len(row) > var_idx else None)
            if var_idx is not None
            else None
        )
        invested = current_value - var_value if var_value is not None else None
        gains = var_value
        items.append(
            {
                "holder": holder,
                "invested": invested,
                "current_value": current_value,
                "gains": gains,
                "category": "Retirement Plans",
            }
        )

    if not items:
        raise HTTPException(status_code=400, detail="BancoInvest has no rows found.")

    return {"items": items, "snapshot_date": snapshot_date}


def _xtb_account_type(filename: str) -> str:
    lowered = filename.lower()
    mapping = {
        "account_50721856": "Brokers",
        "account_50727616": "Beni",
        "account_50727618": "Magui",
    }
    for prefix, label in mapping.items():
        if lowered.startswith(prefix):
            return label
    raise HTTPException(
        status_code=400,
        detail=f"Unrecognized XTB file name: {filename}.",
    )


def _xtb_sheet_by_name_xlsx(workbook, name: str):
    target = _normalize_text(name)
    for sheet in workbook.worksheets:
        if _normalize_text(sheet.title) == target:
            return sheet
    return None


def _xtb_sheet_by_name_xls(workbook, name: str):
    target = _normalize_text(name)
    for sheet_name in workbook.sheet_names():
        if _normalize_text(sheet_name) == target:
            return workbook.sheet_by_name(sheet_name)
    return None


def _xtb_operation_kind(value: str | None) -> str:
    text = _normalize_text(value or "")
    if "stock purchase" in text:
        return "buy"
    if "stock sale" in text:
        return "sell"
    if "dividend" in text:
        return "dividend"
    if "interest" in text:
        return "interest"
    if "tax" in text:
        return "tax"
    if "fee" in text or "commission" in text:
        return "fee"
    return "other"


def _parse_operation_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        raw = value.strip()
        if " " in raw:
            raw = raw.split(" ")[0]
        return _parse_date_value(raw)
    return _parse_date_value(value)


def _xtb_cash_operations_from_xls(sheet) -> list[dict]:
    header_row = None
    header_map: dict[str, int] = {}
    max_scan = min(sheet.nrows, 60)
    for row_idx in range(max_scan):
        row = sheet.row_values(row_idx)
        for col_idx, value in enumerate(row):
            if not isinstance(value, str):
                continue
            key = _normalize_text(value)
            if key in {"type"}:
                header_map["type"] = col_idx
                header_row = row_idx
            if key in {"time", "date"}:
                header_map["date"] = col_idx
                header_row = row_idx
            if key in {"comment", "description"}:
                header_map["comment"] = col_idx
                header_row = row_idx
            if key in {"symbol", "ticker"}:
                header_map["ticker"] = col_idx
                header_row = row_idx
            if key in {"amount"}:
                header_map["amount"] = col_idx
                header_row = row_idx
        if header_row is not None and {"type", "amount"}.issubset(header_map):
            break
    if header_row is None or not {"type", "amount"}.issubset(header_map):
        return []
    operations: list[dict] = []
    for row_idx in range(header_row + 1, sheet.nrows):
        type_value = sheet.cell_value(row_idx, header_map["type"])
        if type_value is None or str(type_value).strip() == "":
            continue
        amount_value = sheet.cell_value(row_idx, header_map["amount"])
        amount = _parse_number(amount_value)
        if amount is None:
            continue
        date_value = (
            sheet.cell_value(row_idx, header_map["date"])
            if "date" in header_map
            else None
        )
        ticker_value = (
            sheet.cell_value(row_idx, header_map["ticker"])
            if "ticker" in header_map
            else None
        )
        comment_value = (
            sheet.cell_value(row_idx, header_map["comment"])
            if "comment" in header_map
            else None
        )
        operations.append(
            {
                "operation_type": str(type_value).strip(),
                "operation_kind": _xtb_operation_kind(str(type_value)),
                "trade_date": _parse_operation_date(date_value),
                "ticker": str(ticker_value).strip() if ticker_value else None,
                "description": str(comment_value).strip() if comment_value else None,
                "amount": float(amount),
            }
        )
    return operations


def _xtb_cash_operations_from_xlsx(sheet) -> list[dict]:
    header_row = None
    header_map: dict[str, int] = {}
    max_scan = min(sheet.max_row or 0, 60)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan):
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            key = _normalize_text(value)
            if key in {"type"}:
                header_map["type"] = cell.column - 1
                header_row = cell.row
            if key in {"time", "date"}:
                header_map["date"] = cell.column - 1
                header_row = cell.row
            if key in {"comment", "description"}:
                header_map["comment"] = cell.column - 1
                header_row = cell.row
            if key in {"symbol", "ticker"}:
                header_map["ticker"] = cell.column - 1
                header_row = cell.row
            if key in {"amount"}:
                header_map["amount"] = cell.column - 1
                header_row = cell.row
        if header_row is not None and {"type", "amount"}.issubset(header_map):
            break
    if header_row is None or not {"type", "amount"}.issubset(header_map):
        return []
    operations: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        type_cell = row[header_map["type"]] if len(row) > header_map["type"] else None
        if not type_cell or type_cell.value is None:
            continue
        amount_cell = (
            row[header_map["amount"]] if len(row) > header_map["amount"] else None
        )
        amount = _parse_number(amount_cell.value if amount_cell else None)
        if amount is None:
            continue
        date_value = (
            row[header_map["date"]].value
            if "date" in header_map and len(row) > header_map["date"]
            else None
        )
        ticker_value = (
            row[header_map["ticker"]].value
            if "ticker" in header_map and len(row) > header_map["ticker"]
            else None
        )
        comment_value = (
            row[header_map["comment"]].value
            if "comment" in header_map and len(row) > header_map["comment"]
            else None
        )
        operations.append(
            {
                "operation_type": str(type_cell.value).strip(),
                "operation_kind": _xtb_operation_kind(str(type_cell.value)),
                "trade_date": _parse_operation_date(date_value),
                "ticker": str(ticker_value).strip() if ticker_value else None,
                "description": str(comment_value).strip() if comment_value else None,
                "amount": float(amount),
            }
        )
    return operations


def _xtb_positions_from_xls(sheet) -> list[dict]:
    header_row = None
    header_map: dict[str, int] = {}
    max_scan = min(sheet.nrows, 40)
    for row_idx in range(max_scan):
        row = sheet.row_values(row_idx)
        for col_idx, value in enumerate(row):
            if not isinstance(value, str):
                continue
            key = _normalize_text(value)
            if key in {"symbol", "ticker"}:
                header_map["ticker"] = col_idx
                header_row = row_idx
            if key in {"volume", "shares"}:
                header_map["shares"] = col_idx
                header_row = row_idx
            if key in {"open price", "openprice"}:
                header_map["open_price"] = col_idx
                header_row = row_idx
            if key in {"purchase value", "purchasevalue"}:
                header_map["purchase_value"] = col_idx
                header_row = row_idx
            if key in {"market price", "marketprice", "current price"}:
                header_map["current_price"] = col_idx
                header_row = row_idx
            if key in {"name", "company"}:
                header_map["name"] = col_idx
                header_row = row_idx
        if header_row is not None and {"ticker", "shares", "open_price"}.issubset(
            header_map
        ):
            break
    if header_row is None or not {"ticker", "shares", "open_price"}.issubset(header_map):
        return []
    positions: list[dict] = []
    for row_idx in range(header_row + 1, sheet.nrows):
        ticker_value = sheet.cell_value(row_idx, header_map["ticker"])
        if ticker_value is None or str(ticker_value).strip() == "":
            continue
        if _normalize_text(str(ticker_value)) == "total":
            continue
        shares = _parse_number(sheet.cell_value(row_idx, header_map["shares"]))
        open_price = _parse_number(sheet.cell_value(row_idx, header_map["open_price"]))
        purchase_value = (
            _parse_number(sheet.cell_value(row_idx, header_map["purchase_value"]))
            if "purchase_value" in header_map
            else None
        )
        current_price = (
            _parse_number(sheet.cell_value(row_idx, header_map["current_price"]))
            if "current_price" in header_map
            else None
        )
        name_value = (
            sheet.cell_value(row_idx, header_map["name"])
            if "name" in header_map
            else None
        )
        if shares is None or open_price is None:
            continue
        positions.append(
            {
                "ticker": str(ticker_value).strip(),
                "shares": float(shares),
                "open_price": float(open_price),
                "purchase_value": float(purchase_value)
                if purchase_value is not None
                else None,
                "current_price": float(current_price) if current_price is not None else None,
                "name": str(name_value).strip() if isinstance(name_value, str) else None,
            }
        )
    return positions


def _xtb_positions_from_xlsx(sheet) -> list[dict]:
    header_row = None
    header_map: dict[str, int] = {}
    max_scan = min(sheet.max_row or 0, 40)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan):
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            key = _normalize_text(value)
            if key in {"symbol", "ticker"}:
                header_map["ticker"] = cell.column - 1
                header_row = cell.row
            if key in {"volume", "shares"}:
                header_map["shares"] = cell.column - 1
                header_row = cell.row
            if key in {"open price", "openprice"}:
                header_map["open_price"] = cell.column - 1
                header_row = cell.row
            if key in {"purchase value", "purchasevalue"}:
                header_map["purchase_value"] = cell.column - 1
                header_row = cell.row
            if key in {"market price", "marketprice", "current price"}:
                header_map["current_price"] = cell.column - 1
                header_row = cell.row
            if key in {"name", "company"}:
                header_map["name"] = cell.column - 1
                header_row = cell.row
        if header_row is not None and {"ticker", "shares", "open_price"}.issubset(
            header_map
        ):
            break
    if header_row is None or not {"ticker", "shares", "open_price"}.issubset(header_map):
        return []
    positions: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        ticker_cell = row[header_map["ticker"]] if len(row) > header_map["ticker"] else None
        if not ticker_cell or ticker_cell.value is None:
            continue
        if _normalize_text(str(ticker_cell.value)) == "total":
            continue
        shares_cell = row[header_map["shares"]] if len(row) > header_map["shares"] else None
        open_cell = (
            row[header_map["open_price"]] if len(row) > header_map["open_price"] else None
        )
        if not shares_cell or not open_cell:
            continue
        shares = _parse_number(shares_cell.value)
        open_price = _parse_number(open_cell.value)
        current_price = None
        if "current_price" in header_map and len(row) > header_map["current_price"]:
            current_price = _parse_number(row[header_map["current_price"]].value)
        purchase_value = None
        if "purchase_value" in header_map and len(row) > header_map["purchase_value"]:
            purchase_value = _parse_number(row[header_map["purchase_value"]].value)
        name_value = None
        if "name" in header_map and len(row) > header_map["name"]:
            name_value = row[header_map["name"]].value
        if shares is None or open_price is None:
            continue
        positions.append(
            {
                "ticker": str(ticker_cell.value).strip(),
                "shares": float(shares),
                "open_price": float(open_price),
                "purchase_value": float(purchase_value)
                if purchase_value is not None
                else None,
                "current_price": float(current_price) if current_price is not None else None,
                "name": str(name_value).strip() if isinstance(name_value, str) else None,
            }
        )
    return positions


def _aggregate_xtb_positions(positions: list[dict]) -> list[dict]:
    aggregated: dict[str, dict] = {}
    for position in positions:
        ticker = str(position.get("ticker") or "").strip().upper()
        if not ticker:
            continue
        shares = float(position.get("shares") or 0)
        if shares <= 0:
            continue
        entry = aggregated.setdefault(
            ticker,
            {
                "ticker": ticker,
                "name": position.get("name"),
                "shares": 0.0,
                "purchase_value": 0.0,
                "open_price_total": 0.0,
                "open_price_shares": 0.0,
                "current_price_total": 0.0,
                "current_price_shares": 0.0,
            },
        )
        entry["shares"] += shares
        purchase_value = position.get("purchase_value")
        if purchase_value is not None:
            entry["purchase_value"] += float(purchase_value)
        else:
            open_price = position.get("open_price")
            if open_price is not None:
                entry["purchase_value"] += float(open_price) * shares
        open_price = position.get("open_price")
        if open_price is not None:
            entry["open_price_total"] += float(open_price) * shares
            entry["open_price_shares"] += shares
        current_price = position.get("current_price")
        if current_price is not None:
            entry["current_price_total"] += float(current_price) * shares
            entry["current_price_shares"] += shares
        if not entry.get("name") and position.get("name"):
            entry["name"] = position.get("name")
    items: list[dict] = []
    for entry in aggregated.values():
        shares = float(entry["shares"] or 0)
        if shares <= 0:
            continue
        purchase_value = float(entry["purchase_value"] or 0)
        avg_price = (
            purchase_value / shares
            if purchase_value
            else (
                entry["open_price_total"] / entry["open_price_shares"]
                if entry["open_price_shares"]
                else 0.0
            )
        )
        avg_current_price = (
            entry["current_price_total"] / entry["current_price_shares"]
            if entry["current_price_shares"]
            else None
        )
        items.append(
            {
                "ticker": entry["ticker"],
                "name": entry.get("name"),
                "shares": shares,
                "open_price": avg_price,
                "purchase_value": purchase_value if purchase_value else None,
                "current_price": avg_current_price,
            }
        )
    return items


def _parse_xtb_file(file_bytes: bytes, filename: str) -> dict:
    warnings: list[str] = []
    operations: list[dict] = []
    if filename.lower().endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        cash_sheet = _xtb_sheet_by_name_xls(workbook, "CASH OPERATION HISTORY")
        if cash_sheet is None:
            raise HTTPException(
                status_code=400,
                detail=f"Missing CASH OPERATION HISTORY sheet in {filename}.",
            )
        cash_value = _parse_number(cash_sheet.cell_value(7, 3))
        current_value = _parse_number(cash_sheet.cell_value(7, 4))
        if current_value is None:
            raise HTTPException(
                status_code=400,
                detail=f"Missing current value (E8) in {filename}.",
            )
        invested_total = 0.0
        for row_idx in range(cash_sheet.nrows):
            type_value = cash_sheet.cell_value(row_idx, 2)
            if isinstance(type_value, str) and type_value.strip().lower() in {
                "stock purchase",
                "stock sale",
            }:
                amount = _parse_number(cash_sheet.cell_value(row_idx, 6))
                if amount is not None:
                    invested_total += float(amount)
        operations = _xtb_cash_operations_from_xls(cash_sheet)
        open_sheets = [
            workbook.sheet_by_name(name)
            for name in workbook.sheet_names()
            if _normalize_text(name).startswith("open position")
        ]
        positions: list[dict] = []
        profit_values: list[float] = []
        for sheet in open_sheets:
            positions.extend(_xtb_positions_from_xls(sheet))
            for row_idx in range(sheet.nrows):
                if sheet.ncols <= 15:
                    continue
                value = _parse_number(sheet.cell_value(row_idx, 15))
                if value is not None:
                    profit_values.append(float(value))
    else:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            cash_sheet = _xtb_sheet_by_name_xlsx(workbook, "CASH OPERATION HISTORY")
            if cash_sheet is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing CASH OPERATION HISTORY sheet in {filename}.",
                )
            cash_value = _parse_number(cash_sheet.cell(row=8, column=4).value)
            current_value = _parse_number(cash_sheet.cell(row=8, column=5).value)
            if current_value is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing current value (E8) in {filename}.",
                )
            invested_total = 0.0
            for row in cash_sheet.iter_rows(
                min_row=1, max_row=cash_sheet.max_row, max_col=7
            ):
                type_value = row[2].value if len(row) > 2 else None
                if isinstance(type_value, str) and type_value.strip().lower() in {
                    "stock purchase",
                    "stock sale",
                }:
                    amount = _parse_number(row[6].value if len(row) > 6 else None)
                    if amount is not None:
                        invested_total += float(amount)
            operations = _xtb_cash_operations_from_xlsx(cash_sheet)
            open_sheets = [
                sheet
                for sheet in workbook.worksheets
                if _normalize_text(sheet.title).startswith("open position")
            ]
            positions: list[dict] = []
            profit_values: list[float] = []
            for sheet in open_sheets:
                positions.extend(_xtb_positions_from_xlsx(sheet))
                for row in sheet.iter_rows(
                    min_row=1, max_row=sheet.max_row, max_col=16
                ):
                    value = _parse_number(row[15].value if len(row) > 15 else None)
                    if value is not None:
                        profit_values.append(float(value))
        finally:
            workbook.close()

    if not open_sheets:
        warnings.append("No OPEN POSITION sheets found.")
    if open_sheets and not positions:
        warnings.append("No OPEN POSITION rows parsed for holdings.")

    profit_value = 0.0
    if profit_values:
        profit_value = sum(profit_values[:-1]) if len(profit_values) > 1 else sum(profit_values)

    cash_value = float(cash_value or 0)
    current_value = float(current_value or 0)
    invested = abs(float(invested_total))
    profit_percent = (profit_value / current_value * 100) if current_value else None
    return {
        "current_value": current_value,
        "cash_value": cash_value,
        "invested": invested,
        "profit_value": profit_value,
        "profit_percent": profit_percent,
        "warnings": warnings,
        "positions": positions,
        "operations": operations,
    }


def _save_savengrow_import(
    portfolio_id: int,
    payload: SaveNGrowCommitRequest,
    totals: dict,
    currency: str,
) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO save_ngrow_imports (
                portfolio_id,
                invested_total,
                current_value_total,
                profit_value_total,
                profit_percent_total,
                currency,
                source_file,
                file_hash,
                snapshot_date,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                portfolio_id,
                totals["invested_total"],
                totals["current_value_total"],
                totals["profit_value_total"],
                totals.get("profit_percent_total"),
                currency,
                payload.filename,
                payload.file_hash,
                payload.snapshot_date,
                now,
            ),
        )
        import_id = cursor.lastrowid
    return {"id": import_id, "created_at": now}


def _save_savengrow_items(import_id: int, items: list[SaveNGrowItem]) -> None:
    with _db_connection() as conn:
        for item in items:
            conn.execute(
                """
                INSERT INTO save_ngrow_items (
                    import_id,
                    item_name,
                    invested,
                    current_value,
                    profit_value,
                    profit_percent,
                    category
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    item.name,
                    item.invested,
                    item.current_value,
                    item.profit_value,
                    item.profit_percent,
                    item.category,
                ),
            )


def _list_savengrow_imports(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   invested_total,
                   current_value_total,
                   profit_value_total,
                   profit_percent_total,
                   currency,
                   source_file,
                   snapshot_date,
                   created_at
            FROM save_ngrow_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
    results: list[dict] = []
    for row in rows:
        invested_total = float(row["invested_total"] or 0)
        current_value_total = float(row["current_value_total"] or 0)
        raw_profit = (
            float(row["profit_value_total"])
            if row["profit_value_total"] is not None
            else None
        )
        profit_value_total = _normalize_profit_value(
            current_value_total, invested_total, raw_profit
        )
        profit_percent_total = (
            float(row["profit_percent_total"])
            if row["profit_percent_total"] is not None
            else None
        )
        if (
            profit_percent_total is None
            and profit_value_total is not None
            and current_value_total
        ):
            profit_percent_total = (profit_value_total / current_value_total) * 100
        results.append(
            {
                "id": row["id"],
                "invested_total": invested_total,
                "current_value_total": current_value_total,
                "profit_value_total": profit_value_total,
                "profit_percent_total": profit_percent_total,
                "currency": row["currency"],
                "source_file": row["source_file"],
                "snapshot_date": row["snapshot_date"],
                "created_at": row["created_at"],
            }
        )
    return results


def _save_aforronet_import(
    portfolio_id: int,
    payload: AforroNetCommitRequest,
    totals: dict,
    currency: str,
    category: str,
) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO aforronet_imports (
                portfolio_id,
                invested_total,
                current_value_total,
                category,
                currency,
                source_file,
                file_hash,
                snapshot_date,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                portfolio_id,
                totals["invested_total"],
                totals["current_value_total"],
                category,
                currency,
                payload.filename,
                payload.file_hash,
                payload.snapshot_date,
                now,
            ),
        )
        import_id = cursor.lastrowid
    return {"id": import_id, "created_at": now}


def _save_aforronet_items(import_id: int, items: list[AforroNetItem]) -> None:
    with _db_connection() as conn:
        for item in items:
            conn.execute(
                """
                INSERT INTO aforronet_items (
                    import_id,
                    item_name,
                    invested,
                    current_value,
                    category
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    item.name,
                    item.invested,
                    item.current_value,
                    item.category,
                ),
            )


def _list_aforronet_imports(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   invested_total,
                   current_value_total,
                   category,
                   currency,
                   source_file,
                   snapshot_date,
                   created_at
            FROM aforronet_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
    return [
        {
            "id": row["id"],
            "invested_total": float(row["invested_total"] or 0),
            "current_value_total": float(row["current_value_total"] or 0),
            "category": row["category"],
            "currency": row["currency"],
            "source_file": row["source_file"],
            "snapshot_date": row["snapshot_date"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def _delete_aforronet_import(portfolio_id: int, import_id: int) -> int:
    with _db_connection() as conn:
        conn.execute(
            """
            DELETE FROM aforronet_items
            WHERE import_id = ? AND import_id IN (
                SELECT id FROM aforronet_imports WHERE portfolio_id = ?
            )
            """,
            (import_id, portfolio_id),
        )
        cursor = conn.execute(
            "DELETE FROM aforronet_imports WHERE id = ? AND portfolio_id = ?",
            (import_id, portfolio_id),
        )
        return cursor.rowcount


def _latest_aforronet_category(portfolio_id: int) -> str | None:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT category
            FROM aforronet_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
    return row["category"] if row else None


def _save_bancoinvest_import(
    portfolio_id: int, payload: BancoInvestCommitRequest
) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO bancoinvest_imports (
                portfolio_id,
                source_file,
                file_hash,
                snapshot_date,
                imported_at
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                portfolio_id,
                payload.filename,
                payload.file_hash,
                payload.snapshot_date,
                now,
            ),
        )
        import_id = cursor.lastrowid
    return {"id": import_id, "imported_at": now}


def _save_bancoinvest_items(import_id: int, items: list[BancoInvestItem]) -> None:
    with _db_connection() as conn:
        for item in items:
            conn.execute(
                """
                INSERT INTO bancoinvest_items (
                    import_id,
                    holder,
                    invested,
                    current_value,
                    gains,
                    category
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    item.holder,
                    item.invested,
                    item.current_value,
                    item.gains,
                    item.category,
                ),
            )


def _list_bancoinvest_imports(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT imp.id,
                   imp.source_file,
                   imp.snapshot_date,
                   imp.imported_at,
                   SUM(items.current_value) AS total_value
            FROM bancoinvest_imports AS imp
            JOIN bancoinvest_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id
            ORDER BY imp.imported_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
        imports = []
        for row in rows:
            item_rows = conn.execute(
                """
                SELECT holder, invested, current_value, gains, category
                FROM bancoinvest_items
                WHERE import_id = ?
                ORDER BY holder ASC
                """,
                (row["id"],),
            ).fetchall()
            imports.append(
                {
                    "id": row["id"],
                    "source_file": row["source_file"],
                    "snapshot_date": row["snapshot_date"],
                    "imported_at": row["imported_at"],
                    "total_value": float(row["total_value"] or 0),
                    "items": [
                        {
                            "holder": item["holder"],
                            "invested": float(item["invested"] or 0)
                            if item["invested"] is not None
                            else None,
                            "current_value": float(item["current_value"] or 0),
                            "gains": float(item["gains"] or 0)
                            if item["gains"] is not None
                            else None,
                            "category": item["category"],
                        }
                        for item in item_rows
                    ],
                }
            )
    return imports


def _delete_bancoinvest_import(portfolio_id: int, import_id: int) -> int:
    with _db_connection() as conn:
        conn.execute(
            "DELETE FROM bancoinvest_items WHERE import_id = ?",
            (import_id,),
        )
        cursor = conn.execute(
            """
            DELETE FROM bancoinvest_imports
            WHERE id = ? AND portfolio_id = ?
            """,
            (import_id, portfolio_id),
        )
        return cursor.rowcount


def _load_bancoinvest_category_map(portfolio_id: int) -> dict[str, str]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT holder_key, category
            FROM bancoinvest_category_map
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
    return {row["holder_key"]: row["category"] for row in rows}


def _upsert_bancoinvest_category_map(portfolio_id: int, items: list[BancoInvestItem]) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        for item in items:
            holder_key = _normalize_text(item.holder).replace(" ", "")
            conn.execute(
                """
                INSERT OR REPLACE INTO bancoinvest_category_map (
                    portfolio_id, holder_key, category, updated_at
                )
                VALUES (?, ?, ?, ?)
                """,
                (portfolio_id, holder_key, item.category, now),
            )


def _save_xtb_imports(portfolio_id: int, items: list[XtbImportItem]) -> list[dict]:
    now = datetime.utcnow().isoformat()
    saved: list[dict] = []
    with _db_connection() as conn:
        for item in items:
            cursor = conn.execute(
                """
                INSERT INTO xtb_imports (
                    portfolio_id,
                    account_type,
                    category,
                    current_value,
                    cash_value,
                    invested,
                    profit_value,
                    profit_percent,
                    source_file,
                    file_hash,
                    imported_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    portfolio_id,
                    item.account_type,
                    item.category,
                    item.current_value,
                    item.cash_value,
                    item.invested,
                    item.profit_value,
                    item.profit_percent,
                    item.filename,
                    item.file_hash,
                    now,
                ),
            )
            saved.append({"id": cursor.lastrowid, "imported_at": now})
    return saved


def _list_xtb_imports(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   account_type,
                   category,
                   current_value,
                   cash_value,
                   invested,
                   profit_value,
                   profit_percent,
                   source_file,
                   imported_at
            FROM xtb_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
    return [
        {
            "id": row["id"],
            "account_type": row["account_type"],
            "category": row["category"],
            "current_value": float(row["current_value"] or 0),
            "cash_value": float(row["cash_value"] or 0),
            "invested": float(row["invested"] or 0),
            "profit_value": float(row["profit_value"] or 0)
            if row["profit_value"] is not None
            else None,
            "profit_percent": float(row["profit_percent"] or 0)
            if row["profit_percent"] is not None
            else None,
            "source_file": row["source_file"],
            "imported_at": row["imported_at"],
        }
        for row in rows
    ]


def _delete_xtb_import(portfolio_id: int, import_id: int) -> int:
    with _db_connection() as conn:
        file_row = conn.execute(
            """
            SELECT file_hash
            FROM xtb_imports
            WHERE id = ? AND portfolio_id = ?
            """,
            (import_id, portfolio_id),
        ).fetchone()
        cursor = conn.execute(
            """
            DELETE FROM xtb_imports
            WHERE id = ? AND portfolio_id = ?
            """,
            (import_id, portfolio_id),
        )
        deleted = cursor.rowcount
        if deleted and file_row:
            holding_row = conn.execute(
                """
                SELECT id
                FROM holdings_imports
                WHERE portfolio_id = ? AND institution = ? AND file_hash = ?
                """,
                (portfolio_id, "XTB", file_row["file_hash"]),
            ).fetchone()
            if holding_row:
                conn.execute(
                    "DELETE FROM holdings_items WHERE import_id = ?",
                    (holding_row["id"],),
                )
                conn.execute(
                    "DELETE FROM holdings_operations WHERE import_id = ?",
                    (holding_row["id"],),
                )
                conn.execute(
                    "DELETE FROM holdings_imports WHERE id = ?",
                    (holding_row["id"],),
                )
        return deleted


def _clear_portfolio_data(portfolio_id: int) -> None:
    with _db_connection() as conn:
        conn.execute(
            """
            DELETE FROM santander_items
            WHERE import_id IN (
                SELECT id FROM santander_imports WHERE portfolio_id = ?
            )
            """,
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM santander_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM trade_republic_entries WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM save_ngrow_entries WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            """
            DELETE FROM save_ngrow_items
            WHERE import_id IN (
                SELECT id FROM save_ngrow_imports WHERE portfolio_id = ?
            )
            """,
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM save_ngrow_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            """
            DELETE FROM aforronet_items
            WHERE import_id IN (
                SELECT id FROM aforronet_imports WHERE portfolio_id = ?
            )
            """,
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM aforronet_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM xtb_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            """
            DELETE FROM holdings_items
            WHERE import_id IN (
                SELECT id FROM holdings_imports WHERE portfolio_id = ?
            )
            """,
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM holdings_operations WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM holdings_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM holding_transactions WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            """
            DELETE FROM bancoinvest_items
            WHERE import_id IN (
                SELECT id FROM bancoinvest_imports WHERE portfolio_id = ?
            )
            """,
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM bancoinvest_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        )


def _delete_portfolio(portfolio_id: int) -> bool:
    _clear_portfolio_data(portfolio_id)
    with _db_connection() as conn:
        conn.execute(
            "DELETE FROM holding_tags WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM holding_tag_suppressed WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM holdings_metadata WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM santander_category_map WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM bancoinvest_category_map WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        conn.execute(
            "DELETE FROM portfolio_category_settings WHERE portfolio_id = ?",
            (portfolio_id,),
        )
        cursor = conn.execute(
            "DELETE FROM portfolios WHERE id = ?",
            (portfolio_id,),
        )
        return cursor.rowcount > 0


def _to_datetime(value: str | None) -> datetime:
    if not value:
        return datetime.min
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return datetime.min
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", text)
    if match:
        day, month, year = (int(part) for part in match.groups())
        try:
            return datetime(year, month, day)
        except ValueError:
            return datetime.min
    return datetime.min



def _date_key(value: str | None) -> str | None:
    if not value:
        return None
    timestamp = _to_datetime(value)
    if timestamp == datetime.min:
        return None
    return timestamp.date().isoformat()



def _month_key(value: str | None) -> str | None:
    if not value:
        return None
    timestamp = _to_datetime(value)
    if timestamp == datetime.min:
        return None
    return timestamp.strftime("%Y-%m")


def _latest_month_values(entries: list[tuple[datetime, float]]) -> tuple[float | None, float | None]:
    by_month: dict[str, tuple[datetime, float]] = {}
    for timestamp, value in entries:
        if timestamp == datetime.min:
            continue
        month = timestamp.strftime("%Y-%m")
        current = by_month.get(month)
        if not current or timestamp > current[0]:
            by_month[month] = (timestamp, value)
    if not by_month:
        return None, None
    months = sorted(by_month.keys())
    latest_value = by_month[months[-1]][1]
    previous_value = by_month[months[-2]][1] if len(months) > 1 else None
    vs_last_month = latest_value - previous_value if previous_value is not None else None
    return latest_value, vs_last_month


def _save_holdings_import(
    portfolio_id: int,
    institution: str,
    filename: str,
    file_hash: str,
    snapshot_date: str | None,
) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO holdings_imports (
                portfolio_id,
                institution,
                source_file,
                file_hash,
                snapshot_date,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (portfolio_id, institution, filename, file_hash, snapshot_date, now),
        )
        import_id = cursor.lastrowid
    return {"id": import_id, "created_at": now}


def _save_holdings_items(import_id: int, items: list[HoldingImportItem]) -> None:
    aggregated: dict[tuple[str, str], dict] = {}
    for item in items:
        ticker = item.ticker.strip().upper()
        category = item.category or "Stocks"
        key = (ticker, category)
        entry = aggregated.setdefault(
            key,
            {
                "ticker": ticker,
                "category": category,
                "name": item.name,
                "shares": 0.0,
                "purchase_value": 0.0,
                "open_price_total": 0.0,
                "open_price_shares": 0.0,
                "current_price_total": 0.0,
                "current_price_shares": 0.0,
            },
        )
        shares = float(item.shares or 0)
        if shares <= 0:
            continue
        entry["shares"] += shares
        if item.purchase_value is not None:
            entry["purchase_value"] += float(item.purchase_value)
        else:
            entry["purchase_value"] += float(item.open_price or 0) * shares
        if item.open_price is not None:
            entry["open_price_total"] += float(item.open_price) * shares
            entry["open_price_shares"] += shares
        if item.current_price is not None:
            entry["current_price_total"] += float(item.current_price) * shares
            entry["current_price_shares"] += shares
        if not entry.get("name") and item.name:
            entry["name"] = item.name

    with _db_connection() as conn:
        for entry in aggregated.values():
            shares = float(entry["shares"] or 0)
            if shares <= 0:
                continue
            purchase_value = float(entry["purchase_value"] or 0)
            avg_price = (
                purchase_value / shares
                if purchase_value
                else (
                    entry["open_price_total"] / entry["open_price_shares"]
                    if entry["open_price_shares"]
                    else 0.0
                )
            )
            avg_current_price = (
                entry["current_price_total"] / entry["current_price_shares"]
                if entry["current_price_shares"]
                else None
            )
            cost_basis = purchase_value if purchase_value else avg_price * shares
            conn.execute(
                """
                INSERT INTO holdings_items (
                    import_id,
                    ticker,
                    name,
                    shares,
                    avg_price,
                    cost_basis,
                    current_price,
                    current_value,
                    profit_value,
                    profit_percent,
                    category
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    entry["ticker"],
                    entry.get("name"),
                    shares,
                    avg_price,
                    cost_basis,
                    avg_current_price,
                    None,
                    None,
                    None,
                    entry["category"],
                ),
            )


def _save_holdings_operations(
    import_id: int,
    portfolio_id: int,
    source_file: str,
    operations: list[HoldingOperationItem],
    currency: str | None = None,
) -> None:
    if not operations:
        return
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        for op in operations:
            conn.execute(
                """
                INSERT INTO holdings_operations (
                    import_id,
                    portfolio_id,
                    source_file,
                    operation_type,
                    operation_kind,
                    ticker,
                    description,
                    amount,
                    currency,
                    trade_date,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    portfolio_id,
                    source_file,
                    op.operation_type,
                    op.operation_kind,
                    op.ticker.strip().upper() if op.ticker else None,
                    op.description,
                    float(op.amount) if op.amount is not None else None,
                    op.currency or currency,
                    op.trade_date,
                    now,
                ),
            )


def _list_holdings_operations(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   source_file,
                   operation_type,
                   operation_kind,
                   ticker,
                   description,
                   amount,
                   currency,
                   trade_date,
                   created_at
            FROM holdings_operations
            WHERE portfolio_id = ?
            ORDER BY trade_date DESC, created_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
    tags_map = _list_holding_tags(portfolio_id)
    items: list[dict] = []
    for row in rows:
        ticker = row["ticker"]
        tags = tags_map.get(ticker.upper(), []) if ticker else []
        items.append(
            {
                "id": row["id"],
                "source_file": row["source_file"],
                "operation_type": row["operation_type"],
                "operation_kind": row["operation_kind"],
                "ticker": ticker,
                "description": row["description"],
                "amount": float(row["amount"] or 0),
                "currency": row["currency"],
                "trade_date": row["trade_date"],
                "created_at": row["created_at"],
                "tags": sorted(tags, key=str.lower),
            }
        )
    return items


def _save_holding_transaction(portfolio_id: int, payload: HoldingTransactionRequest) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO holding_transactions (
                portfolio_id,
                institution,
                ticker,
                name,
                operation,
                trade_date,
                shares,
                price,
                fee,
                note,
                category,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                portfolio_id,
                payload.institution,
                payload.ticker.strip().upper(),
                payload.name,
                payload.operation,
                payload.trade_date,
                payload.shares,
                payload.price,
                payload.fee,
                payload.note,
                payload.category or "Stocks",
                now,
            ),
        )
        entry_id = cursor.lastrowid
    return {"id": entry_id, "created_at": now}


def _list_holding_transactions(portfolio_id: int) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id,
                   institution,
                   ticker,
                   name,
                   operation,
                   trade_date,
                   shares,
                   price,
                   fee,
                   note,
                   category,
                   created_at
            FROM holding_transactions
            WHERE portfolio_id = ?
            ORDER BY trade_date DESC, created_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
    return [
        {
            "id": row["id"],
            "institution": row["institution"],
            "ticker": row["ticker"],
            "name": row["name"],
            "operation": row["operation"],
            "trade_date": row["trade_date"],
            "shares": float(row["shares"] or 0),
            "price": float(row["price"] or 0),
            "fee": float(row["fee"] or 0) if row["fee"] is not None else None,
            "note": row["note"],
            "category": row["category"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def _get_cached_prices(tickers: list[str]) -> dict[str, dict]:
    if not tickers:
        return {}
    placeholders = ",".join("?" * len(tickers))
    with _db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT ticker, price, currency, updated_at
            FROM holdings_prices
            WHERE ticker IN ({placeholders})
            """,
            [ticker.upper() for ticker in tickers],
        ).fetchall()
    return {
        row["ticker"]: {
            "price": float(row["price"] or 0),
            "currency": row["currency"],
            "updated_at": row["updated_at"],
        }
        for row in rows
    }


def _upsert_price(ticker: str, price: float, currency: str | None = None) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO holdings_prices (ticker, price, currency, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(ticker)
            DO UPDATE SET price = excluded.price,
                          currency = excluded.currency,
                          updated_at = excluded.updated_at
            """,
            (ticker.upper(), price, currency, now),
        )


def _list_holdings_metadata(portfolio_id: int) -> dict[str, dict]:
    """Lista metadados dos holdings, priorizando ticker_metadata global sobre holdings_metadata por portfolio."""
    tags_map = _list_holding_tags(portfolio_id)
    
    with _db_connection() as conn:
        # Get portfolio-specific metadata
        portfolio_rows = conn.execute(
            """
            SELECT ticker, sector, industry, country, asset_type
            FROM holdings_metadata
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        
        # Get global ticker metadata with more complete information
        global_rows = conn.execute(
            """
            SELECT ticker, name, sector, industry, country, region, 
                   currency, exchange, asset_class
            FROM ticker_metadata
            """,
        ).fetchall()
    
    # Build portfolio-specific metadata map
    metadata_map = {
        row["ticker"].upper(): {
            "sector": row["sector"],
            "industry": row["industry"],
            "country": row["country"],
            "region": None,
            "currency": None,
            "exchange": None,
            "asset_type": row["asset_type"],
            "tags": tags_map.get(row["ticker"].upper(), []),
        }
        for row in portfolio_rows
    }
    
    # Override with global ticker metadata (priority)
    for row in global_rows:
        ticker_key = row["ticker"].upper()
        existing = metadata_map.get(ticker_key, {})
        metadata_map[ticker_key] = {
            "sector": row["sector"] or existing.get("sector"),
            "industry": row["industry"] or existing.get("industry"),
            "country": row["country"] or existing.get("country"),
            "region": row["region"],
            "currency": row["currency"],
            "exchange": row["exchange"],
            "asset_type": row["asset_class"] or existing.get("asset_type"),
            "tags": existing.get("tags", tags_map.get(ticker_key, [])),
        }
    
    return metadata_map


def _save_banking_import(
    portfolio_id: int,
    institution: str,
    source_file: str,
    file_hash: str,
    row_count: int,
) -> int:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO banking_imports (
                portfolio_id,
                institution,
                source_file,
                file_hash,
                imported_at,
                row_count
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (portfolio_id, institution, source_file, file_hash, now, row_count),
        )
        return int(cursor.lastrowid)


def _save_banking_transactions(
    portfolio_id: int, import_id: int, institution: str, items: list[dict]
) -> None:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.executemany(
            """
            INSERT INTO banking_transactions (
                import_id,
                portfolio_id,
                institution,
                tx_date,
                description,
                amount,
                balance,
                currency,
                category,
                subcategory,
                raw_json,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    import_id,
                    portfolio_id,
                    institution,
                    item["tx_date"],
                    item["description"],
                    float(item["amount"]),
                    float(item["balance"]) if item.get("balance") is not None else None,
                    item.get("currency") or "EUR",
                    item.get("category") or BANKING_DEFAULT_CATEGORY,
                    item.get("subcategory") or BANKING_DEFAULT_SUBCATEGORY,
                    json.dumps(item.get("raw") or {}),
                    now,
                )
                for item in items
            ],
        )


def _list_banking_transactions(
    portfolio_id: int,
    month: str | None = None,
    category: str | None = None,
    subcategory: str | None = None,
    institution: str | None = None,
) -> list[dict]:
    query = [
        "SELECT id, tx_date, description, amount, balance, currency, category, subcategory, institution",
        "FROM banking_transactions",
        "WHERE portfolio_id = ?",
    ]
    params: list[object] = [portfolio_id]
    if month:
        query.append("AND substr(tx_date, 1, 7) = ?")
        params.append(month)
    if category:
        query.append("AND lower(category) = ?")
        params.append(_normalize_text(category))
    if subcategory:
        query.append("AND lower(subcategory) = ?")
        params.append(_normalize_text(subcategory))
    if institution:
        query.append("AND lower(institution) = ?")
        params.append(_normalize_text(institution))
    query.append("ORDER BY tx_date DESC")
    with _db_connection() as conn:
        rows = conn.execute("\n".join(query), params).fetchall()
    return [
        {
            "id": row["id"],
            "tx_date": row["tx_date"],
            "description": row["description"],
            "amount": float(row["amount"] or 0),
            "balance": float(row["balance"]) if row["balance"] is not None else None,
            "currency": row["currency"],
            "category": row["category"],
            "subcategory": row["subcategory"],
            "institution": row["institution"],
        }
        for row in rows
    ]


def _update_banking_transaction_category(
    portfolio_id: int,
    tx_id: int,
    category: str,
    subcategory: str | None,
) -> dict:
    category_name = (category or BANKING_DEFAULT_CATEGORY).strip()
    subcategory_name = (subcategory or BANKING_DEFAULT_SUBCATEGORY).strip()
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT institution, description
            FROM banking_transactions
            WHERE id = ? AND portfolio_id = ?
            """,
            (tx_id, portfolio_id),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Transaction not found.")
        conn.execute(
            """
            UPDATE banking_transactions
            SET category = ?, subcategory = ?
            WHERE id = ? AND portfolio_id = ?
            """,
            (category_name, subcategory_name, tx_id, portfolio_id),
        )
    _learn_banking_rule(
        portfolio_id,
        row["institution"],
        row["description"],
        category_name,
        subcategory_name,
    )
    return {
        "id": tx_id,
        "category": category_name,
        "subcategory": subcategory_name,
    }


def _clear_banking_transactions(portfolio_id: int) -> dict[str, int]:
    with _db_connection() as conn:
        tx_deleted = conn.execute(
            "DELETE FROM banking_transactions WHERE portfolio_id = ?",
            (portfolio_id,),
        ).rowcount
        import_deleted = conn.execute(
            "DELETE FROM banking_imports WHERE portfolio_id = ?",
            (portfolio_id,),
        ).rowcount
        institutions_deleted = conn.execute(
            "DELETE FROM banking_institutions WHERE portfolio_id = ?",
            (portfolio_id,),
        ).rowcount
    return {
        "transactions": tx_deleted or 0,
        "imports": import_deleted or 0,
        "institutions": institutions_deleted or 0,
    }


def _upsert_banking_budget(
    portfolio_id: int, category: str, month: str, amount: float
) -> dict:
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO banking_budgets (
                portfolio_id,
                category,
                month,
                amount,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(portfolio_id, category, month)
            DO UPDATE SET
                amount = excluded.amount,
                updated_at = excluded.updated_at
            """,
            (portfolio_id, category, month, amount, now, now),
        )
        row = conn.execute(
            """
            SELECT id, category, month, amount, created_at, updated_at
            FROM banking_budgets
            WHERE portfolio_id = ? AND category = ? AND month = ?
            """,
            (portfolio_id, category, month),
        ).fetchone()
    return dict(row) if row else {}


def _delete_banking_budget(portfolio_id: int, budget_id: int) -> bool:
    with _db_connection() as conn:
        row = conn.execute(
            """
            SELECT id
            FROM banking_budgets
            WHERE id = ? AND portfolio_id = ?
            """,
            (budget_id, portfolio_id),
        ).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM banking_budgets WHERE id = ?", (budget_id,))
    return True


def _list_banking_budgets(portfolio_id: int, month: str) -> list[dict]:
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, category, month, amount
            FROM banking_budgets
            WHERE portfolio_id = ? AND month = ?
            ORDER BY category
            """,
            (portfolio_id, month),
        ).fetchall()
        results: list[dict] = []
        for row in rows:
            spent_row = conn.execute(
                """
                SELECT SUM(ABS(amount)) AS spent
                FROM banking_transactions
                WHERE portfolio_id = ?
                  AND lower(category) = ?
                  AND amount < 0
                  AND substr(tx_date, 1, 7) = ?
                """,
                (portfolio_id, _normalize_text(row["category"]), month),
            ).fetchone()
            spent = float(spent_row["spent"] or 0)
            amount = float(row["amount"] or 0)
            remaining = amount - spent
            percent = round(spent / amount * 100, 2) if amount else 0.0
            results.append(
                {
                    "id": row["id"],
                    "category": row["category"],
                    "month": row["month"],
                    "amount": amount,
                    "spent": spent,
                    "remaining": remaining,
                    "percent": percent,
                }
            )
    return results


def _upsert_holdings_metadata(
    portfolio_id: int, payload: HoldingMetadataRequest
) -> dict:
    ticker = payload.ticker.strip().upper()
    now = datetime.utcnow().isoformat()
    sector = _normalize_optional_text(payload.sector)
    industry = _normalize_optional_text(payload.industry)
    country = _normalize_optional_text(payload.country)
    asset_type = _normalize_optional_text(payload.asset_type)
    tags = payload.tags
    saved_tags: list[str] | None = None
    if tags is not None:
        saved_tags = _set_holding_tags(portfolio_id, ticker, tags)
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO holdings_metadata (
                portfolio_id,
                ticker,
                sector,
                industry,
                country,
                asset_type,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(portfolio_id, ticker)
            DO UPDATE SET
                sector = excluded.sector,
                industry = excluded.industry,
                country = excluded.country,
                asset_type = excluded.asset_type,
                updated_at = excluded.updated_at
            """,
            (portfolio_id, ticker, sector, industry, country, asset_type, now),
        )
    return {
        "ticker": ticker,
        "sector": sector,
        "industry": industry,
        "country": country,
        "asset_type": asset_type,
        "tags": saved_tags,
        "updated_at": now,
    }


def _price_is_fresh(updated_at: str | None) -> bool:
    if not updated_at:
        return False
    timestamp = _to_datetime(updated_at)
    if timestamp == datetime.min:
        return False
    age = datetime.utcnow() - timestamp
    return age.total_seconds() < PRICE_CACHE_TTL_MINUTES * 60


def _fetch_price_twelvedata(ticker: str) -> float:
    if not PRICE_API_KEY:
        raise HTTPException(status_code=400, detail="Price API not configured.")
    query = urllib.parse.urlencode({"symbol": ticker, "apikey": PRICE_API_KEY})
    url = f"https://api.twelvedata.com/price?{query}"
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            response_text = response.read().decode("utf-8")
            if not response_text.strip():
                print(f"Twelve Data returned empty response for {ticker}")
                raise HTTPException(status_code=400, detail=f"Empty response from Twelve Data for {ticker}.")
            data = json.loads(response_text)
        
        # Check for error in response
        if "code" in data or "status" in data:
            error_msg = data.get("message", "Unknown error")
            print(f"Twelve Data error for {ticker}: {error_msg}")
            raise HTTPException(status_code=400, detail=f"Twelve Data error: {error_msg}")
        
        price_value = data.get("price")
        if price_value is None:
            print(f"Twelve Data returned no price for {ticker}: {data}")
            raise HTTPException(status_code=400, detail=f"Price unavailable for {ticker}.")
        return float(price_value)
    except json.JSONDecodeError as e:
        print(f"Twelve Data JSON decode error for {ticker}: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid response from Twelve Data for {ticker}.")
    except urllib.error.HTTPError as e:
        print(f"Twelve Data HTTP error for {ticker}: {e.code} {e.reason}")
        raise HTTPException(status_code=400, detail=f"Twelve Data HTTP error for {ticker}.")


def _fetch_price_finnhub(ticker: str) -> float:
    """Busca preço atual usando Finnhub /quote endpoint."""
    if not PRICE_API_KEY:
        raise HTTPException(status_code=400, detail="Price API not configured.")
    query = urllib.parse.urlencode({"symbol": ticker, "token": PRICE_API_KEY})
    url = f"https://finnhub.io/api/v1/quote?{query}"
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            response_text = response.read().decode("utf-8")
            if not response_text.strip():
                print(f"Finnhub returned empty response for {ticker}")
                raise HTTPException(status_code=400, detail=f"Empty response from Finnhub for {ticker}.")
            data = json.loads(response_text)
        
        # Check for error in response
        if "error" in data:
            error_msg = data.get("error", "Unknown error")
            print(f"Finnhub error for {ticker}: {error_msg}")
            raise HTTPException(status_code=400, detail=f"Finnhub error: {error_msg}")
        
        # Get current price (c = current price)
        price_value = data.get("c")
        if price_value is None or price_value == 0:
            print(f"Finnhub returned no valid price for {ticker}: {data}")
            raise HTTPException(status_code=400, detail=f"Price unavailable for {ticker}.")
        
        price_value = float(price_value)
        if price_value <= 0:
            raise HTTPException(status_code=400, detail=f"Invalid price for {ticker}.")
        
        return price_value
    except json.JSONDecodeError as e:
        print(f"Finnhub JSON decode error for {ticker}: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid response from Finnhub for {ticker}.")
    except urllib.error.HTTPError as e:
        print(f"Finnhub HTTP error for {ticker}: {e.code} {e.reason}")
        raise HTTPException(status_code=400, detail=f"Finnhub HTTP error for {ticker}.")


def _fetch_profile_finnhub(ticker: str) -> dict | None:
    """Busca perfil da empresa usando Finnhub /stock/profile2 endpoint."""
    if not PRICE_API_KEY:
        return None
    query = urllib.parse.urlencode({"symbol": ticker, "token": PRICE_API_KEY})
    url = f"https://finnhub.io/api/v1/stock/profile2?{query}"
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            response_text = response.read().decode("utf-8")
            if not response_text.strip():
                return None
            data = json.loads(response_text)
        
        if not data or "error" in data:
            return None
        
        return {
            "name": data.get("name"),
            "country": data.get("country"),
            "currency": data.get("currency"),
            "exchange": data.get("exchange"),
            "industry": data.get("finnhubIndustry"),
            "sector": data.get("finnhubIndustry"),  # Finnhub uses same field
        }
    except:
        return None


def _fetch_metrics_finnhub(ticker: str) -> dict | None:
    """Busca métricas usando Finnhub /stock/metric endpoint."""
    if not PRICE_API_KEY:
        return None
    query = urllib.parse.urlencode({"symbol": ticker, "metric": "all", "token": PRICE_API_KEY})
    url = f"https://finnhub.io/api/v1/stock/metric?{query}"
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            response_text = response.read().decode("utf-8")
            if not response_text.strip():
                return None
            data = json.loads(response_text)
        
        if not data or "error" in data:
            return None
        
        metrics = data.get("metric", {})
        return {
            "dividend_yield": metrics.get("dividendYieldIndicatedAnnual"),
        }
    except:
        return None


def _fetch_dividends_finnhub(ticker: str) -> dict | None:
    """Busca dividendos usando Finnhub /stock/dividend2 endpoint."""
    if not PRICE_API_KEY:
        return None
    
    # Get dividends for next 12 months
    from datetime import datetime, timedelta
    today = datetime.now()
    to_date = (today + timedelta(days=365)).strftime("%Y-%m-%d")
    from_date = today.strftime("%Y-%m-%d")
    
    query = urllib.parse.urlencode({
        "symbol": ticker,
        "from": from_date,
        "to": to_date,
        "token": PRICE_API_KEY
    })
    url = f"https://finnhub.io/api/v1/stock/dividend?{query}"
    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            response_text = response.read().decode("utf-8")
            if not response_text.strip():
                return None
            data = json.loads(response_text)
        
        if not data or "error" in data or not isinstance(data, list):
            return None
        
        # Get next dividend date (first future dividend)
        future_dividends = [d for d in data if d.get("payDate") and d.get("payDate") >= from_date]
        if future_dividends:
            next_div = min(future_dividends, key=lambda x: x.get("payDate", "9999-12-31"))
            return {
                "next_dividend_date": next_div.get("payDate"),
                "next_dividend_amount": next_div.get("amount")
            }
        
        return None
    except:
        return None


def _fetch_metadata_finnhub(ticker: str) -> dict | None:
    """Busca metadados completos usando múltiplos endpoints do Finnhub."""
    if not PRICE_API_KEY:
        return None
    
    import time
    
    metadata = {
        "ticker": ticker.upper(),
        "name": None,
        "asset_class": "Stock",
        "sector": None,
        "industry": None,
        "country": None,
        "region": None,
        "currency": "USD",
        "exchange": None,
        "dividend_yield": None,
        "dividend_frequency": None,
        "next_dividend_date": None,
        "next_dividend_amount": None
    }
    
    try:
        # 1. Get profile (name, country, exchange, industry)
        profile = _fetch_profile_finnhub(ticker)
        if profile:
            metadata.update({
                "name": profile.get("name") or ticker,
                "country": profile.get("country"),
                "currency": profile.get("currency") or "USD",
                "exchange": profile.get("exchange"),
                "industry": profile.get("industry"),
                "sector": profile.get("sector")
            })
        
        # Small delay between API calls (rate limit: 59/min)
        time.sleep(1.05)
        
        # 2. Get metrics (dividend yield)
        metrics = _fetch_metrics_finnhub(ticker)
        if metrics:
            metadata["dividend_yield"] = metrics.get("dividend_yield")
        
        time.sleep(1.05)
        
        # 3. Get dividends (next payment date)
        dividends = _fetch_dividends_finnhub(ticker)
        if dividends:
            metadata["next_dividend_date"] = dividends.get("next_dividend_date")
            metadata["next_dividend_amount"] = dividends.get("next_dividend_amount")
            if metadata["next_dividend_date"]:
                metadata["dividend_frequency"] = "Quarterly"  # Assume quarterly if has dividends
        
        time.sleep(1.05)
        
        # 4. Get current price
        try:
            price = _fetch_price_finnhub(ticker)
            if price:
                metadata["price"] = price
        except Exception as price_error:
            print(f"Could not fetch price for {ticker}: {price_error}")
            metadata["price"] = None
        
        return metadata
    except Exception as e:
        print(f"Error fetching Finnhub metadata for {ticker}: {e}")
        return None


def _fetch_latest_price(ticker: str) -> float:
    """Fetch latest price with fallback to yfinance if configured provider fails."""
    
    # Try configured provider first (also check for typo "finhub")
    if PRICE_API_PROVIDER in ("twelvedata",):
        try:
            return _fetch_price_twelvedata(ticker)
        except Exception as e:
            print(f"Twelve Data failed for {ticker}, falling back to yfinance: {str(e)[:100]}")
    
    if PRICE_API_PROVIDER in ("finnhub", "finhub"):
        try:
            return _fetch_price_finnhub(ticker)
        except Exception as e:
            print(f"Finnhub failed for {ticker}, falling back to yfinance: {str(e)[:100]}")
    
    # Fallback to yfinance (free, works for most tickers including European)
    print(f"Attempting yfinance for {ticker}...")
    try:
        price_data = _fetch_ticker_price_yfinance(ticker)
        if price_data and price_data.get("price"):
            print(f"✓ yfinance success for {ticker}: {price_data['price']}")
            return float(price_data["price"])
    except Exception as e:
        print(f"✗ yfinance also failed for {ticker}: {e}")
    
    raise HTTPException(status_code=400, detail=f"Price unavailable for {ticker}.")


def _aggregate_transactions_by_ticker(
    transactions: list[dict], snapshot_date: str | None
) -> dict[tuple[str, str | None, str], dict]:
    aggregated: dict[tuple[str, str | None, str], dict] = {}
    for tx in sorted(transactions, key=lambda item: item["trade_date"]):
        if snapshot_date and _date_key(tx["trade_date"]) > snapshot_date:
            continue
        key = (tx["ticker"], tx["institution"], tx["category"])
        state = aggregated.setdefault(
            key,
            {
                "ticker": tx["ticker"],
                "institution": tx["institution"],
                "category": tx["category"],
                "name": tx["name"],
                "shares": 0.0,
                "cost_basis": 0.0,
            },
        )
        shares = float(tx["shares"] or 0)
        price = float(tx["price"] or 0)
        fee = float(tx["fee"] or 0) if tx["fee"] is not None else 0.0
        operation = tx["operation"].strip().lower()
        if operation == "buy":
            state["shares"] += shares
            state["cost_basis"] += shares * price + fee
        elif operation == "sell":
            if state["shares"] > 0:
                avg_cost = state["cost_basis"] / state["shares"]
                state["shares"] -= shares
                state["cost_basis"] = max(state["cost_basis"] - avg_cost * shares, 0.0)
        if tx["name"] and not state.get("name"):
            state["name"] = tx["name"]
    return aggregated


def _convert_currency(amount: float, from_currency: str, to_currency: str) -> float:
    """Convert currency amount using simple fixed rates.
    In production, this should use a real exchange rate API."""
    if not amount or from_currency == to_currency:
        return amount
    
    # Fixed rates (USD as base)
    rates = {
        "USD": 1.0,
        "EUR": 0.92,  # 1 USD = 0.92 EUR
        "GBP": 0.79,
        "CHF": 0.85,
        "JPY": 149.0,
    }
    
    from_rate = rates.get(from_currency, 1.0)
    to_rate = rates.get(to_currency, 1.0)
    
    # Convert to USD first, then to target currency
    amount_usd = amount / from_rate
    return amount_usd * to_rate


def _list_holdings_for_portfolio(
    portfolio_id: int,
    category_settings: dict[str, bool],
    category: str | None = None,
    institution: str | None = None,
    ticker: str | None = None,
) -> dict:
    # Get portfolio currency for price conversion
    with _db_connection() as conn:
        portfolio_row = conn.execute(
            "SELECT currency FROM portfolios WHERE id = ?",
            (portfolio_id,),
        ).fetchone()
        portfolio_currency = portfolio_row["currency"] if portfolio_row else "USD"
    
    history_items = _list_portfolio_history(portfolio_id, category_settings)
    if not history_items:
        return {"items": [], "total_value": 0.0}
    latest_snapshot = history_items[-1]["date"]
    holdings: dict[tuple[str, str | None, str], dict] = {}

    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT imp.institution,
                   imp.source_file,
                   imp.snapshot_date,
                   imp.created_at,
                   items.ticker,
                   items.name,
                   items.shares,
                   items.avg_price,
                   items.cost_basis,
                   items.current_price,
                   items.category
            FROM holdings_imports AS imp
            JOIN holdings_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            snapshot_value = row["snapshot_date"] or row["created_at"]
            if _date_key(snapshot_value) != latest_snapshot:
                continue
            key = (row["ticker"], row["institution"], row["category"])
            holdings[key] = {
                "ticker": row["ticker"],
                "name": row["name"],
                "institution": row["institution"],
                "category": row["category"],
                "shares": float(row["shares"] or 0),
                "cost_basis": float(row["cost_basis"] or 0),
                "avg_price": float(row["avg_price"] or 0),
                "current_price": float(row["current_price"] or 0)
                if row["current_price"] is not None
                else None,
                "source": "import",
            }

        tx_rows = conn.execute(
            """
            SELECT institution,
                   ticker,
                   name,
                   operation,
                   trade_date,
                   shares,
                   price,
                   fee,
                   note,
                   category
            FROM holding_transactions
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        transactions = [
            {
                "institution": row["institution"],
                "ticker": row["ticker"],
                "name": row["name"],
                "operation": row["operation"],
                "trade_date": row["trade_date"],
                "shares": float(row["shares"] or 0),
                "price": float(row["price"] or 0),
                "fee": float(row["fee"] or 0) if row["fee"] is not None else None,
                "category": row["category"],
            }
            for row in tx_rows
        ]
        aggregated = _aggregate_transactions_by_ticker(transactions, latest_snapshot)
        for key, state in aggregated.items():
            if state["shares"] <= 0:
                continue
            if key in holdings:
                holdings[key]["shares"] += state["shares"]
                holdings[key]["cost_basis"] += state["cost_basis"]
                if not holdings[key].get("name") and state.get("name"):
                    holdings[key]["name"] = state.get("name")
                holdings[key]["source"] = "import+manual"
            else:
                holdings[key] = {
                    "ticker": state["ticker"],
                    "name": state.get("name"),
                    "institution": state.get("institution"),
                    "category": state.get("category"),
                    "shares": state["shares"],
                    "cost_basis": state["cost_basis"],
                    "avg_price": state["cost_basis"] / state["shares"]
                    if state["shares"]
                    else 0.0,
                    "current_price": None,
                    "source": "manual",
                }

    metadata_map = _list_holdings_metadata(portfolio_id)
    tags_map = _list_holding_tags(portfolio_id)
    suppressed_map = _list_suppressed_tags(portfolio_id)
    for entry in holdings.values():
        ticker_key = entry["ticker"].upper()
        meta = metadata_map.get(ticker_key, {})
        entry["sector"] = meta.get("sector")
        entry["industry"] = meta.get("industry")
        entry["country"] = meta.get("country")
        entry["region"] = meta.get("region")
        entry["ticker_currency"] = meta.get("currency")
        entry["exchange"] = meta.get("exchange")
        entry["asset_type"] = meta.get("asset_type")
        entry["tags"] = tags_map.get(ticker_key, [])
        auto_tags = _auto_tags_from_entry(entry)
        if auto_tags:
            entry["tags"] = _ensure_auto_tags(
                portfolio_id,
                ticker_key,
                auto_tags,
                entry["tags"],
                suppressed_map.get(ticker_key, set()),
            )
            tags_map[ticker_key] = entry["tags"]

    filtered_entries: list[dict] = []
    for entry in holdings.values():
        if category and _normalize_text(entry["category"] or "") != _normalize_text(category):
            continue
        if institution and _normalize_text(entry["institution"] or "") != _normalize_text(
            institution
        ):
            continue
        if ticker and entry["ticker"].upper() != ticker.strip().upper():
            continue
        filtered_entries.append(entry)

    if not filtered_entries:
        return {"items": [], "total_value": 0.0}

    tickers = [entry["ticker"] for entry in filtered_entries]
    price_cache = _get_cached_prices(tickers)
    
    # No auto-refresh - only use cached prices
    # User must click "Update All" button to refresh prices via /holdings/refresh-prices endpoint

    items = []
    total_value = 0.0
    for entry in filtered_entries:
        cache_key = entry["ticker"].upper()
        price_info = price_cache.get(cache_key)
        cached_price = (
            price_info["price"] if price_info and _price_is_fresh(price_info["updated_at"]) else None
        )
        if cached_price is None:
            cached_price = entry["current_price"] or entry["avg_price"]
        
        # Convert price to portfolio currency if needed
        ticker_currency = entry.get("ticker_currency") or "USD"
        if cached_price and ticker_currency != portfolio_currency:
            cached_price = _convert_currency(cached_price, ticker_currency, portfolio_currency)
        
        current_value = float(entry["shares"] or 0) * float(cached_price or 0)
        avg_price = (
            float(entry["cost_basis"] or 0) / float(entry["shares"] or 1)
            if entry["shares"]
            else 0.0
        )
        profit_value = current_value - float(entry["cost_basis"] or 0)
        profit_percent = (
            profit_value / float(entry["cost_basis"] or 1) * 100
            if entry["cost_basis"]
            else None
        )
        item = {
            "ticker": entry["ticker"],
            "name": entry.get("name"),
            "institution": entry.get("institution"),
            "category": entry.get("category"),
            "shares": float(entry["shares"] or 0),
            "avg_price": round(avg_price, 4),
            "cost_basis": round(float(entry["cost_basis"] or 0), 2),
            "current_price": round(float(cached_price or 0), 4),
            "current_value": round(current_value, 2),
            "profit_value": round(profit_value, 2),
            "profit_percent": round(profit_percent, 2) if profit_percent is not None else None,
            "sector": entry.get("sector"),
            "industry": entry.get("industry"),
            "country": entry.get("country"),
            "region": entry.get("region"),
            "currency": entry.get("ticker_currency"),
            "exchange": entry.get("exchange"),
            "asset_type": entry.get("asset_type"),
            "tags": sorted(entry.get("tags", []), key=str.lower),
            "source": entry.get("source"),
        }
        items.append(item)
        total_value += current_value

    items.sort(key=lambda value: value["current_value"], reverse=True)
    for item in items:
        item["share_percent"] = (
            round(item["current_value"] / total_value * 100, 2) if total_value else 0.0
        )
    return {"items": items, "total_value": round(total_value, 2)}


def _list_institutions(
    portfolio_id: int, category_settings: dict[str, bool]
) -> list[dict]:
    items: list[dict] = []
    with _db_connection() as conn:
        latest_santander = conn.execute(
            """
            SELECT id
            FROM santander_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if latest_santander:
            totals = conn.execute(
                """
                SELECT
                    SUM(balance) AS total,
                    SUM(CASE WHEN gains IS NOT NULL THEN gains ELSE 0 END) AS gains
                FROM santander_items
                WHERE import_id = ?
                """,
                (latest_santander["id"],),
            ).fetchone()
            category_rows = conn.execute(
                """
                SELECT lower(category) AS category, SUM(balance) AS total
                FROM santander_items
                WHERE import_id = ?
                GROUP BY lower(category)
                """,
                (latest_santander["id"],),
            ).fetchall()
            category_totals = {
                row["category"]: float(row["total"] or 0) for row in category_rows
            }
            month_rows = conn.execute(
                """
                SELECT imp.imported_at AS date_value, SUM(items.balance) AS total
                FROM santander_imports AS imp
                JOIN santander_items AS items ON items.import_id = imp.id
                WHERE imp.portfolio_id = ?
                GROUP BY imp.id
                """,
                (portfolio_id,),
            ).fetchall()
            month_entries = [
                (_to_datetime(row["date_value"]), float(row["total"] or 0))
                for row in month_rows
            ]
            _, vs_last_month = _latest_month_values(month_entries)
            total_value = float(totals["total"] or 0)
            if total_value or totals["total"] is not None:
                gains = float(totals["gains"] or 0)
                items.append(
                    {
                        "institution": "Santander",
                        "total": total_value,
                        "gains": gains,
                        "profit_percent": (gains / total_value * 100)
                        if total_value
                        else None,
                        "vs_last_month": vs_last_month,
                        "beni": category_totals.get("beni"),
                        "magui": category_totals.get("magui"),
                    }
                )

        trade_latest = conn.execute(
            """
            SELECT value, gains, category, snapshot_date, created_at
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if trade_latest:
            trade_rows = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value, value
                FROM trade_republic_entries
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
            trade_entries = [
                (_to_datetime(row["date_value"]), float(row["value"] or 0))
                for row in trade_rows
            ]
            _, vs_last_month = _latest_month_values(trade_entries)
            total_value = float(trade_latest["value"] or 0)
            gains = float(trade_latest["gains"] or 0)
            category_key = (trade_latest["category"] or "").lower()
            items.append(
                {
                    "institution": "Trade Republic",
                    "total": total_value,
                    "gains": gains,
                    "profit_percent": (gains / total_value * 100)
                    if total_value
                    else None,
                    "vs_last_month": vs_last_month,
                    "beni": total_value if category_key == "beni" else None,
                    "magui": total_value if category_key == "magui" else None,
                }
            )

        save_latest = conn.execute(
            """
            SELECT id, current_value_total, profit_value_total, snapshot_date, created_at
            FROM save_ngrow_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if save_latest:
            save_rows = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value,
                       current_value_total AS total
                FROM save_ngrow_imports
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
            save_entries = [
                (_to_datetime(row["date_value"]), float(row["total"] or 0))
                for row in save_rows
            ]
            _, vs_last_month = _latest_month_values(save_entries)
            save_items = conn.execute(
                """
                SELECT category, current_value, invested, profit_value
                FROM save_ngrow_items
                WHERE import_id = ?
                """,
                (save_latest["id"],),
            ).fetchall()
            save_category_totals: dict[str, float] = {}
            gains = 0.0
            for row in save_items:
                category_key = _normalize_text(row["category"])
                current_value = float(row["current_value"] or 0)
                invested_value = (
                    float(row["invested"]) if row["invested"] is not None else None
                )
                profit_value = _normalize_profit_value(
                    current_value, invested_value, row["profit_value"]
                )
                save_category_totals[category_key] = (
                    save_category_totals.get(category_key, 0.0) + current_value
                )
                gains += float(profit_value or 0)
            total_value = float(save_latest["current_value_total"] or 0)
            items.append(
                {
                    "institution": "Save N Grow",
                    "total": total_value,
                    "gains": gains,
                    "profit_percent": (gains / total_value * 100)
                    if total_value
                    else None,
                    "vs_last_month": vs_last_month,
                    "beni": save_category_totals.get("beni"),
                    "magui": save_category_totals.get("magui"),
                }
            )
        else:
            save_fallback = conn.execute(
                """
                SELECT current_value, profit_value, snapshot_date, created_at
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if save_fallback:
                save_rows = conn.execute(
                    """
                    SELECT COALESCE(snapshot_date, created_at) AS date_value,
                           current_value AS total
                    FROM save_ngrow_entries
                    WHERE portfolio_id = ?
                    """,
                    (portfolio_id,),
                ).fetchall()
                save_entries = [
                    (_to_datetime(row["date_value"]), float(row["total"] or 0))
                    for row in save_rows
                ]
                _, vs_last_month = _latest_month_values(save_entries)
                total_value = float(save_fallback["current_value"] or 0)
                gains = float(save_fallback["profit_value"] or 0)
                items.append(
                    {
                        "institution": "Save N Grow",
                        "total": total_value,
                        "gains": gains,
                        "profit_percent": (gains / total_value * 100)
                        if total_value
                        else None,
                        "vs_last_month": vs_last_month,
                        "beni": None,
                        "magui": None,
                    }
                )

        aforronet_latest = conn.execute(
            """
            SELECT current_value_total, invested_total, category, snapshot_date, created_at
            FROM aforronet_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if aforronet_latest:
            aforro_rows = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value,
                       current_value_total AS total
                FROM aforronet_imports
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
            aforro_entries = [
                (_to_datetime(row["date_value"]), float(row["total"] or 0))
                for row in aforro_rows
            ]
            _, vs_last_month = _latest_month_values(aforro_entries)
            total_value = float(aforronet_latest["current_value_total"] or 0)
            invested_value = float(aforronet_latest["invested_total"] or 0)
            gains = total_value - invested_value
            category_key = (aforronet_latest["category"] or "").lower()
            items.append(
                {
                    "institution": "AforroNet",
                    "total": total_value,
                    "gains": gains,
                    "profit_percent": (gains / total_value * 100)
                    if total_value
                    else None,
                    "vs_last_month": vs_last_month,
                    "beni": total_value if category_key == "beni" else None,
                    "magui": total_value if category_key == "magui" else None,
                }
            )

        xtb_rows = conn.execute(
            """
            SELECT account_type, current_value, profit_value, imported_at
            FROM xtb_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
        if xtb_rows:
            latest_by_account: dict[str, sqlite3.Row] = {}
            latest_by_month: dict[str, dict[str, tuple[datetime, float]]] = {}
            for row in xtb_rows:
                key = row["account_type"]
                if key not in latest_by_account:
                    latest_by_account[key] = row
                timestamp = _to_datetime(row["imported_at"])
                if timestamp != datetime.min:
                    month_key = timestamp.strftime("%Y-%m")
                    month_accounts = latest_by_month.setdefault(month_key, {})
                    current = month_accounts.get(key)
                    if not current or timestamp > current[0]:
                        month_accounts[key] = (timestamp, float(row["current_value"] or 0))
            if latest_by_account:
                total_value = sum(
                    float(row["current_value"] or 0) for row in latest_by_account.values()
                )
                total_gains = sum(
                    float(row["profit_value"] or 0) for row in latest_by_account.values()
                )
                monthly_totals = []
                for entries in latest_by_month.values():
                    if not entries:
                        continue
                    latest_timestamp = max(value[0] for value in entries.values())
                    month_total = sum(val for _, val in entries.values())
                    monthly_totals.append((latest_timestamp, month_total))
                _, vs_last_month = _latest_month_values(monthly_totals)
                beni_value = (
                    float(latest_by_account["Beni"]["current_value"] or 0)
                    if "Beni" in latest_by_account
                    else None
                )
                magui_value = (
                    float(latest_by_account["Magui"]["current_value"] or 0)
                    if "Magui" in latest_by_account
                    else None
                )
                items.append(
                    {
                        "institution": "XTB",
                        "total": total_value,
                        "gains": total_gains,
                        "profit_percent": (total_gains / total_value * 100)
                        if total_value
                        else None,
                        "vs_last_month": vs_last_month,
                        "beni": beni_value,
                        "magui": magui_value,
                    }
                )

        bancoinvest_import = conn.execute(
            """
            SELECT id, snapshot_date, imported_at
            FROM bancoinvest_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if bancoinvest_import:
            totals = conn.execute(
                """
                SELECT
                    SUM(current_value) AS total,
                    SUM(CASE WHEN gains IS NOT NULL THEN gains ELSE 0 END) AS gains
                FROM bancoinvest_items
                WHERE import_id = ?
                """,
                (bancoinvest_import["id"],),
            ).fetchone()
            category_rows = conn.execute(
                """
                SELECT lower(category) AS category, SUM(current_value) AS total
                FROM bancoinvest_items
                WHERE import_id = ?
                GROUP BY lower(category)
                """,
                (bancoinvest_import["id"],),
            ).fetchall()
            category_totals = {
                row["category"]: float(row["total"] or 0) for row in category_rows
            }
            month_rows = conn.execute(
                """
                SELECT COALESCE(imp.snapshot_date, imp.imported_at) AS date_value,
                       SUM(items.current_value) AS total
                FROM bancoinvest_imports AS imp
                JOIN bancoinvest_items AS items ON items.import_id = imp.id
                WHERE imp.portfolio_id = ?
                GROUP BY imp.id
                """,
                (portfolio_id,),
            ).fetchall()
            month_entries = [
                (_to_datetime(row["date_value"]), float(row["total"] or 0))
                for row in month_rows
            ]
            _, vs_last_month = _latest_month_values(month_entries)
            total_value = float(totals["total"] or 0)
            gains = float(totals["gains"] or 0)
            items.append(
                {
                    "institution": "BancoInvest",
                    "total": total_value,
                    "gains": gains,
                    "profit_percent": (gains / total_value * 100)
                    if total_value
                    else None,
                    "vs_last_month": vs_last_month,
                    "beni": category_totals.get("beni"),
                    "magui": category_totals.get("magui"),
                }
            )

    return items


def _list_portfolio_monthly_history(portfolio_id: int) -> list[dict]:
    latest_santander: dict[str, tuple[datetime, float]] = {}
    latest_trade: dict[str, tuple[datetime, float]] = {}
    latest_save: dict[str, tuple[datetime, float]] = {}
    latest_aforronet: dict[str, tuple[datetime, float]] = {}
    latest_bancoinvest: dict[str, tuple[datetime, float]] = {}
    latest_xtb: dict[str, dict[str, tuple[datetime, float]]] = {}

    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT imp.imported_at AS imported_at,
                   SUM(items.balance) AS total
            FROM santander_imports AS imp
            JOIN santander_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            month = _month_key(row["imported_at"])
            if not month:
                continue
            timestamp = _to_datetime(row["imported_at"])
            total = float(row["total"] or 0)
            current = latest_santander.get(month)
            if not current or timestamp > current[0]:
                latest_santander[month] = (timestamp, total)

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value, value
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            month = _month_key(row["date_value"])
            if not month:
                continue
            timestamp = _to_datetime(row["date_value"])
            total = float(row["value"] or 0)
            current = latest_trade.get(month)
            if not current or timestamp > current[0]:
                latest_trade[month] = (timestamp, total)

        save_imports = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   current_value_total AS total
            FROM save_ngrow_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        save_entries = []
        if not save_imports:
            save_entries = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value,
                       current_value AS total
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
        for row in save_imports or save_entries:
            month = _month_key(row["date_value"])
            if not month:
                continue
            timestamp = _to_datetime(row["date_value"])
            total = float(row["total"] or 0)
            current = latest_save.get(month)
            if not current or timestamp > current[0]:
                latest_save[month] = (timestamp, total)

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   current_value_total AS total
            FROM aforronet_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            month = _month_key(row["date_value"])
            if not month:
                continue
            timestamp = _to_datetime(row["date_value"])
            total = float(row["total"] or 0)
            current = latest_aforronet.get(month)
            if not current or timestamp > current[0]:
                latest_aforronet[month] = (timestamp, total)

        rows = conn.execute(
            """
            SELECT COALESCE(imp.snapshot_date, imp.imported_at) AS date_value,
                   SUM(items.current_value) AS total
            FROM bancoinvest_imports AS imp
            JOIN bancoinvest_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            month = _month_key(row["date_value"])
            if not month:
                continue
            timestamp = _to_datetime(row["date_value"])
            total = float(row["total"] or 0)
            current = latest_bancoinvest.get(month)
            if not current or timestamp > current[0]:
                latest_bancoinvest[month] = (timestamp, total)

        rows = conn.execute(
            """
            SELECT account_type, imported_at, current_value
            FROM xtb_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            month = _month_key(row["imported_at"])
            if not month:
                continue
            timestamp = _to_datetime(row["imported_at"])
            total = float(row["current_value"] or 0)
            month_accounts = latest_xtb.setdefault(month, {})
            current = month_accounts.get(row["account_type"])
            if not current or timestamp > current[0]:
                month_accounts[row["account_type"]] = (timestamp, total)

    months = set()
    months.update(latest_santander.keys())
    months.update(latest_trade.keys())
    months.update(latest_save.keys())
    months.update(latest_aforronet.keys())
    months.update(latest_bancoinvest.keys())
    months.update(latest_xtb.keys())

    items: list[dict] = []
    for month in months:
        total = 0.0
        if month in latest_santander:
            total += latest_santander[month][1]
        if month in latest_trade:
            total += latest_trade[month][1]
        if month in latest_save:
            total += latest_save[month][1]
        if month in latest_aforronet:
            total += latest_aforronet[month][1]
        if month in latest_bancoinvest:
            total += latest_bancoinvest[month][1]
        if month in latest_xtb:
            total += sum(value for _, value in latest_xtb[month].values())
        items.append({"month": month, "total": round(total, 2)})
    items.sort(key=lambda item: item["month"])
    return items


def _list_portfolio_history(
    portfolio_id: int,
    category_settings: dict[str, bool],
) -> list[dict]:
    history: dict[str, dict[str, float | str]] = {}

    def normalize_key(category: str) -> str:
        cleaned = _normalize_text(category)
        cleaned = cleaned.replace("/", " ").replace("-", " ")
        return " ".join(cleaned.split())

    def is_investment(category: str | None) -> bool:
        if not category:
            return False
        raw_key = _normalize_text(category)
        if raw_key in category_settings:
            return category_settings[raw_key]
        key = normalize_key(category)
        if key in category_settings:
            return category_settings[key]
        return _default_is_investment(category)

    def category_bucket(category: str | None) -> str | None:
        if not category:
            return None
        key = normalize_key(category)
        if key.startswith("cash"):
            return "cash"
        if key.startswith("emergency"):
            return "emergency"
        return None

    def ensure_row(date_key: str) -> dict[str, float | str]:
        row = history.get(date_key)
        if not row:
            row = {
                "date": date_key,
                "total": 0.0,
                "cash": 0.0,
                "emergency": 0.0,
                "invested": 0.0,
            }
            history[date_key] = row
        return row

    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT imp.imported_at AS imported_at,
                   items.category AS category,
                   SUM(items.balance) AS total,
                   SUM(COALESCE(items.invested, 0)) AS invested
            FROM santander_imports AS imp
            JOIN santander_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id, items.category
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            date_key = _date_key(row["imported_at"])
            if not date_key:
                continue
            entry = ensure_row(date_key)
            amount = float(row["total"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(row["category"])
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(row["category"]):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   category,
                   value,
                   invested
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            date_key = _date_key(row["date_value"])
            if not date_key:
                continue
            category = row["category"] or "Cash"
            entry = ensure_row(date_key)
            amount = float(row["value"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(category)
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(category):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

        save_import_exists = conn.execute(
            "SELECT 1 FROM save_ngrow_imports WHERE portfolio_id = ? LIMIT 1",
            (portfolio_id,),
        ).fetchone()
        if save_import_exists:
            rows = conn.execute(
                """
                SELECT COALESCE(imp.snapshot_date, imp.created_at) AS date_value,
                       items.category AS category,
                       SUM(items.current_value) AS total,
                       SUM(COALESCE(items.invested, 0)) AS invested
                FROM save_ngrow_imports AS imp
                JOIN save_ngrow_items AS items ON items.import_id = imp.id
                WHERE imp.portfolio_id = ?
                GROUP BY imp.id, items.category
                """,
                (portfolio_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value,
                       current_value AS total,
                       invested AS invested
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
        for row in rows:
            date_key = _date_key(row["date_value"])
            if not date_key:
                continue
            category = row["category"] if "category" in row.keys() else "Retirement Plans"
            entry = ensure_row(date_key)
            amount = float(row["total"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(category)
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(category):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   current_value_total AS total,
                   invested_total AS invested,
                   category
            FROM aforronet_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            date_key = _date_key(row["date_value"])
            if not date_key:
                continue
            category = row["category"] or "Emergency Funds"
            entry = ensure_row(date_key)
            amount = float(row["total"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(category)
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(category):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

        rows = conn.execute(
            """
            SELECT COALESCE(imp.snapshot_date, imp.imported_at) AS date_value,
                   items.category AS category,
                   SUM(items.current_value) AS total,
                   SUM(COALESCE(items.invested, 0)) AS invested
            FROM bancoinvest_imports AS imp
            JOIN bancoinvest_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id, items.category
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            date_key = _date_key(row["date_value"])
            if not date_key:
                continue
            category = row["category"] or "Retirement Plans"
            entry = ensure_row(date_key)
            amount = float(row["total"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(category)
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(category):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

        rows = conn.execute(
            """
            SELECT imported_at, category, current_value, invested
            FROM xtb_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            date_key = _date_key(row["imported_at"])
            if not date_key:
                continue
            category = row["category"] or "Stocks"
            entry = ensure_row(date_key)
            amount = float(row["current_value"] or 0)
            entry["total"] = float(entry["total"]) + amount
            bucket = category_bucket(category)
            if bucket:
                entry[bucket] = float(entry[bucket]) + amount
            if is_investment(category):
                entry["invested"] = float(entry["invested"]) + float(row["invested"] or 0)

    items = []
    for date_key, row in history.items():
        items.append(
            {
                "date": date_key,
                "total": round(float(row["total"]), 2),
                "cash": round(float(row["cash"]), 2),
                "emergency": round(float(row["emergency"]), 2),
                "invested": round(float(row["invested"]), 2),
            }
        )
    items.sort(key=lambda item: item["date"])
    return items


def _aggregate_snapshot_totals(
    portfolio_id: int,
    snapshot_date: str,
    category_settings: dict[str, bool],
) -> tuple[dict[str, float], float, float]:
    totals: dict[str, float] = {}
    total_invested = 0.0
    total_profit = 0.0

    def normalize_key(category: str) -> str:
        cleaned = _normalize_text(category)
        cleaned = cleaned.replace("/", " ").replace("-", " ")
        return " ".join(cleaned.split())

    def is_investment(category: str | None) -> bool:
        if not category:
            return False
        raw_key = _normalize_text(category)
        if raw_key in category_settings:
            return category_settings[raw_key]
        normalized = normalize_key(category)
        if normalized in category_settings:
            return category_settings[normalized]
        return _default_is_investment(category)

    def add_entry(
        category: str | None,
        current_value: float | None,
        invested: float | None = None,
        profit: float | None = None,
    ) -> None:
        nonlocal total_invested, total_profit
        if not category:
            return
        value = float(current_value or 0)
        totals[category] = totals.get(category, 0.0) + value
        if is_investment(category):
            total_invested += float(invested or 0)
            total_profit += float(profit or 0)

    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT imp.imported_at AS imported_at,
                   items.category AS category,
                   SUM(items.balance) AS total,
                   SUM(COALESCE(items.invested, 0)) AS invested,
                   SUM(COALESCE(items.gains, 0)) AS gains
            FROM santander_imports AS imp
            JOIN santander_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id, items.category
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            if _date_key(row["imported_at"]) != snapshot_date:
                continue
            add_entry(row["category"], row["total"], row["invested"], row["gains"])

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   category,
                   value,
                   invested,
                   gains
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            if _date_key(row["date_value"]) != snapshot_date:
                continue
            category = row["category"] or "Cash"
            add_entry(category, row["value"], row["invested"], row["gains"])

        save_imports = conn.execute(
            "SELECT 1 FROM save_ngrow_imports WHERE portfolio_id = ? LIMIT 1",
            (portfolio_id,),
        ).fetchone()
        if save_imports:
            rows = conn.execute(
                """
                SELECT COALESCE(imp.snapshot_date, imp.created_at) AS date_value,
                       items.category AS category,
                       SUM(items.current_value) AS total,
                       SUM(COALESCE(items.invested, 0)) AS invested,
                       SUM(COALESCE(items.profit_value, 0)) AS profit
                FROM save_ngrow_imports AS imp
                JOIN save_ngrow_items AS items ON items.import_id = imp.id
                WHERE imp.portfolio_id = ?
                GROUP BY imp.id, items.category
                """,
                (portfolio_id,),
            ).fetchall()
            for row in rows:
                if _date_key(row["date_value"]) != snapshot_date:
                    continue
                add_entry(row["category"], row["total"], row["invested"], row["profit"])
        else:
            rows = conn.execute(
                """
                SELECT COALESCE(snapshot_date, created_at) AS date_value,
                       current_value AS total,
                       invested,
                       profit_value
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchall()
            for row in rows:
                if _date_key(row["date_value"]) != snapshot_date:
                    continue
                add_entry(
                    "Retirement Plans",
                    row["total"],
                    row["invested"],
                    row["profit_value"],
                )

        rows = conn.execute(
            """
            SELECT COALESCE(snapshot_date, created_at) AS date_value,
                   current_value_total AS total,
                   invested_total AS invested,
                   category
            FROM aforronet_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            if _date_key(row["date_value"]) != snapshot_date:
                continue
            category = row["category"] or "Emergency Funds"
            add_entry(category, row["total"], row["invested"], None)

        rows = conn.execute(
            """
            SELECT COALESCE(imp.snapshot_date, imp.imported_at) AS date_value,
                   items.category AS category,
                   SUM(items.current_value) AS total,
                   SUM(COALESCE(items.invested, 0)) AS invested,
                   SUM(COALESCE(items.gains, 0)) AS gains
            FROM bancoinvest_imports AS imp
            JOIN bancoinvest_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
            GROUP BY imp.id, items.category
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            if _date_key(row["date_value"]) != snapshot_date:
                continue
            category = row["category"] or "Retirement Plans"
            add_entry(category, row["total"], row["invested"], row["gains"])

        rows = conn.execute(
            """
            SELECT imported_at,
                   category,
                   current_value,
                   invested,
                   profit_value
            FROM xtb_imports
            WHERE portfolio_id = ?
            """,
            (portfolio_id,),
        ).fetchall()
        for row in rows:
            if _date_key(row["imported_at"]) != snapshot_date:
                continue
            category = row["category"] or "Stocks"
            add_entry(category, row["current_value"], row["invested"], row["profit_value"])

    return totals, total_invested, total_profit


def _aggregate_latest_totals(
    portfolio_id: int,
    category_settings: dict[str, bool],
) -> tuple[dict[str, float], float, float, float, bool]:
    totals: dict[str, float] = {}
    total_invested = 0.0
    total_profit = 0.0
    investment_current_total = 0.0
    cash_investment = False

    def normalize_key(category: str) -> str:
        cleaned = _normalize_text(category)
        cleaned = cleaned.replace("/", " ").replace("-", " ")
        return " ".join(cleaned.split())

    def is_investment(category: str | None) -> bool:
        if not category:
            return False
        raw_key = _normalize_text(category)
        if raw_key in category_settings:
            return category_settings[raw_key]
        normalized = normalize_key(category)
        if normalized in category_settings:
            return category_settings[normalized]
        return _default_is_investment(category)

    def add_entry(
        category: str | None,
        current_value: float | None,
        invested: float | None = None,
        profit: float | None = None,
    ) -> None:
        nonlocal total_invested, total_profit, investment_current_total, cash_investment
        if not category:
            return
        value = float(current_value or 0)
        totals[category] = totals.get(category, 0.0) + value
        is_cash_profit = _normalize_text(category) == "cash" and profit not in (None, 0)
        if is_cash_profit:
            cash_investment = True
        if is_investment(category) or is_cash_profit:
            total_invested += float(invested or 0)
            total_profit += float(profit or 0)
            investment_current_total += value

    with _db_connection() as conn:
        latest_santander = conn.execute(
            """
            SELECT id
            FROM santander_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if latest_santander:
            rows = conn.execute(
                """
                SELECT category,
                       SUM(balance) AS total,
                       SUM(COALESCE(invested, 0)) AS invested,
                       SUM(COALESCE(gains, 0)) AS gains
                FROM santander_items
                WHERE import_id = ?
                GROUP BY category
                """,
                (latest_santander["id"],),
            ).fetchall()
            for row in rows:
                add_entry(row["category"], row["total"], row["invested"], row["gains"])

        trade_latest = conn.execute(
            """
            SELECT value, invested, gains, category
            FROM trade_republic_entries
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if trade_latest:
            category = trade_latest["category"] or "Cash"
            add_entry(
                category,
                trade_latest["value"],
                trade_latest["invested"],
                trade_latest["gains"],
            )

        save_latest = conn.execute(
            """
            SELECT id
            FROM save_ngrow_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if save_latest:
            rows = conn.execute(
                """
                SELECT category, current_value, invested, profit_value
                FROM save_ngrow_items
                WHERE import_id = ?
                """,
                (save_latest["id"],),
            ).fetchall()
            totals_by_category: dict[str, float] = {}
            invested_by_category: dict[str, float] = {}
            profit_by_category: dict[str, float] = {}
            for row in rows:
                category = row["category"]
                current_value = float(row["current_value"] or 0)
                invested_value = (
                    float(row["invested"]) if row["invested"] is not None else None
                )
                profit_value = _normalize_profit_value(
                    current_value, invested_value, row["profit_value"]
                )
                totals_by_category[category] = (
                    totals_by_category.get(category, 0.0) + current_value
                )
                invested_by_category[category] = (
                    invested_by_category.get(category, 0.0)
                    + float(invested_value or 0)
                )
                profit_by_category[category] = (
                    profit_by_category.get(category, 0.0)
                    + float(profit_value or 0)
                )
            for category, total in totals_by_category.items():
                add_entry(
                    category,
                    total,
                    invested_by_category.get(category, 0.0),
                    profit_by_category.get(category, 0.0),
                )
        else:
            save_entry = conn.execute(
                """
                SELECT current_value, invested, profit_value
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if save_entry:
                normalized_profit = _normalize_profit_value(
                    float(save_entry["current_value"] or 0),
                    float(save_entry["invested"]) if save_entry["invested"] is not None else None,
                    save_entry["profit_value"],
                )
                add_entry(
                    "Retirement Plans",
                    save_entry["current_value"],
                    save_entry["invested"],
                    normalized_profit,
                )

        aforronet_latest = conn.execute(
            """
            SELECT current_value_total, invested_total, category
            FROM aforronet_imports
            WHERE portfolio_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if aforronet_latest:
            current_value = float(aforronet_latest["current_value_total"] or 0)
            invested_value = float(aforronet_latest["invested_total"] or 0)
            profit = current_value - invested_value
            category = aforronet_latest["category"] or "Emergency Funds"
            add_entry(category, current_value, invested_value, profit)

        bancoinvest_import = conn.execute(
            """
            SELECT id
            FROM bancoinvest_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            LIMIT 1
            """,
            (portfolio_id,),
        ).fetchone()
        if bancoinvest_import:
            rows = conn.execute(
                """
                SELECT category,
                       SUM(current_value) AS total,
                       SUM(COALESCE(invested, 0)) AS invested,
                       SUM(COALESCE(gains, 0)) AS gains
                FROM bancoinvest_items
                WHERE import_id = ?
                GROUP BY category
                """,
                (bancoinvest_import["id"],),
            ).fetchall()
            for row in rows:
                add_entry(row["category"], row["total"], row["invested"], row["gains"])

        xtb_rows = conn.execute(
            """
            SELECT account_type, category, current_value, invested, profit_value, imported_at
            FROM xtb_imports
            WHERE portfolio_id = ?
            ORDER BY imported_at DESC
            """,
            (portfolio_id,),
        ).fetchall()
        latest_by_account: dict[str, sqlite3.Row] = {}
        for row in xtb_rows:
            key = row["account_type"]
            if key not in latest_by_account:
                latest_by_account[key] = row
        for row in latest_by_account.values():
            category = row["category"] or "Stocks"
            add_entry(
                category,
                row["current_value"],
                row["invested"],
                row["profit_value"],
            )

    return totals, total_invested, total_profit, investment_current_total, cash_investment


def _first_investment_date(
    portfolio_id: int, investment_categories: set[str]
) -> datetime | None:
    if not investment_categories:
        return None
    categories = [item.lower() for item in investment_categories]
    placeholders = ",".join(["?"] * len(categories))
    earliest = datetime.min

    def update(candidate: str | None) -> None:
        nonlocal earliest
        timestamp = _to_datetime(candidate)
        if timestamp == datetime.min:
            return
        if earliest == datetime.min or timestamp < earliest:
            earliest = timestamp

    with _db_connection() as conn:
        row = conn.execute(
            f"""
            SELECT MIN(imp.imported_at) AS date_value
            FROM santander_imports AS imp
            JOIN santander_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
              AND lower(items.category) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

        row = conn.execute(
            f"""
            SELECT MIN(COALESCE(snapshot_date, created_at)) AS date_value
            FROM trade_republic_entries
            WHERE portfolio_id = ?
              AND lower(COALESCE(category, 'Cash')) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

        row = conn.execute(
            f"""
            SELECT MIN(COALESCE(imp.snapshot_date, imp.created_at)) AS date_value
            FROM save_ngrow_imports AS imp
            JOIN save_ngrow_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
              AND lower(items.category) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

        if "retirement plans" in categories:
            row = conn.execute(
                """
                SELECT MIN(COALESCE(snapshot_date, created_at)) AS date_value
                FROM save_ngrow_entries
                WHERE portfolio_id = ?
                """,
                (portfolio_id,),
            ).fetchone()
            update(row["date_value"] if row else None)

        row = conn.execute(
            f"""
            SELECT MIN(COALESCE(snapshot_date, created_at)) AS date_value
            FROM aforronet_imports
            WHERE portfolio_id = ?
              AND lower(COALESCE(category, 'Emergency Funds')) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

        row = conn.execute(
            f"""
            SELECT MIN(COALESCE(imp.snapshot_date, imp.imported_at)) AS date_value
            FROM bancoinvest_imports AS imp
            JOIN bancoinvest_items AS items ON items.import_id = imp.id
            WHERE imp.portfolio_id = ?
              AND lower(items.category) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

        row = conn.execute(
            f"""
            SELECT MIN(imported_at) AS date_value
            FROM xtb_imports
            WHERE portfolio_id = ?
              AND lower(COALESCE(category, 'Stocks')) IN ({placeholders})
            """,
            (portfolio_id, *categories),
        ).fetchone()
        update(row["date_value"] if row else None)

    return None if earliest == datetime.min else earliest


def _calculate_irr(
    portfolio_id: int,
    investment_categories: set[str],
    total_invested: float,
    current_value: float,
) -> float | None:
    if total_invested <= 0 or current_value <= 0:
        return None
    start_date = _first_investment_date(portfolio_id, investment_categories)
    if not start_date:
        return None
    days = (datetime.utcnow() - start_date).days
    if days <= 0:
        return None
    years = days / 365.25
    if years <= 0:
        return None
    try:
        irr = (current_value / total_invested) ** (1 / years) - 1
    except (ZeroDivisionError, ValueError):
        return None
    return irr * 100


def _list_institution_detail(portfolio_id: int, institution: str) -> dict:
    key = _institution_key(institution)
    with _db_connection() as conn:
        if key == "santander":
            latest_import = conn.execute(
                """
                SELECT id, imported_at, filename
                FROM santander_imports
                WHERE portfolio_id = ?
                ORDER BY imported_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if not latest_import:
                raise HTTPException(status_code=404, detail="No Santander data.")
            rows = conn.execute(
                """
                SELECT account, description, balance, category, invested, gains
                FROM santander_items
                WHERE import_id = ?
                """,
                (latest_import["id"],),
            ).fetchall()
            entries = [
                {
                    "label": row["account"],
                    "description": row["description"],
                    "current_value": float(row["balance"] or 0),
                    "invested": row["invested"],
                    "gains": row["gains"],
                    "category": row["category"],
                }
                for row in rows
            ]
            totals_by_category: dict[str, float] = {}
            for entry in entries:
                category = entry["category"]
                totals_by_category[category] = totals_by_category.get(category, 0.0) + float(
                    entry["current_value"] or 0
                )
            return {
                "institution": "Santander",
                "source": latest_import["filename"],
                "date": latest_import["imported_at"],
                "total": round(sum(totals_by_category.values()), 2),
                "totals_by_category": totals_by_category,
                "entries": entries,
            }

        if key == "traderepublic":
            latest = conn.execute(
                """
                SELECT available_cash,
                       interests_received,
                       invested,
                       value,
                       gains,
                       currency,
                       category,
                       source,
                       source_file,
                       snapshot_date,
                       created_at
                FROM trade_republic_entries
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if not latest:
                raise HTTPException(status_code=404, detail="No Trade Republic data.")
            category = latest["category"] or "Cash"
            source_file = latest["source_file"]
            source_label = source_file or latest["source"]
            entry = {
                "label": "Available cash",
                "description": source_label,
                "current_value": float(latest["value"] or 0),
                "invested": float(latest["invested"] or 0),
                "gains": float(latest["gains"] or 0),
                "category": category,
            }
            return {
                "institution": "Trade Republic",
                "source": source_label,
                "date": latest["snapshot_date"] or latest["created_at"],
                "total": round(entry["current_value"], 2),
                "totals_by_category": {category: round(entry["current_value"], 2)},
                "entries": [entry],
            }

        if key == "savengrow":
            latest_import = conn.execute(
                """
                SELECT id, source_file, snapshot_date, created_at
                FROM save_ngrow_imports
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            entries: list[dict] = []
            source = None
            date_value = None
            if latest_import:
                rows = conn.execute(
                    """
                    SELECT item_name, invested, current_value, profit_value, profit_percent, category
                    FROM save_ngrow_items
                    WHERE import_id = ?
                    """,
                    (latest_import["id"],),
                ).fetchall()
                entries = [
                    {
                        "label": row["item_name"],
                        "description": None,
                        "current_value": float(row["current_value"] or 0),
                        "invested": row["invested"],
                        "gains": _normalize_profit_value(
                            float(row["current_value"] or 0),
                            float(row["invested"]) if row["invested"] is not None else None,
                            row["profit_value"],
                        ),
                        "category": row["category"],
                    }
                    for row in rows
                ]
                source = latest_import["source_file"]
                date_value = latest_import["snapshot_date"] or latest_import["created_at"]
            else:
                latest_entry = conn.execute(
                    """
                    SELECT invested, current_value, profit_value, profit_percent, source_file, snapshot_date, created_at
                    FROM save_ngrow_entries
                    WHERE portfolio_id = ?
                    ORDER BY created_at DESC
                    LIMIT 1
                    """,
                    (portfolio_id,),
                ).fetchone()
                if latest_entry:
                    entries = [
                        {
                            "label": "Save N Grow",
                            "description": None,
                            "current_value": float(latest_entry["current_value"] or 0),
                            "invested": latest_entry["invested"],
                            "gains": _normalize_profit_value(
                                float(latest_entry["current_value"] or 0),
                                float(latest_entry["invested"])
                                if latest_entry["invested"] is not None
                                else None,
                                latest_entry["profit_value"],
                            ),
                            "category": "Retirement Plans",
                        }
                    ]
                    source = latest_entry["source_file"]
                    date_value = latest_entry["snapshot_date"] or latest_entry["created_at"]
            if not entries:
                raise HTTPException(status_code=404, detail="No Save N Grow data.")
            totals_by_category: dict[str, float] = {}
            for entry in entries:
                category = entry["category"]
                totals_by_category[category] = totals_by_category.get(category, 0.0) + float(
                    entry["current_value"] or 0
                )
            return {
                "institution": "Save N Grow",
                "source": source,
                "date": date_value,
                "total": round(sum(totals_by_category.values()), 2),
                "totals_by_category": totals_by_category,
                "entries": entries,
            }

        if key == "aforronet":
            latest_import = conn.execute(
                """
                SELECT id, source_file, snapshot_date, created_at, category, invested_total, current_value_total
                FROM aforronet_imports
                WHERE portfolio_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if not latest_import:
                raise HTTPException(status_code=404, detail="No AforroNet data.")
            entries = [
                {
                    "label": "AforroNet",
                    "description": None,
                    "current_value": float(latest_import["current_value_total"] or 0),
                    "invested": float(latest_import["invested_total"] or 0),
                    "gains": None,
                    "category": latest_import["category"] or "Emergency Funds",
                }
            ]
            totals_by_category = {
                entries[0]["category"]: float(entries[0]["current_value"] or 0)
            }
            return {
                "institution": "AforroNet",
                "source": latest_import["source_file"],
                "date": latest_import["snapshot_date"] or latest_import["created_at"],
                "total": round(sum(totals_by_category.values()), 2),
                "totals_by_category": totals_by_category,
                "entries": entries,
            }

        if key == "bancoinvest":
            latest_import = conn.execute(
                """
                SELECT id, source_file, snapshot_date, imported_at
                FROM bancoinvest_imports
                WHERE portfolio_id = ?
                ORDER BY imported_at DESC
                LIMIT 1
                """,
                (portfolio_id,),
            ).fetchone()
            if not latest_import:
                raise HTTPException(status_code=404, detail="No BancoInvest data.")
            rows = conn.execute(
                """
                SELECT holder, invested, current_value, gains, category
                FROM bancoinvest_items
                WHERE import_id = ?
                """,
                (latest_import["id"],),
            ).fetchall()
            entries = [
                {
                    "label": row["holder"],
                    "description": None,
                    "current_value": float(row["current_value"] or 0),
                    "invested": row["invested"],
                    "gains": row["gains"],
                    "category": row["category"],
                }
                for row in rows
            ]
            totals_by_category: dict[str, float] = {}
            for entry in entries:
                category = entry["category"]
                totals_by_category[category] = totals_by_category.get(category, 0.0) + float(
                    entry["current_value"] or 0
                )
            return {
                "institution": "BancoInvest",
                "source": latest_import["source_file"],
                "date": latest_import["snapshot_date"] or latest_import["imported_at"],
                "total": round(sum(totals_by_category.values()), 2),
                "totals_by_category": totals_by_category,
                "entries": entries,
            }

        if key == "xtb":
            rows = conn.execute(
                """
                SELECT account_type, category, current_value, invested, profit_value, source_file, imported_at
                FROM xtb_imports
                WHERE portfolio_id = ?
                ORDER BY imported_at DESC
                """,
                (portfolio_id,),
            ).fetchall()
            latest_by_account: dict[str, sqlite3.Row] = {}
            for row in rows:
                account = row["account_type"]
                if account not in latest_by_account:
                    latest_by_account[account] = row
            entries = [
                {
                    "label": row["account_type"],
                    "description": row["source_file"],
                    "current_value": float(row["current_value"] or 0),
                    "invested": row["invested"],
                    "gains": row["profit_value"],
                    "category": row["category"] or "Stocks",
                }
                for row in latest_by_account.values()
            ]
            if not entries:
                raise HTTPException(status_code=404, detail="No XTB data.")
            totals_by_category: dict[str, float] = {}
            for entry in entries:
                category = entry["category"]
                totals_by_category[category] = totals_by_category.get(category, 0.0) + float(
                    entry["current_value"] or 0
                )
            return {
                "institution": "XTB",
                "source": "latest",
                "date": rows[0]["imported_at"] if rows else None,
                "total": round(sum(totals_by_category.values()), 2),
                "totals_by_category": totals_by_category,
                "entries": entries,
            }

    raise HTTPException(status_code=404, detail="Institution not found.")


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.post("/auth/register")
def register(payload: RegisterRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password too short.")

    existing = _get_user(email)
    if existing and existing["verified"]:
        raise HTTPException(status_code=409, detail="User already exists.")

    salt = existing["salt"] if existing else secrets.token_hex(12)
    password_hash = _hash_password(payload.password, salt)
    created_at = existing["created_at"] if existing else datetime.utcnow().isoformat()
    _save_user(email, salt, password_hash, False, created_at)
    code = _issue_code(email)
    text, html = _verification_email(code)
    _send_email_async(email, "MyFAInance verification code", text, html)
    return {
        "status": "verification_required",
        "message": "Verification code sent.",
    }


@app.post("/auth/login")
def login(payload: LoginRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    user = _get_user(email)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")

    expected = _hash_password(payload.password, user["salt"])
    if expected != user["password_hash"]:
        raise HTTPException(status_code=401, detail="Invalid credentials.")

    if not user["verified"]:
        code = _issue_code(email)
        text, html = _verification_email(code)
        _send_email_async(email, "MyFAInance verification code", text, html)
        return {
            "status": "verification_required",
            "message": "Email verification required.",
        }

    token = _issue_session(email)
    return {"status": "ok", "token": token}


@app.post("/auth/verify")
def verify(payload: VerifyRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    _verify_code(email, payload.code.strip())
    user = _get_user(email)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    _set_verified(email, True)
    return {"status": "verified"}


@app.post("/auth/resend-code")
def resend_code(payload: ResendRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    if not _get_user(email):
        raise HTTPException(status_code=404, detail="User not found.")
    code = _issue_code(email)
    text, html = _verification_email(code)
    _send_email_async(email, "MyFAInance verification code", text, html)
    return {"status": "code_sent"}


@app.post("/auth/password/reset-request")
def reset_request(payload: ResendRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    if _get_user(email):
        code = _issue_reset_code(email)
        text, html = _reset_email(code)
        _send_email_async(email, "MyFAInance password reset", text, html)
    return {"status": "ok"}


class ResetPasswordRequest(BaseModel):
    email: str
    code: str
    new_password: str


@app.post("/auth/password/reset")
def reset_password(payload: ResetPasswordRequest) -> dict:
    email = payload.email.strip().lower()
    _validate_email(email)
    if not _get_user(email):
        raise HTTPException(status_code=404, detail="User not found.")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password too short.")
    _verify_reset_code(email, payload.code.strip())
    salt = secrets.token_hex(12)
    _set_password(email, salt, _hash_password(payload.new_password, salt))
    _delete_sessions_for_email(email)
    return {"status": "reset"}


@app.get("/auth/me")
def auth_me(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    user = _get_user(session["email"])
    if not user:
        raise HTTPException(status_code=401, detail="User not found.")
    return {
        "status": "ok",
        "email": user["email"],
        "verified": bool(user["verified"]),
        "created_at": user["created_at"],
        "expires_at": session["expires_at"],
    }


@app.post("/auth/logout")
def logout(authorization: str | None = Header(default=None)) -> dict:
    token = _require_token(authorization)
    _delete_session(token)
    return {"status": "logged_out"}


@app.get("/profile")
def get_profile(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    return {"age": _get_profile_age(session["email"])}


@app.post("/profile")
def update_profile(payload: ProfileUpdateRequest, authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    age_value = payload.age
    if age_value is not None and age_value <= 0:
        raise HTTPException(status_code=400, detail="Age must be greater than 0.")
    _set_profile_age(session["email"], age_value)
    return {"status": "saved", "age": age_value}


@app.get("/debts")
def list_debts(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    return {"items": _list_debts(session["email"])}


@app.post("/debts")
def create_debt(payload: DebtRequest, authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    if payload.original_amount <= 0:
        raise HTTPException(status_code=400, detail="Original amount must be greater than 0.")
    if payload.current_balance < 0:
        raise HTTPException(status_code=400, detail="Current balance cannot be negative.")
    if payload.monthly_payment <= 0:
        raise HTTPException(status_code=400, detail="Monthly payment must be greater than 0.")
    debt = _create_debt(session["email"], payload)
    return {"status": "saved", "debt": debt}


@app.put("/debts/{debt_id}")
def update_debt(
    debt_id: int,
    payload: DebtRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if payload.original_amount <= 0:
        raise HTTPException(status_code=400, detail="Original amount must be greater than 0.")
    if payload.current_balance < 0:
        raise HTTPException(status_code=400, detail="Current balance cannot be negative.")
    if payload.monthly_payment <= 0:
        raise HTTPException(status_code=400, detail="Monthly payment must be greater than 0.")
    debt = _update_debt(session["email"], debt_id, payload)
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found.")
    return {"status": "saved", "debt": debt}


@app.delete("/debts/{debt_id}")
def delete_debt(debt_id: int, authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    if not _delete_debt(session["email"], debt_id):
        raise HTTPException(status_code=404, detail="Debt not found.")
    return {"status": "deleted"}


@app.get("/goals")
def list_goals(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    return {"items": _list_goals(session["email"])}


@app.post("/goals")
def create_goal(payload: GoalCreateRequest, authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Goal name is required.")
    goal = _create_goal(session["email"], name)
    return {"status": "saved", "goal": goal}


@app.put("/goals/{goal_id}")
def update_goal(
    goal_id: int,
    payload: GoalUpdateRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Goal name is required.")
    goal = _update_goal_name(session["email"], goal_id, name)
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found.")
    return {"status": "saved", "goal": goal}


@app.delete("/goals/{goal_id}")
def delete_goal(goal_id: int, authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    if not _delete_goal(session["email"], goal_id):
        raise HTTPException(status_code=404, detail="Goal not found.")
    return {"status": "deleted"}


@app.get("/goals/{goal_id}")
def get_goal(
    goal_id: int,
    portfolio_id: int | None = None,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    goal = _get_goal(goal_id, session["email"])
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found.")
    summary = _goal_summary(session["email"], goal_id, portfolio_id)
    return {
        "goal": {
            "id": goal["id"],
            "name": goal["name"],
            "is_default": bool(goal["is_default"]),
            "created_at": goal["created_at"],
            "updated_at": goal["updated_at"],
        },
        **summary,
    }


@app.post("/goals/{goal_id}/inputs")
def update_goal_inputs(
    goal_id: int,
    payload: GoalInputRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_goal(goal_id, session["email"]):
        raise HTTPException(status_code=404, detail="Goal not found.")
    inputs = _update_goal_inputs(goal_id, payload)
    return {"status": "saved", "inputs": inputs}


@app.post("/goals/{goal_id}/contributions")
def add_goal_contribution(
    goal_id: int,
    payload: GoalContributionRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_goal(goal_id, session["email"]):
        raise HTTPException(status_code=404, detail="Goal not found.")
    item = _add_goal_contribution(goal_id, payload)
    return {"status": "saved", "contribution": item}


@app.delete("/goals/{goal_id}/contributions/{contribution_id}")
def delete_goal_contribution(
    goal_id: int,
    contribution_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_goal(goal_id, session["email"]):
        raise HTTPException(status_code=404, detail="Goal not found.")
    if not _delete_goal_contribution(goal_id, contribution_id):
        raise HTTPException(status_code=404, detail="Contribution not found.")
    return {"status": "deleted"}


@app.get("/portfolios")
def list_portfolios(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    items = _list_portfolios(session["email"])
    return {"items": items}


@app.get("/portfolios/aggregated/summary")
def aggregated_portfolio_summary(authorization: str | None = Header(default=None)) -> dict:
    """Retorna o sumário agregado de todos os portfolios do utilizador."""
    session = _require_session(authorization)
    portfolios = _list_portfolios(session["email"])
    
    if not portfolios:
        return {
            "totals_by_category": {},
            "total": 0.0,
            "total_invested": 0.0,
            "total_profit": 0.0,
            "profit_percent": 0.0,
            "irr": None,
        }
    
    # Aggregate totals from all portfolios
    all_totals = {}
    total_invested_all = 0.0
    total_profit_all = 0.0
    
    for portfolio in portfolios:
        portfolio_id = portfolio["id"]
        categories = portfolio["categories"]  # Already filtered by _list_portfolios
        _ensure_category_settings(portfolio_id, categories)
        settings = _get_category_settings(portfolio_id)
        settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
        
        (
            totals,
            total_invested,
            total_profit,
            investment_current_total,
            cash_investment,
        ) = _aggregate_latest_totals(portfolio_id, settings_lookup)
        
        if totals:
            for category, value in totals.items():
                all_totals[category] = all_totals.get(category, 0.0) + value
            total_invested_all += total_invested or 0.0
            total_profit_all += total_profit or 0.0
    
    total_value = sum(all_totals.values())
    profit_percent = (
        round((total_profit_all / total_invested_all) * 100, 2)
        if total_invested_all
        else 0.0
    )
    
    return {
        "totals_by_category": all_totals,
        "total": round(total_value, 2),
        "total_invested": round(total_invested_all, 2),
        "total_profit": round(total_profit_all, 2),
        "profit_percent": profit_percent,
        "irr": None,  # IRR calculation would need all portfolio transactions
    }


@app.post("/portfolios/aggregated/snapshot")
def create_aggregated_snapshot(authorization: str | None = Header(default=None)) -> dict:
    """Cria um snapshot do portfolio agregado para tracking de evolução."""
    session = _require_session(authorization)
    portfolios = _list_portfolios(session["email"])
    
    if not portfolios:
        raise HTTPException(status_code=400, detail="No portfolios found.")
    
    # Get aggregated summary
    summary = aggregated_portfolio_summary(authorization)
    
    now = datetime.utcnow().isoformat()
    snapshot_date = datetime.utcnow().date().isoformat()
    
    with _db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO aggregated_snapshots (
                owner_email,
                snapshot_date,
                total_value,
                total_invested,
                total_profit,
                profit_percent,
                totals_by_category_json,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session["email"],
                snapshot_date,
                summary["total"],
                summary["total_invested"],
                summary["total_profit"],
                summary["profit_percent"],
                json.dumps(summary["totals_by_category"]),
                now,
            ),
        )
        snapshot_id = cursor.lastrowid
    
    return {
        "status": "created",
        "snapshot_id": snapshot_id,
        "snapshot_date": snapshot_date,
        "total_value": summary["total"],
    }


@app.get("/portfolios/aggregated/snapshots")
def list_aggregated_snapshots(authorization: str | None = Header(default=None)) -> dict:
    """Lista todos os snapshots do portfolio agregado."""
    session = _require_session(authorization)
    
    with _db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, snapshot_date, total_value, total_invested, 
                   total_profit, profit_percent, created_at
            FROM aggregated_snapshots
            WHERE owner_email = ?
            ORDER BY snapshot_date DESC
            """,
            (session["email"],),
        ).fetchall()
    
    items = [dict(row) for row in rows]
    return {"items": items}


@app.get("/portfolios/{portfolio_id}/summary")
def portfolio_summary(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
    investment_categories = {
        category.lower()
        for category in categories
        if settings_lookup.get(_normalize_text(category), _default_is_investment(category))
    }

    def is_investment_category(category: str) -> bool:
        return settings_lookup.get(_normalize_text(category), _default_is_investment(category))

    (
        totals,
        total_invested,
        total_profit,
        investment_current_total,
        cash_investment,
    ) = _aggregate_latest_totals(portfolio_id, settings_lookup)
    if not totals:
        return {
            "totals_by_category": {},
            "total": 0.0,
            "total_invested": 0.0,
            "total_profit": 0.0,
            "profit_percent": 0.0,
            "irr": None,
        }
    total_value = sum(totals.values())
    if cash_investment:
        investment_categories.add("cash")
    investment_total = (
        investment_current_total
        if investment_current_total
        else sum(
            value
            for category, value in totals.items()
            if is_investment_category(category)
        )
    )
    irr = _calculate_irr(
        portfolio_id, investment_categories, total_invested, investment_total
    )
    profit_percent = (total_profit / total_value * 100) if total_value else 0.0
    return {
        "totals_by_category": totals,
        "total": total_value,
        "total_invested": total_invested,
        "total_profit": total_profit,
        "profit_percent": profit_percent,
        "irr": irr,
    }


@app.get("/portfolios/{portfolio_id}/history")
def portfolio_history(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
    return {"items": _list_portfolio_history(portfolio_id, settings_lookup)}


@app.get("/portfolios/{portfolio_id}/history/monthly")
def portfolio_history_monthly(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_portfolio_monthly_history(portfolio_id)}


@app.get("/portfolios/{portfolio_id}/institutions")
def portfolio_institutions(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
    return {"items": _list_institutions(portfolio_id, settings_lookup)}


@app.get("/portfolios/{portfolio_id}/holdings")
def portfolio_holdings(
    portfolio_id: int,
    category: str | None = None,
    institution: str | None = None,
    ticker: str | None = None,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
    return _list_holdings_for_portfolio(
        portfolio_id,
        settings_lookup,
        category=category,
        institution=institution,
        ticker=ticker,
    )


@app.get("/holdings")
def holdings_overall(
    category: str | None = None,
    institution: str | None = None,
    ticker: str | None = None,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolios = _list_portfolios(session["email"])
    all_items: list[dict] = []
    total_value = 0.0
    for portfolio in portfolios:
        categories = portfolio.get("categories") or list(DEFAULT_CATEGORIES)
        categories = _filter_categories(categories)
        _ensure_category_settings(portfolio["id"], categories)
        settings = _get_category_settings(portfolio["id"])
        settings_lookup = {
            _normalize_text(key): value for key, value in settings.items()
        }
        data = _list_holdings_for_portfolio(
            portfolio["id"],
            settings_lookup,
            category=category,
            institution=institution,
            ticker=ticker,
        )
        for item in data["items"]:
            item["portfolio_id"] = portfolio["id"]
            item["portfolio_name"] = portfolio["name"]
            all_items.append(item)
            total_value += float(item["current_value"] or 0)
    for item in all_items:
        item["share_percent"] = (
            round(float(item["current_value"] or 0) / total_value * 100, 2)
            if total_value
            else 0.0
        )
    return {"items": all_items, "total_value": round(total_value, 2)}


@app.get("/portfolios/{portfolio_id}/holdings/transactions")
def holdings_transactions(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_holding_transactions(portfolio_id)}


@app.get("/portfolios/{portfolio_id}/holdings/operations")
def holdings_operations(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_holdings_operations(portfolio_id)}


@app.post("/portfolios/{portfolio_id}/holdings/transactions")
def create_holding_transaction(
    portfolio_id: int,
    payload: HoldingTransactionRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.ticker.strip():
        raise HTTPException(status_code=400, detail="Ticker is required.")
    operation = payload.operation.strip().lower()
    if operation not in {"buy", "sell"}:
        raise HTTPException(status_code=400, detail="Operation must be Buy or Sell.")
    if payload.shares <= 0:
        raise HTTPException(status_code=400, detail="Shares must be greater than 0.")
    if payload.price <= 0:
        raise HTTPException(status_code=400, detail="Price must be greater than 0.")
    if not _date_key(payload.trade_date):
        raise HTTPException(status_code=400, detail="Invalid trade date.")
    payload.operation = operation
    payload.category = payload.category or "Stocks"
    meta = _save_holding_transaction(portfolio_id, payload)
    return {"status": "created", "item": meta}


@app.post("/portfolios/{portfolio_id}/holdings/metadata")
def upsert_holding_metadata(
    portfolio_id: int,
    payload: HoldingMetadataRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.ticker.strip():
        raise HTTPException(status_code=400, detail="Ticker is required.")
    item = _upsert_holdings_metadata(portfolio_id, payload)
    return {"status": "saved", "item": item}


@app.get("/holdings/tags")
def list_holding_tags(authorization: str | None = Header(default=None)) -> dict:
    session = _require_session(authorization)
    return _list_investment_tags(session["email"])


@app.post("/holdings/tags")
def create_holding_tag(
    payload: TagRequest, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    item = _save_investment_tag(session["email"], payload.name)
    return {"status": "created", "item": item}


@app.delete("/holdings/tags/{tag_name}")
def delete_holding_tag(
    tag_name: str, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    _delete_investment_tag(session["email"], urllib.parse.unquote(tag_name))
    return {"status": "deleted"}


@app.post("/portfolios/{portfolio_id}/holdings/refresh-prices")
def refresh_holding_prices(
    portfolio_id: int,
    payload: PriceRefreshRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
    holdings = _list_holdings_for_portfolio(portfolio_id, settings_lookup)
    tickers = payload.tickers or [item["ticker"] for item in holdings["items"]]
    results: list[dict] = []
    cache = _get_cached_prices(tickers)
    
    import time
    
    # Rate limiting: max 59 calls/min = 1 call per 1.02 seconds
    rate_limit_delay = 1.05  # slightly over 1 second to be safe
    
    for idx, ticker in enumerate(tickers):
        cached = cache.get(ticker.upper())
        if cached and _price_is_fresh(cached["updated_at"]) and not payload.force:
            results.append(
                {
                    "ticker": ticker,
                    "status": "cached",
                    "price": cached["price"],
                    "progress": int((idx + 1) / len(tickers) * 100)
                }
            )
            continue
        
        try:
            price_value = _fetch_latest_price(ticker)
            _upsert_price(ticker, price_value)
            results.append(
                {"ticker": ticker, "status": "updated", "price": price_value, "progress": int((idx + 1) / len(tickers) * 100)}
            )
        except HTTPException as exc:
            results.append({"ticker": ticker, "status": "error", "error": exc.detail, "progress": int((idx + 1) / len(tickers) * 100)})
        
        # Rate limiting: wait between API calls (not for cached)
        if idx < len(tickers) - 1 and not cached:
            time.sleep(rate_limit_delay)
    
    return {"items": results}


@app.post("/holdings/refresh-prices")
def refresh_holding_prices_overall(
    payload: PriceRefreshRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolios = _list_portfolios(session["email"])
    tickers: list[str] = []
    for portfolio in portfolios:
        categories = portfolio.get("categories") or list(DEFAULT_CATEGORIES)
        categories = _filter_categories(categories)
        _ensure_category_settings(portfolio["id"], categories)
        settings = _get_category_settings(portfolio["id"])
        settings_lookup = {_normalize_text(key): value for key, value in settings.items()}
        data = _list_holdings_for_portfolio(portfolio["id"], settings_lookup)
        tickers.extend([item["ticker"] for item in data["items"]])
    unique_tickers = sorted({ticker.upper() for ticker in tickers})
    if payload.tickers:
        unique_tickers = [ticker.upper() for ticker in payload.tickers]
    results: list[dict] = []
    cache = _get_cached_prices(unique_tickers)
    
    import time
    
    # Rate limiting: max 59 calls/min = 1 call per 1.02 seconds
    rate_limit_delay = 1.05
    
    for idx, ticker in enumerate(unique_tickers):
        cached = cache.get(ticker.upper())
        if cached and _price_is_fresh(cached["updated_at"]) and not payload.force:
            results.append(
                {
                    "ticker": ticker,
                    "status": "cached",
                    "price": cached["price"],
                    "progress": int((idx + 1) / len(unique_tickers) * 100)
                }
            )
            continue
        try:
            price_value = _fetch_latest_price(ticker)
            _upsert_price(ticker, price_value)
            results.append(
                {"ticker": ticker, "status": "updated", "price": price_value, "progress": int((idx + 1) / len(unique_tickers) * 100)}
            )
        except HTTPException as exc:
            results.append({"ticker": ticker, "status": "error", "error": exc.detail, "progress": int((idx + 1) / len(unique_tickers) * 100)})
        
        # Rate limiting: wait between API calls (not for cached)
        if idx < len(unique_tickers) - 1 and not cached:
            time.sleep(rate_limit_delay)
    
    return {"items": results}


@app.get("/portfolios/{portfolio_id}/banking/categories")
def list_banking_categories(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    _ensure_banking_categories(portfolio_id)
    return {"items": _list_banking_categories(portfolio_id)}


@app.get("/portfolios/{portfolio_id}/banking/institutions")
def list_banking_institutions(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_banking_institutions(portfolio_id)}


@app.post("/portfolios/{portfolio_id}/banking/preview")
def banking_preview(
    portfolio_id: int,
    authorization: str | None = Header(default=None),
    institution: str | None = Form(default=None),
    text: str | None = Form(default=None),
    file: UploadFile | None = File(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not file and not text:
        raise HTTPException(status_code=400, detail="Provide a file or pasted data.")
    _ensure_banking_categories(portfolio_id)
    file_bytes = file.file.read() if file else None
    filename = file.filename if file else "pasted.csv"
    content_hash = hashlib.sha256(file_bytes or text.encode("utf-8")).hexdigest()
    rows = _load_banking_rows(file_bytes, filename, text)
    if not rows:
        raise HTTPException(status_code=400, detail="No rows found in the file.")
    columns, preview_rows, mapping, warnings = _build_banking_preview(
        rows, portfolio["currency"]
    )
    return {
        "source_file": filename,
        "file_hash": content_hash,
        "institution": institution,
        "columns": columns,
        "rows": preview_rows,
        "mapping": mapping,
        "warnings": warnings,
    }


@app.post("/portfolios/{portfolio_id}/banking/commit")
def banking_commit(
    portfolio_id: int,
    payload: BankingCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    institution = payload.institution.strip() if payload.institution else ""
    if not institution:
        raise HTTPException(status_code=400, detail="Institution is required.")
    _ensure_banking_categories(portfolio_id)
    rows = [row.cells for row in payload.rows if row.include]
    items, warnings = _build_banking_items(
        rows, payload.columns, payload.mapping, portfolio["currency"]
    )
    if not items:
        raise HTTPException(status_code=400, detail="No valid rows to import.")
    _apply_banking_category_rules(portfolio_id, institution, items)
    try:
        import_id = _save_banking_import(
            portfolio_id, institution, payload.source_file, payload.file_hash, len(items)
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="File already imported.") from exc
    _save_banking_transactions(portfolio_id, import_id, institution, items)
    _upsert_banking_institution(portfolio_id, institution)
    return {"status": "imported", "import_id": import_id, "warnings": warnings}


@app.get("/portfolios/{portfolio_id}/banking/transactions")
def list_banking_transactions(
    portfolio_id: int,
    authorization: str | None = Header(default=None),
    month: str | None = None,
    category: str | None = None,
    subcategory: str | None = None,
    institution: str | None = None,
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    items = _list_banking_transactions(
        portfolio_id, month=month, category=category, subcategory=subcategory, institution=institution
    )
    return {"items": items}


@app.post("/portfolios/{portfolio_id}/banking/transactions/{tx_id}/category")
def update_banking_transaction_category(
    portfolio_id: int,
    tx_id: int,
    payload: BankingCategoryUpdateRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    _ensure_banking_categories(portfolio_id)
    result = _update_banking_transaction_category(
        portfolio_id, tx_id, payload.category, payload.subcategory
    )
    return {"status": "updated", "transaction": result}


@app.post("/portfolios/{portfolio_id}/banking/clear")
def clear_banking_transactions(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    result = _clear_banking_transactions(portfolio_id)
    return {"status": "cleared", "counts": result}


@app.get("/portfolios/{portfolio_id}/banking/budgets")
def list_banking_budgets(
    portfolio_id: int,
    month: str | None = None,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    month_value = month or datetime.utcnow().strftime("%Y-%m")
    return {"items": _list_banking_budgets(portfolio_id, month_value)}


@app.post("/portfolios/{portfolio_id}/banking/budgets")
def upsert_banking_budget(
    portfolio_id: int,
    payload: BankingBudgetRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    category = (payload.category or "").strip()
    if not category:
        raise HTTPException(status_code=400, detail="Category is required.")
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0.")
    month_value = payload.month or datetime.utcnow().strftime("%Y-%m")
    _upsert_banking_category(portfolio_id, category, None)
    budget = _upsert_banking_budget(portfolio_id, category, month_value, payload.amount)
    return {"status": "saved", "budget": budget}


@app.delete("/portfolios/{portfolio_id}/banking/budgets/{budget_id}")
def delete_banking_budget(
    portfolio_id: int,
    budget_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not _delete_banking_budget(portfolio_id, budget_id):
        raise HTTPException(status_code=404, detail="Budget not found.")
    return {"status": "deleted"}


@app.get("/portfolios/{portfolio_id}/institutions/{institution}/detail")
def portfolio_institution_detail(
    portfolio_id: int,
    institution: str,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return _list_institution_detail(portfolio_id, institution)


@app.post("/portfolios")
def create_portfolio(
    payload: PortfolioCreateRequest, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Portfolio name required.")
    currency = payload.currency.strip().upper()
    if currency not in ALLOWED_CURRENCIES:
        raise HTTPException(status_code=400, detail="Unsupported currency.")
    # Combine DEFAULT_CATEGORIES with custom categories
    custom_categories = _normalize_categories(payload.custom_categories)
    categories = _filter_categories(custom_categories)
    try:
        portfolio = _create_portfolio(session["email"], name, currency, categories)
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Portfolio already exists.") from exc
    return {"status": "created", "portfolio": portfolio}


@app.post("/portfolios/{portfolio_id}/clear-data")
def clear_portfolio_data(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    _clear_portfolio_data(portfolio_id)
    return {"status": "cleared"}


@app.delete("/portfolios/{portfolio_id}")
def delete_portfolio(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    deleted = _delete_portfolio(portfolio_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"status": "deleted"}


@app.get("/portfolios/{portfolio_id}/imports/trade-republic/manual")
def list_trade_republic_manual(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_trade_republic_entries(portfolio_id, source="manual")}


@app.post("/portfolios/{portfolio_id}/imports/trade-republic/manual")
def create_trade_republic_manual(
    portfolio_id: int,
    payload: TradeRepublicManualRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    currency = payload.currency or portfolio["currency"]
    category = (
        payload.category.strip()
        if payload.category and payload.category.strip()
        else (_latest_trade_republic_category(portfolio_id) or "Cash")
    )
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    if category.lower() not in allowed:
        categories.append(category)
        _update_portfolio_categories(portfolio_id, categories)
    entry = _build_trade_republic_entry(
        payload.available_cash,
        payload.interests_received,
        currency,
        category=category,
        source="manual",
    )
    meta = _save_trade_republic_entry(portfolio_id, entry)
    return {
        "status": "created",
        "entry": {
            **entry,
            "id": meta["id"],
            "created_at": meta["created_at"],
        },
    }


@app.post("/portfolios/{portfolio_id}/imports/trade-republic/preview")
def trade_republic_preview(
    portfolio_id: int,
    files: list[UploadFile] = File(...),
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not files:
        raise HTTPException(status_code=400, detail="Files required.")
    default_category = _latest_trade_republic_category(portfolio_id) or "Cash"
    items: list[dict] = []
    for file in files:
        if not file.filename:
            continue
        if not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Invalid file type.")
        file_bytes = file.file.read()
        parsed = _parse_trade_republic_pdf(file_bytes, file.filename)
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        entry = _build_trade_republic_entry(
            parsed["available_cash"],
            parsed["interests_received"],
            portfolio["currency"],
            category=default_category,
            source="file",
            source_file=file.filename,
            file_hash=file_hash,
            snapshot_date=parsed["snapshot_date"],
        )
        items.append(
            {
                "filename": file.filename,
                "file_hash": file_hash,
                "snapshot_date": parsed["snapshot_date"],
                "available_cash": entry["available_cash"],
                "interests_received": entry["interests_received"],
                "invested": entry["invested"],
                "value": entry["value"],
                "gains": entry["gains"],
                "category": entry["category"],
            }
        )
    if not items:
        raise HTTPException(status_code=400, detail="Files required.")
    return {"status": "ok", "items": items}


@app.post("/portfolios/{portfolio_id}/imports/trade-republic/commit")
def trade_republic_commit(
    portfolio_id: int,
    payload: TradeRepublicImportRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    currency = portfolio["currency"]
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    created: list[dict] = []
    for item in payload.items:
        if not item.filename or not item.file_hash:
            raise HTTPException(status_code=400, detail="Invalid payload.")
        if _trade_republic_file_exists(portfolio_id, item.file_hash):
            raise HTTPException(status_code=409, detail="File already imported.")
        category = item.category.strip() if item.category else "Cash"
        if category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        entry = _build_trade_republic_entry(
            item.available_cash,
            item.interests_received,
            currency,
            category=category,
            source="file",
            source_file=item.filename,
            file_hash=item.file_hash,
            snapshot_date=item.snapshot_date,
        )
        meta = _save_trade_republic_entry(portfolio_id, entry)
        created.append(
            {
                "id": meta["id"],
                "created_at": meta["created_at"],
                "filename": item.filename,
                "available_cash": entry["available_cash"],
                "interests_received": entry["interests_received"],
                "invested": entry["invested"],
                "value": entry["value"],
                "gains": entry["gains"],
                "category": entry["category"],
                "snapshot_date": item.snapshot_date,
            }
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    return {"status": "created", "items": created}


@app.get("/portfolios/{portfolio_id}/imports/trade-republic")
def list_trade_republic_imports(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_trade_republic_entries(portfolio_id)}


@app.delete("/portfolios/{portfolio_id}/imports/trade-republic/{entry_id}")
def delete_trade_republic_import(
    portfolio_id: int,
    entry_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    with _db_connection() as conn:
        cursor = conn.execute(
            "DELETE FROM trade_republic_entries WHERE id = ? AND portfolio_id = ?",
            (entry_id, portfolio_id),
        )
        deleted = cursor.rowcount
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Import not found.")
    return {"status": "deleted"}


@app.post("/portfolios/{portfolio_id}/imports/xtb/preview")
def xtb_preview(
    portfolio_id: int,
    files: list[UploadFile] = File(...),
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not files:
        raise HTTPException(status_code=400, detail="Files required.")
    items: list[dict] = []
    warnings: list[dict] = []
    holdings: list[dict] = []
    operations: list[dict] = []
    for file in files:
        if not file.filename:
            continue
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Invalid file type.")
        account_type = _xtb_account_type(file.filename)
        file_bytes = file.file.read()
        parsed = _parse_xtb_file(file_bytes, file.filename)
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        if parsed["warnings"]:
            warnings.append({"filename": file.filename, "warnings": parsed["warnings"]})
        aggregated_positions = _aggregate_xtb_positions(parsed.get("positions", []))
        for position in aggregated_positions:
            holdings.append(
                {
                    "source_file": file.filename,
                    "ticker": position["ticker"],
                    "name": position.get("name"),
                    "shares": position["shares"],
                    "open_price": position["open_price"],
                    "purchase_value": position.get("purchase_value"),
                    "current_price": position.get("current_price"),
                    "category": "Stocks",
                }
            )
        items.append(
            {
                "filename": file.filename,
                "file_hash": file_hash,
                "account_type": account_type,
                "category": "Stocks",
                "current_value": parsed["current_value"],
                "cash_value": parsed["cash_value"],
                "invested": parsed["invested"],
                "profit_value": parsed["profit_value"],
                "profit_percent": parsed["profit_percent"],
            }
        )
        for operation in parsed.get("operations", []):
            operations.append(
                {
                    "source_file": file.filename,
                    "ticker": operation.get("ticker"),
                    "operation_type": operation.get("operation_type"),
                    "operation_kind": operation.get("operation_kind"),
                    "description": operation.get("description"),
                    "amount": operation.get("amount"),
                    "trade_date": operation.get("trade_date"),
                    "currency": "EUR",
                }
            )
    if not items:
        raise HTTPException(status_code=400, detail="Files required.")
    return {
        "status": "ok",
        "items": items,
        "warnings": warnings,
        "holdings": holdings,
        "operations": operations,
    }


@app.post("/portfolios/{portfolio_id}/imports/xtb/commit")
def xtb_commit(
    portfolio_id: int,
    payload: XtbImportCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    sanitized_items: list[XtbImportItem] = []
    for item in payload.items:
        category = item.category.strip() if item.category else "Stocks"
        if category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        profit_value = item.profit_value if item.profit_value is not None else 0.0
        profit_percent = item.profit_percent
        if profit_percent is None and item.current_value:
            profit_percent = (profit_value / item.current_value) * 100
        sanitized_items.append(
            XtbImportItem(
                filename=item.filename,
                file_hash=item.file_hash,
                account_type=item.account_type,
                category=category,
                current_value=item.current_value,
                cash_value=item.cash_value,
                invested=item.invested,
                profit_value=profit_value,
                profit_percent=profit_percent,
            )
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    try:
        meta = _save_xtb_imports(portfolio_id, sanitized_items)
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="File already imported.") from exc
    holdings_payload = payload.holdings or []
    operations_payload = payload.operations or []
    file_category = {item.filename: item.category for item in sanitized_items}
    file_hash_lookup = {item.filename: item.file_hash for item in sanitized_items}
    holdings_by_file: dict[str, list[HoldingImportItem]] = {}
    for holding in holdings_payload:
        source_file = holding.source_file
        if not source_file or source_file not in file_hash_lookup:
            continue
        category = holding.category or file_category.get(source_file) or "Stocks"
        holdings_by_file.setdefault(source_file, []).append(
            HoldingImportItem(
                source_file=source_file,
                ticker=holding.ticker,
                name=holding.name,
                shares=holding.shares,
                open_price=holding.open_price,
                current_price=holding.current_price,
                purchase_value=holding.purchase_value,
                category=category,
            )
        )
    operations_by_file: dict[str, list[HoldingOperationItem]] = {}
    for op in operations_payload:
        if not op.source_file:
            continue
        if op.source_file not in file_hash_lookup:
            continue
        operations_by_file.setdefault(op.source_file, []).append(op)
    for source_file in set(list(holdings_by_file.keys()) + list(operations_by_file.keys())):
        entries = holdings_by_file.get(source_file, [])
        ops = operations_by_file.get(source_file, [])
        if not entries and not ops:
            continue
        snapshot_date = _extract_date_from_filename(source_file)
        try:
            holding_import = _save_holdings_import(
                portfolio_id,
                "XTB",
                source_file,
                file_hash_lookup[source_file],
                snapshot_date,
            )
        except sqlite3.IntegrityError:
            continue
        if entries:
            _save_holdings_items(holding_import["id"], entries)
        if ops:
            _save_holdings_operations(
                holding_import["id"],
                portfolio_id,
                source_file,
                ops,
                currency=portfolio["currency"],
            )
    return {"status": "imported", "items": len(meta)}


@app.get("/portfolios/{portfolio_id}/imports/xtb")
def xtb_list(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_xtb_imports(portfolio_id)}


@app.delete("/portfolios/{portfolio_id}/imports/xtb/{import_id}")
def xtb_delete(
    portfolio_id: int,
    import_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    deleted = _delete_xtb_import(portfolio_id, import_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="XTB import not found.")
    return {"status": "deleted", "id": import_id}


@app.post("/portfolios/{portfolio_id}/imports/save-n-grow/preview")
def savengrow_preview(
    portfolio_id: int,
    file: UploadFile,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File required.")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file type.")
    file_bytes = file.file.read()
    parsed = _read_savengrow_cells(file_bytes, file.filename)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    return {
        "status": "ok",
        "filename": file.filename,
        "file_hash": file_hash,
        **parsed,
    }


@app.post("/portfolios/{portfolio_id}/imports/bancoinvest/preview")
def bancoinvest_preview(
    portfolio_id: int,
    file: UploadFile,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File required.")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file type.")
    file_bytes = file.file.read()
    parsed = _read_bancoinvest_cells(file_bytes, file.filename)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    category_map = _load_bancoinvest_category_map(portfolio_id)
    for item in parsed["items"]:
        key = _normalize_text(item["holder"]).replace(" ", "")
        mapped = category_map.get(key)
        if mapped:
            item["category"] = mapped
    return {
        "status": "ok",
        "filename": file.filename,
        "file_hash": file_hash,
        **parsed,
    }


@app.post("/portfolios/{portfolio_id}/imports/bancoinvest/commit")
def bancoinvest_commit(
    portfolio_id: int,
    payload: BancoInvestCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    sanitized_items: list[BancoInvestItem] = []
    for item in payload.items:
        category = item.category.strip() if item.category else "Retirement Plans"
        if category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        sanitized_items.append(
            BancoInvestItem(
                holder=item.holder,
                invested=item.invested,
                current_value=item.current_value,
                gains=item.gains,
                category=category,
            )
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    _upsert_bancoinvest_category_map(portfolio_id, sanitized_items)
    try:
        meta = _save_bancoinvest_import(portfolio_id, payload)
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="File already imported.") from exc
    _save_bancoinvest_items(meta["id"], sanitized_items)
    return {
        "status": "imported",
        "import_id": meta["id"],
        "imported_at": meta["imported_at"],
        "items": len(sanitized_items),
        "categories": categories,
    }


@app.get("/portfolios/{portfolio_id}/imports/bancoinvest")
def bancoinvest_list(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_bancoinvest_imports(portfolio_id)}


@app.delete("/portfolios/{portfolio_id}/imports/bancoinvest/{import_id}")
def bancoinvest_delete(
    portfolio_id: int,
    import_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    deleted = _delete_bancoinvest_import(portfolio_id, import_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="BancoInvest import not found.")
    return {"status": "deleted", "id": import_id}


@app.post("/portfolios/{portfolio_id}/imports/save-n-grow/commit")
def savengrow_commit(
    portfolio_id: int,
    payload: SaveNGrowCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.filename or not payload.file_hash:
        raise HTTPException(status_code=400, detail="Invalid payload.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    currency = portfolio["currency"]
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    sanitized_items: list[SaveNGrowItem] = []
    for item in payload.items:
        category = item.category.strip() if item.category else "Retirement Plans"
        if category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        profit_value = item.profit_value
        if profit_value is None and item.invested is not None:
            profit_value = item.current_value - item.invested
        profit_percent = item.profit_percent
        if (
            profit_percent is None
            and profit_value is not None
            and item.current_value
        ):
            profit_percent = (profit_value / item.current_value) * 100
        sanitized_items.append(
            SaveNGrowItem(
                name=item.name,
                invested=item.invested,
                current_value=item.current_value,
                profit_value=profit_value,
                profit_percent=profit_percent,
                category=category,
            )
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    invested_total = sum(item.invested or 0 for item in sanitized_items)
    current_value_total = sum(item.current_value for item in sanitized_items)
    profit_value_total = sum(item.profit_value or 0 for item in sanitized_items)
    profit_percent_total = (
        (profit_value_total / current_value_total) * 100 if current_value_total else None
    )
    try:
        meta = _save_savengrow_import(
            portfolio_id,
            payload,
            {
                "invested_total": invested_total,
                "current_value_total": current_value_total,
                "profit_value_total": profit_value_total,
                "profit_percent_total": profit_percent_total,
            },
            currency,
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="File already imported.") from exc
    _save_savengrow_items(meta["id"], sanitized_items)
    return {
        "status": "created",
        "entry": {
            "id": meta["id"],
            "created_at": meta["created_at"],
            "filename": payload.filename,
            "invested_total": invested_total,
            "current_value_total": current_value_total,
            "profit_value_total": profit_value_total,
            "profit_percent_total": profit_percent_total,
            "snapshot_date": payload.snapshot_date,
        },
    }


@app.get("/portfolios/{portfolio_id}/imports/save-n-grow")
def savengrow_list(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_savengrow_imports(portfolio_id)}


@app.post("/portfolios/{portfolio_id}/imports/aforronet/preview")
def aforronet_preview(
    portfolio_id: int,
    files: list[UploadFile] | None = File(default=None),
    file: UploadFile | None = File(default=None),
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    category = _latest_aforronet_category(portfolio_id) or "Emergency Funds"
    selected_files: list[UploadFile] = []
    if files:
        selected_files = files
    elif file:
        selected_files = [file]
    if not selected_files:
        raise HTTPException(status_code=400, detail="File required.")
    items: list[dict] = []
    for entry in selected_files:
        if not entry.filename:
            continue
        if not entry.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Invalid file type.")
        file_bytes = entry.file.read()
        parsed = _parse_aforronet_pdf(file_bytes, entry.filename)
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        items.append(
            {
                "filename": entry.filename,
                "file_hash": file_hash,
                "snapshot_date": parsed["snapshot_date"],
                "items": [
                    {
                        "name": "AforroNet",
                        "invested": parsed["invested"],
                        "current_value": parsed["current_value"],
                        "category": category,
                    }
                ],
            }
        )
    if not items:
        raise HTTPException(status_code=400, detail="File required.")
    return {"status": "ok", "items": items}


@app.post("/portfolios/{portfolio_id}/imports/aforronet/commit")
def aforronet_commit(
    portfolio_id: int,
    payload: AforroNetCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.filename or not payload.file_hash:
        raise HTTPException(status_code=400, detail="Invalid payload.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    currency = portfolio["currency"]
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    sanitized_items: list[AforroNetItem] = []
    for item in payload.items:
        category = item.category.strip() if item.category else "Emergency Funds"
        if category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        sanitized_items.append(
            AforroNetItem(
                name=item.name,
                invested=item.invested,
                current_value=item.current_value,
                category=category,
            )
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    invested_total = sum(item.invested for item in sanitized_items)
    current_value_total = sum(item.current_value for item in sanitized_items)
    try:
        meta = _save_aforronet_import(
            portfolio_id,
            payload,
            {
                "invested_total": invested_total,
                "current_value_total": current_value_total,
            },
            currency,
            sanitized_items[0].category if sanitized_items else "Emergency Funds",
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="File already imported.") from exc
    _save_aforronet_items(meta["id"], sanitized_items)
    return {
        "status": "created",
        "entry": {
            "id": meta["id"],
            "created_at": meta["created_at"],
            "filename": payload.filename,
            "invested_total": invested_total,
            "current_value_total": current_value_total,
            "snapshot_date": payload.snapshot_date,
        },
    }


@app.post("/portfolios/{portfolio_id}/imports/aforronet/commit-batch")
def aforronet_commit_batch(
    portfolio_id: int,
    payload: AforroNetBatchCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.imports:
        raise HTTPException(status_code=400, detail="No items to import.")
    currency = portfolio["currency"]
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    entries: list[dict] = []
    for import_payload in payload.imports:
        if not import_payload.filename or not import_payload.file_hash:
            raise HTTPException(status_code=400, detail="Invalid payload.")
        if not import_payload.items:
            raise HTTPException(status_code=400, detail="No items to import.")
        sanitized_items: list[AforroNetItem] = []
        for item in import_payload.items:
            category = item.category.strip() if item.category else "Emergency Funds"
            if category.lower() not in allowed:
                allowed.add(category.lower())
                categories.append(category)
                new_categories.append(category)
            sanitized_items.append(
                AforroNetItem(
                    name=item.name,
                    invested=item.invested,
                    current_value=item.current_value,
                    category=category,
                )
            )
        invested_total = sum(item.invested for item in sanitized_items)
        current_value_total = sum(item.current_value for item in sanitized_items)
        try:
            meta = _save_aforronet_import(
                portfolio_id,
                import_payload,
                {
                    "invested_total": invested_total,
                    "current_value_total": current_value_total,
                },
                currency,
                sanitized_items[0].category if sanitized_items else "Emergency Funds",
            )
        except sqlite3.IntegrityError as exc:
            raise HTTPException(
                status_code=409,
                detail=f"File already imported: {import_payload.filename}",
            ) from exc
        _save_aforronet_items(meta["id"], sanitized_items)
        entries.append(
            {
                "id": meta["id"],
                "created_at": meta["created_at"],
                "filename": import_payload.filename,
                "invested_total": invested_total,
                "current_value_total": current_value_total,
                "snapshot_date": import_payload.snapshot_date,
            }
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    return {"status": "created", "entries": entries}


@app.get("/portfolios/{portfolio_id}/imports/aforronet")
def aforronet_list(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    return {"items": _list_aforronet_imports(portfolio_id)}


@app.delete("/portfolios/{portfolio_id}/imports/aforronet/{import_id}")
def aforronet_delete(
    portfolio_id: int,
    import_id: int,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    deleted = _delete_aforronet_import(portfolio_id, import_id)
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Import not found.")
    return {"status": "deleted"}


@app.get("/portfolios/{portfolio_id}/categories/settings")
def list_category_settings(
    portfolio_id: int, authorization: str | None = Header(default=None)
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    _ensure_category_settings(portfolio_id, categories)
    settings = _get_category_settings(portfolio_id)
    return {
        "items": [
            {
                "category": category,
                "is_investment": settings.get(
                    category, _default_is_investment(category)
                ),
            }
            for category in categories
        ]
    }


@app.post("/portfolios/{portfolio_id}/categories/settings")
def update_category_settings(
    portfolio_id: int,
    payload: CategorySettingRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    category = payload.category.strip()
    if not category:
        raise HTTPException(status_code=400, detail="Category required.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    if category.lower() not in {item.lower() for item in categories}:
        raise HTTPException(status_code=404, detail="Category not found.")
    _set_category_setting(portfolio_id, category, payload.is_investment)
    return {
        "status": "updated",
        "category": category,
        "is_investment": payload.is_investment,
    }


@app.post("/portfolios/{portfolio_id}/categories/add")
def add_portfolio_category(
    portfolio_id: int,
    payload: CategoryAddRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    category = payload.category.strip()
    if not category:
        raise HTTPException(status_code=400, detail="Category required.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    if category.lower() in {item.lower() for item in categories}:
        return {"status": "exists", "categories": categories}
    categories.append(category)
    _update_portfolio_categories(portfolio_id, categories)
    return {"status": "added", "categories": categories}


@app.post("/portfolios/{portfolio_id}/categories/remove")
def remove_portfolio_category(
    portfolio_id: int,
    payload: CategoryRemoveRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    category = payload.category.strip()
    if not category:
        raise HTTPException(status_code=400, detail="Category required.")
    categories = _filter_categories(json.loads(portfolio["categories_json"]))
    if category.lower() in {item.lower() for item in DEFAULT_CATEGORIES}:
        raise HTTPException(status_code=400, detail="Default category cannot be removed.")
    if category.lower() not in {item.lower() for item in categories}:
        raise HTTPException(status_code=404, detail="Category not found.")
    count = _count_santander_category(portfolio_id, category)
    if count > 0 and not payload.clear_data:
        raise HTTPException(
            status_code=409,
            detail="Category has data. Clear data to remove.",
        )
    if count > 0 and payload.clear_data:
        _delete_santander_category(portfolio_id, category)
    _delete_category_map_by_category(portfolio_id, category)
    _delete_category_setting(portfolio_id, category)
    categories = [item for item in categories if item.lower() != category.lower()]
    _update_portfolio_categories(portfolio_id, categories)
    return {"status": "removed", "remaining": categories, "cleared": count}


@app.post("/portfolios/{portfolio_id}/imports/santander/preview")
def santander_preview(
    portfolio_id: int,
    file: UploadFile,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    if not _get_portfolio(portfolio_id, session["email"]):
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File required.")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file type.")
    file_bytes = file.file.read()
    items, warnings = _parse_santander_sheet(file_bytes, file.filename)
    if not items:
        raise HTTPException(
            status_code=400,
            detail=(
                "This file does not match the Santander format. "
                "Please choose a Santander Excel statement."
            ),
        )
    category_map = _load_category_map(portfolio_id)
    for item in items:
        key = _account_key(item.section, item.account)
        mapped = category_map.get(key)
        if mapped:
            item.category = str(mapped["category"])
            item.ignore = bool(mapped["ignore"])
    return {
        "status": "ok",
        "filename": file.filename,
        "items": [item.model_dump() for item in items],
        "warnings": warnings,
    }


@app.post("/portfolios/{portfolio_id}/imports/santander/commit")
def santander_commit(
    portfolio_id: int,
    payload: SantanderCommitRequest,
    authorization: str | None = Header(default=None),
) -> dict:
    session = _require_session(authorization)
    portfolio = _get_portfolio(portfolio_id, session["email"])
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found.")
    if not payload.items:
        raise HTTPException(status_code=400, detail="No items to import.")
    categories = json.loads(portfolio["categories_json"])
    allowed = {item.lower() for item in categories}
    new_categories: list[str] = []
    sanitized: list[SantanderItem] = []
    mapped_items: list[SantanderItem] = []
    for item in payload.items:
        category = item.category.strip() if item.category else "Unknown"
        if not category:
            category = "Unknown"
        ignore = bool(item.ignore and category.lower() == "unknown")
        if category.lower() != "unknown" and category.lower() not in allowed:
            allowed.add(category.lower())
            categories.append(category)
            new_categories.append(category)
        normalized_item = SantanderItem(
            section=item.section,
            account=item.account,
            description=item.description,
            balance=item.balance,
            category=category,
            ignore=ignore,
            invested=item.invested,
            gains=item.gains,
        )
        mapped_items.append(normalized_item)
        if ignore and category.lower() == "unknown":
            continue
        sanitized.append(
            SantanderItem(
                section=item.section,
                account=item.account,
                description=item.description,
                balance=item.balance,
                category=category,
                invested=item.invested,
                gains=item.gains,
            )
        )
    if new_categories:
        _update_portfolio_categories(portfolio_id, categories)
    if mapped_items:
        _upsert_category_map(portfolio_id, mapped_items)
    meta = _save_santander_import(portfolio_id, payload.filename, sanitized)
    return {
        "status": "imported",
        "import_id": meta["import_id"],
        "imported_at": meta["imported_at"],
        "items": len(sanitized),
        "categories": categories,
    }


@app.get("/admin/prices")
def admin_list_prices(
    authorization: str | None = Header(default=None),
    search: str | None = None,
    limit: int = 100
) -> dict:
    """Lista todos os preços de tickers armazenados (apenas admin)."""
    _require_admin(authorization)
    with _db_connection() as conn:
        if search:
            rows = conn.execute(
                """
                SELECT ticker, price, currency, updated_at
                FROM holdings_prices
                WHERE ticker LIKE ?
                ORDER BY ticker
                LIMIT ?
                """,
                (f"%{search.upper()}%", limit),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT ticker, price, currency, updated_at
                FROM holdings_prices
                ORDER BY updated_at DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
    items = [
        {
            "ticker": row["ticker"],
            "price": float(row["price"]),
            "currency": row["currency"],
            "updated_at": row["updated_at"]
        }
        for row in rows
    ]
    return {"items": items, "total": len(items)}


@app.post("/admin/prices")
def admin_upload_price(
    payload: TickerPriceUpload,
    authorization: str | None = Header(default=None)
) -> dict:
    """Carrega/atualiza o preço de um ticker (apenas admin)."""
    _require_admin(authorization)
    ticker = payload.ticker.strip().upper()
    if not ticker:
        raise HTTPException(status_code=400, detail="Ticker is required.")
    if payload.price <= 0:
        raise HTTPException(status_code=400, detail="Price must be greater than 0.")
    
    now = datetime.utcnow().isoformat()
    with _db_connection() as conn:
        conn.execute(
            """
            INSERT INTO holdings_prices (ticker, price, currency, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(ticker) DO UPDATE SET
                price = excluded.price,
                currency = excluded.currency,
                updated_at = excluded.updated_at
            """,
            (ticker, payload.price, payload.currency, now),
        )
    return {
        "status": "success",
        "ticker": ticker,
        "price": payload.price,
        "currency": payload.currency,
        "updated_at": now
    }


@app.post("/admin/prices/bulk")
def admin_bulk_upload_prices(
    payload: BulkTickerPriceUpload,
    authorization: str | None = Header(default=None)
) -> dict:
    """Carrega/atualiza múltiplos preços de tickers (apenas admin)."""
    _require_admin(authorization)
    if not payload.prices:
        raise HTTPException(status_code=400, detail="No prices provided.")
    
    now = datetime.utcnow().isoformat()
    success_count = 0
    errors = []
    
    with _db_connection() as conn:
        for item in payload.prices:
            ticker = item.ticker.strip().upper()
            if not ticker:
                errors.append({"ticker": item.ticker, "error": "Ticker is required"})
                continue
            if item.price <= 0:
                errors.append({"ticker": ticker, "error": "Price must be greater than 0"})
                continue
            
            try:
                conn.execute(
                    """
                    INSERT INTO holdings_prices (ticker, price, currency, updated_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(ticker) DO UPDATE SET
                        price = excluded.price,
                        currency = excluded.currency,
                        updated_at = excluded.updated_at
                    """,
                    (ticker, item.price, item.currency, now),
                )
                success_count += 1
            except Exception as e:
                errors.append({"ticker": ticker, "error": str(e)})
    
    return {
        "status": "completed",
        "success": success_count,
        "errors": errors,
        "updated_at": now
    }


@app.delete("/admin/prices/{ticker}")
def admin_delete_price(
    ticker: str,
    authorization: str | None = Header(default=None)
) -> dict:
    """Remove o preço de um ticker (apenas admin)."""
    _require_admin(authorization)
    ticker = ticker.strip().upper()
    with _db_connection() as conn:
        conn.execute("DELETE FROM holdings_prices WHERE ticker = ?", (ticker,))
    return {"status": "deleted", "ticker": ticker}


@app.post("/admin/tickers/upload")
async def admin_upload_tickers_excel(
    file: UploadFile,
    authorization: str | None = Header(default=None)
) -> dict:
    """Upload de ficheiro Excel com lista de tickers (apenas admin)."""
    _require_admin(authorization)
    
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        import openpyxl
        content = await file.read()
        
        # Criar ficheiro temporário
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        # Ler Excel
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        
        tickers = []
        seen_header = False
        
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            cell_value = str(row[0]).strip().upper()
            
            # Skip common header names
            if cell_value in ['TICKER', 'SYMBOL', 'CODE', 'STOCK', 'NAME']:
                seen_header = True
                continue
            
            # Skip if it looks like a header and we haven't seen data yet
            if not seen_header and not tickers and len(cell_value) > 10:
                seen_header = True
                continue
            
            # Add valid ticker
            if cell_value and len(cell_value) <= 20:  # Reasonable ticker length
                tickers.append(cell_value)
        
        # Limpar ficheiro temporário
        import os
        os.unlink(tmp_path)
        
        if not tickers:
            raise HTTPException(status_code=400, detail="No tickers found in file. Make sure tickers are in the first column.")
        
        return {
            "status": "uploaded",
            "tickers": tickers,
            "count": len(tickers)
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/admin/tickers/update-fixed-from-excel")
async def admin_update_fixed_ticker_data(
    file: UploadFile,
    authorization: str | None = Header(default=None)
) -> dict:
    """Atualiza apenas dados fixos dos tickers (Name, Class, Sector, Country, Currency) via Excel."""
    _require_admin(authorization)
    
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        import openpyxl
        import tempfile
        import os
        
        content = await file.read()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        
        # Expected columns: Ticker, Name, Class, Sector, Country, Currency
        headers = []
        data_rows = []
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:  # Header row
                headers = [str(cell).strip().lower() if cell else "" for cell in row]
                continue
            
            if not row or not row[0]:  # Skip empty rows
                continue
            
            data_rows.append(row)
        
        os.unlink(tmp_path)
        
        # Validate headers
        required = ['ticker', 'name', 'class', 'sector', 'country', 'currency']
        if not all(req in headers for req in required):
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns. Expected: {', '.join(required)}"
            )
        
        # Get column indices
        ticker_idx = headers.index('ticker')
        name_idx = headers.index('name')
        class_idx = headers.index('class')
        sector_idx = headers.index('sector')
        country_idx = headers.index('country')
        currency_idx = headers.index('currency')
        
        success_count = 0
        errors = []
        
        now = datetime.utcnow().isoformat()
        
        with _db_connection() as conn:
            for row in data_rows:
                try:
                    ticker = str(row[ticker_idx]).strip().upper() if row[ticker_idx] else None
                    if not ticker:
                        continue
                    
                    name = str(row[name_idx]).strip() if row[name_idx] else ticker
                    asset_class = str(row[class_idx]).strip() if row[class_idx] else "Stock"
                    sector = str(row[sector_idx]).strip() if row[sector_idx] else None
                    country = str(row[country_idx]).strip() if row[country_idx] else None
                    currency = str(row[currency_idx]).strip().upper() if row[currency_idx] else "USD"
                    
                    # Update only fixed fields, preserve API-fetched data
                    conn.execute(
                        """
                        INSERT INTO ticker_metadata (
                            ticker, name, asset_class, sector, country, currency, 
                            industry, region, exchange, dividend_yield, dividend_frequency, 
                            next_dividend_date, next_dividend_amount, last_updated
                        )
                        VALUES (?, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, ?)
                        ON CONFLICT(ticker) DO UPDATE SET
                            name = excluded.name,
                            asset_class = excluded.asset_class,
                            sector = excluded.sector,
                            country = excluded.country,
                            currency = excluded.currency,
                            last_updated = excluded.last_updated
                        """,
                        (ticker, name, asset_class, sector, country, currency, now),
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append({"ticker": ticker if 'ticker' in locals() else "unknown", "error": str(e)})
        
        return {
            "status": "completed",
            "success": success_count,
            "errors": len(errors),
            "error_details": errors[:10]  # Limit error details
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/admin/tickers/update-all-from-excel")
async def admin_update_all_from_excel(
    file: UploadFile,
    authorization: str | None = Header(default=None)
) -> dict:
    """Atualiza todos os dados dos tickers via Excel (todos os campos)."""
    _require_admin(authorization)
    
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        import openpyxl
        import tempfile
        import os
        
        content = await file.read()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        
        headers = []
        data_rows = []
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:  # Header row
                headers = [str(cell).strip().lower() if cell else "" for cell in row]
                continue
            
            if not row or not row[0]:  # Skip empty rows
                continue
            
            data_rows.append(row)
        
        os.unlink(tmp_path)
        
        # Expected columns (all fields)
        required = ['ticker']
        if 'ticker' not in headers:
            raise HTTPException(status_code=400, detail="Missing 'ticker' column")
        
        # Get all column indices
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        success_count = 0
        errors = []
        
        now = datetime.utcnow().isoformat()
        
        with _db_connection() as conn:
            for row in data_rows:
                try:
                    ticker = str(row[col_map['ticker']]).strip().upper() if col_map.get('ticker') is not None and row[col_map['ticker']] else None
                    if not ticker:
                        continue
                    
                    # Extract all available fields
                    name = str(row[col_map.get('name', -1)]).strip() if col_map.get('name') is not None and len(row) > col_map.get('name', -1) and row[col_map.get('name', -1)] else ticker
                    asset_class = str(row[col_map.get('class', -1)]).strip() if col_map.get('class') is not None and len(row) > col_map.get('class', -1) and row[col_map.get('class', -1)] else "Stock"
                    sector = str(row[col_map.get('sector', -1)]).strip() if col_map.get('sector') is not None and len(row) > col_map.get('sector', -1) and row[col_map.get('sector', -1)] else None
                    industry = str(row[col_map.get('industry', -1)]).strip() if col_map.get('industry') is not None and len(row) > col_map.get('industry', -1) and row[col_map.get('industry', -1)] else None
                    country = str(row[col_map.get('country', -1)]).strip() if col_map.get('country') is not None and len(row) > col_map.get('country', -1) and row[col_map.get('country', -1)] else None
                    region = str(row[col_map.get('region', -1)]).strip() if col_map.get('region') is not None and len(row) > col_map.get('region', -1) and row[col_map.get('region', -1)] else None
                    currency = str(row[col_map.get('currency', -1)]).strip().upper() if col_map.get('currency') is not None and len(row) > col_map.get('currency', -1) and row[col_map.get('currency', -1)] else "USD"
                    exchange = str(row[col_map.get('exchange', -1)]).strip() if col_map.get('exchange') is not None and len(row) > col_map.get('exchange', -1) and row[col_map.get('exchange', -1)] else None
                    
                    # Dividend fields (optional)
                    dividend_yield = None
                    if col_map.get('dividend_yield') is not None and len(row) > col_map.get('dividend_yield', -1) and row[col_map.get('dividend_yield', -1)]:
                        try:
                            dividend_yield = float(row[col_map['dividend_yield']])
                        except:
                            pass
                    
                    dividend_frequency = str(row[col_map.get('dividend_frequency', -1)]).strip() if col_map.get('dividend_frequency') is not None and len(row) > col_map.get('dividend_frequency', -1) and row[col_map.get('dividend_frequency', -1)] else None
                    next_dividend_date = str(row[col_map.get('next_dividend_date', -1)]).strip() if col_map.get('next_dividend_date') is not None and len(row) > col_map.get('next_dividend_date', -1) and row[col_map.get('next_dividend_date', -1)] else None
                    
                    next_dividend_amount = None
                    if col_map.get('next_dividend_amount') is not None and len(row) > col_map.get('next_dividend_amount', -1) and row[col_map.get('next_dividend_amount', -1)]:
                        try:
                            next_dividend_amount = float(row[col_map['next_dividend_amount']])
                        except:
                            pass
                    
                    # Insert or update all fields
                    conn.execute(
                        """
                        INSERT INTO ticker_metadata (
                            ticker, name, asset_class, sector, industry, country, region, 
                            currency, exchange, dividend_yield, dividend_frequency, 
                            next_dividend_date, next_dividend_amount, last_updated
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(ticker) DO UPDATE SET
                            name = excluded.name,
                            asset_class = excluded.asset_class,
                            sector = excluded.sector,
                            industry = excluded.industry,
                            country = excluded.country,
                            region = excluded.region,
                            currency = excluded.currency,
                            exchange = excluded.exchange,
                            dividend_yield = excluded.dividend_yield,
                            dividend_frequency = excluded.dividend_frequency,
                            next_dividend_date = excluded.next_dividend_date,
                            next_dividend_amount = excluded.next_dividend_amount,
                            last_updated = excluded.last_updated
                        """,
                        (ticker, name, asset_class, sector, industry, country, region, 
                         currency, exchange, dividend_yield, dividend_frequency, 
                         next_dividend_date, next_dividend_amount, now),
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append({"ticker": ticker if 'ticker' in locals() else "unknown", "error": str(e)})
        
        return {
            "status": "completed",
            "success": success_count,
            "errors": len(errors),
            "error_details": errors[:10]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/admin/tickers/fetch-metadata")
def admin_fetch_ticker_metadata(
    ticker: str,
    authorization: str | None = Header(default=None)
) -> dict:
    """Busca e armazena metadados completos de um ticker (apenas admin).
    
    Usa Finnhub se configurado para obter dados abrangentes:
    - Company profile (name, country, exchange, industry)
    - Current price
    - Dividend yield
    - Next dividend payment date and amount
    """
    _require_admin(authorization)
    ticker = ticker.strip().upper()
    
    # Buscar metadados completos usando Finnhub se disponível
    if PRICE_API_PROVIDER in ("finnhub", "finhub"):
        metadata = _fetch_metadata_finnhub(ticker)
        
        if not metadata or not metadata.get("ticker"):
            raise HTTPException(status_code=404, detail=f"Could not fetch metadata for {ticker}")
        
        # Salvar metadados na base de dados
        _save_ticker_metadata(metadata)
        
        # Salvar preço se disponível
        price_data = None
        if metadata.get("price"):
            now = datetime.utcnow().isoformat()
            with _db_connection() as conn:
                conn.execute(
                    """
                    INSERT INTO holdings_prices (ticker, price, currency, updated_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(ticker) DO UPDATE SET
                        price = excluded.price,
                        currency = excluded.currency,
                        updated_at = excluded.updated_at
                    """,
                    (ticker, metadata["price"], metadata.get("currency", "USD"), now),
                )
            price_data = {
                "ticker": ticker,
                "price": metadata["price"],
                "currency": metadata.get("currency", "USD")
            }
        
        return {"status": "success", "metadata": metadata, "price": price_data}
    
    else:
        # Fallback para yfinance (legado)
        metadata = _fetch_ticker_metadata_yfinance(ticker)
        if not metadata:
            raise HTTPException(status_code=404, detail=f"Could not fetch metadata for {ticker}")
        
        # Salvar na base de dados
        _save_ticker_metadata(metadata)
        
        # Buscar e salvar preço também
        price_data = _fetch_ticker_price_yfinance(ticker)
        if price_data:
            now = datetime.utcnow().isoformat()
            with _db_connection() as conn:
                conn.execute(
                    """
                    INSERT INTO holdings_prices (ticker, price, currency, updated_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(ticker) DO UPDATE SET
                        price = excluded.price,
                        currency = excluded.currency,
                        updated_at = excluded.updated_at
                    """,
                    (price_data["ticker"], price_data["price"], price_data["currency"], now),
                )
        
        return {"status": "success", "metadata": metadata, "price": price_data}


@app.post("/admin/tickers/fetch-bulk")
async def admin_fetch_bulk_metadata(
    tickers: list[str],
    authorization: str | None = Header(default=None)
) -> dict:
    """Busca metadados completos e preços para múltiplos tickers (apenas admin).
    
    Usa Finnhub API se configurado, com rate limit de 59 calls/min.
    Cada ticker requer 4 chamadas: profile2, quote, metric, dividend (total ~4.2s por ticker).
    """
    _require_admin(authorization)
    
    import time
    
    success = []
    errors = []
    total = len(tickers)
    
    # Estimate time: 4 API calls per ticker with 1.05s between = ~4.2s per ticker
    estimated_time = total * 4.2 if PRICE_API_PROVIDER in ("finnhub", "finhub") else total * 3.0
    
    print(f"\n{'='*60}")
    print(f"Starting comprehensive metadata fetch for {total} tickers...")
    print(f"Provider: {PRICE_API_PROVIDER}")
    print(f"Estimated time: {estimated_time:.0f} seconds (~{estimated_time / 60:.1f} minutes)")
    print(f"{'='*60}\n")
    
    for idx, ticker_raw in enumerate(tickers, 1):
        ticker = ticker_raw.strip().upper()
        print(f"\n[{idx}/{total}] Processing: {ticker}")
        
        try:
            # Fetch comprehensive metadata using Finnhub (includes rate limiting)
            if PRICE_API_PROVIDER in ("finnhub", "finhub"):
                metadata = _fetch_metadata_finnhub(ticker)
                
                # Save metadata to database
                _save_ticker_metadata(metadata)
                
                # Save price if available
                if metadata.get("price"):
                    _upsert_price(ticker, metadata["price"])
                    print(f"✓ {ticker}: ${metadata['price']:.2f} | {metadata.get('name', 'N/A')} | {metadata.get('country', 'N/A')} | Div: {metadata.get('dividend_yield', 0):.2%}")
                else:
                    print(f"⚠ {ticker}: Metadata saved but no price available")
                
                success.append(ticker)
                
            else:
                # Fallback to basic price fetch for other providers
                price_value = _fetch_latest_price(ticker)
                
                # Save basic metadata
                metadata = {
                    "ticker": ticker,
                    "name": ticker,
                    "asset_class": "Stock",
                    "sector": None,
                    "industry": None,
                    "country": None,
                    "region": None,
                    "currency": "USD",
                    "exchange": None,
                    "dividend_yield": None,
                    "dividend_frequency": None,
                    "next_dividend_date": None,
                    "next_dividend_amount": None
                }
                _save_ticker_metadata(metadata)
                _upsert_price(ticker, price_value)
                
                print(f"✓ {ticker}: Price ${price_value:.2f}")
                success.append(ticker)
                
                # Rate limiting for non-Finnhub providers
                if idx < total:
                    time.sleep(3.0)
                
        except Exception as e:
            error_msg = str(e)[:150]
            errors.append({"ticker": ticker, "error": error_msg})
            print(f"✗ {ticker}: {error_msg}")
    
    print(f"\n{'='*60}")
    print(f"Bulk fetch completed!")
    print(f"✓ Success: {len(success)}/{total} ({len(success)/total*100:.1f}%)" if total else "No tickers")
    print(f"✗ Errors: {len(errors)}/{total}")
    if errors:
        print(f"Failed tickers: {', '.join([e['ticker'] for e in errors[:10]])}")
        if len(errors) > 10:
            print(f"... and {len(errors) - 10} more")
    print(f"{'='*60}\n")
    
    return {
        "status": "completed",
        "success": len(success),
        "errors": len(errors),
        "success_tickers": success,
        "error_details": errors
    }


@app.get("/admin/tickers/metadata")
def admin_list_metadata(
    authorization: str | None = Header(default=None),
    search: str | None = None,
    limit: int = 100
) -> dict:
    """Lista metadados de tickers armazenados com preço atual (apenas admin)."""
    _require_admin(authorization)
    
    with _db_connection() as conn:
        if search:
            rows = conn.execute(
                """
                SELECT tm.*, hp.price as current_price
                FROM ticker_metadata tm
                LEFT JOIN holdings_prices hp ON tm.ticker = hp.ticker
                WHERE tm.ticker LIKE ? OR tm.name LIKE ?
                ORDER BY tm.ticker
                LIMIT ?
                """,
                (f"%{search.upper()}%", f"%{search}%", limit),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT tm.*, hp.price as current_price
                FROM ticker_metadata tm
                LEFT JOIN holdings_prices hp ON tm.ticker = hp.ticker
                ORDER BY tm.last_updated DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
    
    items = [dict(row) for row in rows]
    return {"items": items, "total": len(items)}


@app.delete("/admin/tickers/{ticker}")
def admin_delete_ticker(
    ticker: str,
    authorization: str | None = Header(default=None)
) -> dict:
    """Remove ticker e seus metadados (apenas admin)."""
    _require_admin(authorization)
    ticker = ticker.strip().upper()
    
    with _db_connection() as conn:
        conn.execute("DELETE FROM holdings_prices WHERE ticker = ?", (ticker,))
        conn.execute("DELETE FROM ticker_metadata WHERE ticker = ?", (ticker,))
    
    return {"status": "deleted", "ticker": ticker}


@app.get("/admin/api-settings")
def admin_get_api_settings(
    authorization: str | None = Header(default=None)
) -> dict:
    """Lista configurações de API providers (apenas admin)."""
    _require_admin(authorization)
    
    providers = [
        {
            "id": "yfinance",
            "name": "Yahoo Finance (yfinance)",
            "description": "Free API, no key required. Good for US & European stocks.",
            "enabled": True,  # Always enabled as fallback
            "requires_key": False,
            "has_key": True,
            "limits": "Rate limited, ~48 requests/hour",
            "supported_markets": ["US", "EU", "UK", "Asia"]
        },
        {
            "id": "twelvedata",
            "name": "Twelve Data",
            "description": "Free tier: 800 requests/day. Supports global markets.",
            "enabled": PRICE_API_PROVIDER == "twelvedata",
            "requires_key": True,
            "has_key": bool(PRICE_API_KEY),
            "limits": "Free: 800/day, Paid: unlimited",
            "supported_markets": ["US", "EU", "UK", "Asia", "Crypto"]
        },
        {
            "id": "finnhub",
            "name": "Finnhub",
            "description": "Free tier: 60 calls/minute. Focus on US stocks.",
            "enabled": PRICE_API_PROVIDER == "finnhub",
            "requires_key": True,
            "has_key": bool(os.getenv("FINNHUB_API_KEY")),
            "limits": "Free: 60/min, Paid: higher limits",
            "supported_markets": ["US", "EU"]
        }
    ]
    
    return {
        "providers": providers,
        "current_provider": PRICE_API_PROVIDER,
        "fallback": "yfinance"
    }


@app.post("/admin/api-settings/test")
def admin_test_api_provider(
    payload: dict,
    authorization: str | None = Header(default=None)
) -> dict:
    """Testa conectividade com um API provider (apenas admin)."""
    _require_admin(authorization)
    provider = payload.get("provider")
    
    test_ticker = "AAPL"
    
    try:
        if provider == "yfinance":
            price_data = _fetch_ticker_price_yfinance(test_ticker)
            if price_data and price_data.get("price"):
                return {
                    "status": "success",
                    "provider": provider,
                    "test_ticker": test_ticker,
                    "price": price_data["price"],
                    "message": "Connection successful"
                }
        elif provider == "twelvedata":
            if not PRICE_API_KEY:
                return {
                    "status": "error",
                    "provider": provider,
                    "message": "API key not configured. Set PRICE_API_KEY in .env file."
                }
            price = _fetch_price_twelvedata(test_ticker)
            return {
                "status": "success",
                "provider": provider,
                "test_ticker": test_ticker,
                "price": price,
                "message": "Connection successful"
            }
        elif provider == "finnhub":
            finnhub_key = os.getenv("FINNHUB_API_KEY")
            if not finnhub_key:
                return {
                    "status": "error",
                    "provider": provider,
                    "message": "API key not configured. Set FINNHUB_API_KEY in .env file."
                }
            price = _fetch_price_finnhub(test_ticker)
            return {
                "status": "success",
                "provider": provider,
                "test_ticker": test_ticker,
                "price": price,
                "message": "Connection successful"
            }
        else:
            raise HTTPException(status_code=400, detail=f"Unknown provider: {provider}")
    except Exception as e:
        return {
            "status": "error",
            "provider": provider,
            "message": str(e)
        }


@app.post("/auth/google")
def google_oauth() -> dict:
    raise HTTPException(status_code=501, detail="Google OAuth not implemented.")
